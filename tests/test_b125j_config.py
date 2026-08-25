import json
from pathlib import Path


TEMPLATE_ROOT = Path(__file__).parents[1] / "templates" / "BACKPACK"
CONFIG_PATH = TEMPLATE_ROOT / "B125J" / "config.json"


def test_b125j_has_required_backpack_dimensions_and_valid_capacity_unit():
    profile = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    expected_dimensions = {
        "depth_width_side_to_side": "28",
        "depth_width_side_to_side_unit_of_measure": "Centimetres",
        "depth_front_to_back": "19",
        "depth_front_to_back_unit_of_measure": "Centimetres",
        "depth_height_floor_to_top": "38",
        "depth_height_floor_to_top_unit_of_measure": "Centimetres",
    }

    for field_group in ("extra_parent_fields", "extra_child_fields"):
        fields = profile[field_group]
        assert {key: fields.get(key) for key in expected_dimensions} == expected_dimensions
        assert fields["capacity_unit_of_measure"] == "L"


def test_b125j_uses_a_backpack_workbook_with_depth_headers():
    from openpyxl import load_workbook

    profile = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    workbook_path = TEMPLATE_ROOT / profile["template_file"]
    workbook = load_workbook(workbook_path, read_only=True, keep_vba=True)
    try:
        template = workbook["Template"]
        metadata = {str(cell.value) for cell in template[1] if cell.value}
        headers = {str(cell.value) for cell in template[3] if cell.value}
        populated_template_cells = [
            cell.value
            for row in template.iter_rows(min_row=4)
            for cell in row
            if cell.value not in (None, "")
        ]
    finally:
        workbook.close()

    assert "TemplateSignature=QkFDS1BBQ0s=" in metadata
    assert {
        "depth_width_side_to_side",
        "depth_width_side_to_side_unit_of_measure",
        "depth_front_to_back",
        "depth_front_to_back_unit_of_measure",
        "depth_height_floor_to_top",
        "depth_height_floor_to_top_unit_of_measure",
    } <= headers
    assert populated_template_cells == []
