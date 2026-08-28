from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
HOODIE_ROOT = ROOT / "templates" / "HOODIE"


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def test_new_hoodie_profiles_satisfy_family_schema_and_use_shared_workbook() -> None:
    schema = load_json(HOODIE_ROOT / "schema.json")
    workbook_path = HOODIE_ROOT / schema["workbook_file"]
    profiles = [
        load_json(HOODIE_ROOT / "Generic Hoodies" / "config.json"),
        load_json(HOODIE_ROOT / "Generic Sweatshirts" / "config.json"),
        load_json(HOODIE_ROOT / "UC202" / "config.json"),
        load_json(HOODIE_ROOT / "UC503" / "config.json"),
    ]

    assert workbook_path.is_file()
    for profile in profiles:
        missing = [field for field in schema["required_profile_fields"] if not profile.get(field)]
        assert missing == []

    workbook = load_workbook(workbook_path, read_only=True, keep_vba=True)
    try:
        assert "Template" in workbook.sheetnames
        headers = {
            str(workbook["Template"].cell(3, column).value or "")
            for column in range(1, workbook["Template"].max_column + 1)
        }
    finally:
        workbook.close()

    assert {
        "item_sku",
        "item_name",
        "parent_child",
        "relationship_type",
        "variation_theme",
        "main_image_url",
    }.issubset(headers)


def test_h01_h02_folder_codes_select_generic_hoodie_detection_path() -> None:
    app_source = (ROOT / "app.py").read_text(encoding="utf-8")

    assert 'bounded_match("H01") or bounded_match("H02")' in app_source
    assert '== "GENERIC_HOODIES"' in app_source


def test_s01_s02_folder_codes_select_generic_sweatshirt_detection_path() -> None:
    app_source = (ROOT / "app.py").read_text(encoding="utf-8")

    assert 'bounded_match("S01") or bounded_match("S02")' in app_source
    assert '== "GENERIC_SWEATSHIRTS"' in app_source


def test_uc202_profile_matches_supplied_childrens_sweatshirt_contract() -> None:
    profile = load_json(HOODIE_ROOT / "UC202" / "config.json")

    assert profile["template_key"] == "UC202"
    assert profile["sizes"] == [
        "2 YRS", "3/4 YRS", "5/6 YRS", "7/8 YRS", "9/10 YRS", "11/13 YRS"
    ]
    assert profile["colors"] == [
        "Black",
        "Bottle Green",
        "Brown",
        "Heather Grey",
        "Maroon",
        "Navy",
        "Purple",
        "Red",
        "Royal",
        "Sky",
        "Yellow",
    ]
    assert len(profile["colors"]) * len(profile["sizes"]) == 66
