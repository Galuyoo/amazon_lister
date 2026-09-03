from __future__ import annotations
from datetime import datetime
import hashlib
import inspect
import io
import os
import time
import json
import re
import random
import string
import traceback
import zipfile

from copy import copy, deepcopy
from collections import Counter
from pathlib import Path
from typing import Any, Callable
from utils.image_resolver import resolve_one
import streamlit as st
from openpyxl import load_workbook
from itertools import product
from services.listing_memory import (
    DEFAULT_HANDLING_TIME_DAYS,
    DEFAULT_MERCHANT_SHIPPING_GROUP,
    DEFAULT_VARIANT_QUANTITY,
    MERCHANT_SHIPPING_GROUP_OPTIONS,
    build_listing_memory_path,
    build_listing_memory_payload,
    normalize_handling_time_days,
    normalize_merchant_shipping_group,
    normalize_variant_quantity,
)
from services.christmas_project_grouping import (
    build_christmas_group_image_manifest,
    derive_christmas_group_members,
    is_grouped_christmas_memory,
)
from services.christmas_group_publication import (
    ChristmasGroupPublicationStorage,
    is_group_submission_locked,
    is_ready_listing_visible,
    publish_christmas_group,
)
from services.quality_checks import (
    build_child_title_for_validation,
    find_oversized_child_titles,
    validate_listing_quality,
    words_repeated_at_least,
)
from services.runtime_flags import dev_tools_enabled
from services.staged_listing_tasks import create_staged_listing_task
from services.stock_references import (
    MAX_AMAZON_SKU_LENGTH,
    build_child_sku_details,
    get_stock_reference,
    has_stock_reference,
    is_strict_stock_ready,
    resolve_sku_decoration_code,
    validate_stock_ready_skus,
)
from ui.approved_output import render_approved_output
from ui.listing_content import clear_grouped_christmas_session_state, render_listing_content
from ui.product_setup import render_product_setup, render_product_setup_controls
from ui.review_queue import render_review_queue

from utils.dropbox_client import (
    get_or_create_shared_link,
    format_dropbox_error,
    to_direct_url,
    list_folder_files,
    list_folder_names,
    create_folder_if_missing,
    create_folder_exclusive,
    copy_dropbox_file,
    move_dropbox_folder,
    path_exists,
    path_exists_strict,
    upload_text_file,
    upload_binary_file,
    download_text_file,
)

GLOBAL_BRAND_NAME = "sloganitto"
WORKFLOW_ASSIGNEES = ["", "Sal", "Suleman", "Khalid"]

BASE_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = BASE_DIR / "templates"
CONFIG_DIR = BASE_DIR / "config"
OUTPUT_DIR = BASE_DIR / "outputs"

LOAD_EVENT_LIMIT = 160
FORBIDDEN_TITLE_PHRASES = [
    "Mother's Day Gift",
]


def normalize_title_phrase_for_app(value: str) -> str:
    value = (value or "").lower().replace("’", "'").replace("`", "'")
    value = value.replace("'", "")
    return " ".join(re.sub(r"[^a-z0-9]+", " ", value).split())


def find_forbidden_title_phrases_for_app(value: str) -> list[str]:
    normalized_title = normalize_title_phrase_for_app(value)
    return [
        phrase
        for phrase in FORBIDDEN_TITLE_PHRASES
        if normalize_title_phrase_for_app(phrase) in normalized_title
    ]


def reset_load_events() -> None:
    st.session_state["current_load_events"] = []
    st.session_state["current_rerun_started_at"] = time.perf_counter()


def record_load_event(label: str, started_at: float, detail: str = "") -> None:
    try:
        elapsed_ms = round((time.perf_counter() - started_at) * 1000, 1)
        events = st.session_state.setdefault("current_load_events", [])
        events.append({
            "step": label,
            "ms": elapsed_ms,
            "detail": detail,
        })
        if len(events) > LOAD_EVENT_LIMIT:
            st.session_state["current_load_events"] = events[-LOAD_EVENT_LIMIT:]
    except Exception:
        # Loading debug must never break the app.
        pass


def format_folder_detail(folder_path: str) -> str:
    folder_path = str(folder_path or "").rstrip("/")
    return folder_path.split("/")[-1] if folder_path else ""


def get_cached_folder_names(cache_key: str, root_path: str, label: str) -> list[str]:
    cache = st.session_state.setdefault("dropbox_folder_list_cache", {})
    load_errors = st.session_state.setdefault("dropbox_folder_load_errors", {})
    cached = cache.get(cache_key)

    if cached and cached.get("root_path") == root_path:
        folder_names = list(cached.get("folder_names", []))
        record_load_event(
            f"Dropbox: cached {label}",
            time.perf_counter(),
            f"{len(folder_names)} folder(s)",
        )
        return folder_names

    started_at = time.perf_counter()
    try:
        folder_names = list_folder_names(root_path)
    except Exception as exc:
        error_message = format_dropbox_error(exc)
        load_errors[cache_key] = f"{label}: {error_message}"
        if cached and cached.get("root_path") == root_path:
            folder_names = list(cached.get("folder_names", []))
            record_load_event(
                f"Dropbox: cached after failed {label}",
                started_at,
                f"{len(folder_names)} folder(s); {error_message}",
            )
            return folder_names

        record_load_event(
            f"Dropbox: failed {label}",
            started_at,
            error_message,
        )
        return []

    load_errors.pop(cache_key, None)
    record_load_event(
        f"Dropbox: list {label}",
        started_at,
        f"{len(folder_names)} folder(s)",
    )

    cache[cache_key] = {
        "root_path": root_path,
        "folder_names": folder_names,
    }
    return folder_names


def refresh_cached_folder_names(*cache_keys: str) -> None:
    cache = st.session_state.setdefault("dropbox_folder_list_cache", {})
    load_errors = st.session_state.setdefault("dropbox_folder_load_errors", {})
    for cache_key in cache_keys:
        cache.pop(cache_key, None)
        load_errors.pop(cache_key, None)


def clear_cached_listing_memory(*folder_paths: str) -> None:
    cache = st.session_state.setdefault("listing_memory_cache", {})

    if not folder_paths:
        cache.clear()
        return

    for folder_path in folder_paths:
        cache.pop(str(folder_path or "").rstrip("/"), None)


DEBUG_STATE_SKIP_KEYS = {
    "current_load_events",
    "current_rerun_started_at",
    "last_debug_state_snapshot",
    "current_rerun_changed_keys",
    "perf_history",
    "last_perf_saved_signature",
}

DEBUG_STATE_SKIP_PREFIXES = (
    "_",
)


def normalize_debug_state_value(value: Any) -> str:
    try:
        if isinstance(value, (str, int, float, bool, type(None))):
            return repr(value)
        if isinstance(value, (list, tuple, set)):
            return f"{type(value).__name__}(len={len(value)})::{repr(list(value)[:8])}"
        if isinstance(value, dict):
            keys = list(value.keys())[:8]
            return f"dict(len={len(value)}, keys={keys})"
        return f"{type(value).__name__}::{repr(value)[:180]}"
    except Exception:
        return f"{type(value).__name__}::<unreadable>"


def build_debug_state_snapshot() -> dict[str, str]:
    snapshot: dict[str, str] = {}

    for key, value in st.session_state.items():
        key_text = str(key)

        if key_text in DEBUG_STATE_SKIP_KEYS:
            continue

        if key_text.startswith(DEBUG_STATE_SKIP_PREFIXES):
            continue

        snapshot[key_text] = normalize_debug_state_value(value)

    return snapshot


def capture_rerun_cause() -> None:
    try:
        previous = dict(st.session_state.get("last_debug_state_snapshot", {}))
        current = build_debug_state_snapshot()

        changed_keys: list[dict[str, str]] = []
        all_keys = sorted(set(previous.keys()) | set(current.keys()))

        for key in all_keys:
            before = previous.get(key, "<missing>")
            after = current.get(key, "<missing>")

            if before != after:
                changed_keys.append({
                    "key": key,
                    "before": before[:220],
                    "after": after[:220],
                })

        st.session_state["current_rerun_changed_keys"] = changed_keys[:80]
    except Exception as exc:
        st.session_state["current_rerun_changed_keys"] = [{
            "key": "debug_error",
            "before": "",
            "after": str(exc),
        }]


def save_debug_state_snapshot() -> None:
    try:
        st.session_state["last_debug_state_snapshot"] = build_debug_state_snapshot()
    except Exception:
        pass


def consume_pending_perf_action_label() -> None:
    pending_label = str(st.session_state.pop("pending_perf_action_label", "")).strip()
    if pending_label:
        st.session_state["active_perf_action_label"] = pending_label


def infer_perf_action_label_from_changed_keys() -> str:
    changed_keys = [
        str(row.get("key", ""))
        for row in st.session_state.get("current_rerun_changed_keys", [])
    ]

    if not changed_keys:
        return ""

    debug_keys = {
        "show_loading_debug_inline",
        "perf_action_label",
        "clear_perf_history_btn",
        "download_perf_history_csv",
    }

    non_debug_keys = [
        key for key in changed_keys
        if key and key not in debug_keys
    ]

    if not non_debug_keys:
        return "debug/profiler toggle"

    key_set = set(non_debug_keys)

    if "load_image_mappings_now" in key_set or "image_mappings_loaded_folder" in key_set:
        return "load/refresh image mappings"

    if "staged_folder_select" in key_set:
        return "select staged folder"

    if "folder_source_mode" in key_set:
        return "change folder source"

    if "template_family_select" in key_set or "listing_template_select" in key_set:
        return "change template"

    if "parent_main_image_choice" in key_set:
        return "change parent main image"

    if "title_input" in key_set:
        return "edit title"

    if any(key.startswith("bullet_") for key in key_set):
        return "edit bullets"

    if "product_description" in key_set:
        return "edit description"

    if "generic_keywords" in key_set:
        return "edit search terms"

    if "variant_quantity" in key_set:
        return "change quantity"

    if any(key.startswith("price_") for key in key_set):
        return "change price"

    if any("selected_variant" in key or key.startswith("variant_") for key in key_set):
        return "change variants"

    if "ready_queue_review_folder" in key_set:
        return "select review queue item"

    if "review_queue_reviewed_by" in key_set:
        return "change reviewer"

    if "approved_queue_review_folder" in key_set:
        return "select approved review item"

    if "approved_queue_selected_folders" in key_set:
        return "select approved folders"

    if "review_queue_tab_loaded" in key_set:
        return "load review queue"

    if "approved_output_tab_loaded" in key_set:
        return "load approved output"

    preview = ", ".join(non_debug_keys[:4])
    return f"rerun: {preview}"


def get_current_perf_action_label() -> str:
    active_label = str(st.session_state.get("active_perf_action_label", "")).strip()
    manual_label = str(st.session_state.get("perf_action_label", "")).strip()
    inferred_label = infer_perf_action_label_from_changed_keys()

    # One-shot button labels win. Manual label is useful for controlled test sessions.
    # If neither exists, infer from changed Streamlit session-state keys.
    return active_label or manual_label or inferred_label or "(unlabeled)"


def build_current_perf_summary() -> dict[str, Any]:
    events = list(st.session_state.get("current_load_events", []))
    rerun_started_at = st.session_state.get("current_rerun_started_at")

    full_rerun_ms = None
    if rerun_started_at:
        full_rerun_ms = round((time.perf_counter() - float(rerun_started_at)) * 1000, 1)

    recorded_load_ms = round(
        sum(float(event.get("ms", 0) or 0) for event in events),
        1,
    )

    estimated_ui_ms = None
    if full_rerun_ms is not None:
        estimated_ui_ms = round(max(full_rerun_ms - recorded_load_ms, 0), 1)

    slowest_event = ""
    slowest_ms = 0.0
    if events:
        slowest = max(events, key=lambda event: float(event.get("ms", 0) or 0))
        slowest_event = str(slowest.get("step", ""))
        slowest_ms = float(slowest.get("ms", 0) or 0)

    return {
        "events": events,
        "full_rerun_ms": full_rerun_ms,
        "recorded_load_ms": recorded_load_ms,
        "estimated_ui_build_ms": estimated_ui_ms,
        "slowest_event": slowest_event,
        "slowest_ms": round(slowest_ms, 1),
        "event_count": len(events),
    }


def should_skip_perf_history_row(action_label: str) -> bool:
    changed_keys = [
        row.get("key", "")
        for row in st.session_state.get("current_rerun_changed_keys", [])
    ]

    debug_only_keys = {
        "show_loading_debug_inline",
        "perf_action_label",
        "clear_perf_history_btn",
        "download_perf_history_csv",
    }

    if changed_keys and all(key in debug_only_keys for key in changed_keys):
        return True

    return False


def save_current_perf_run() -> None:
    try:
        summary = build_current_perf_summary()
        events = summary["events"]

        if not events:
            return

        action_label = get_current_perf_action_label()

        if should_skip_perf_history_row(action_label):
            return

        run_signature = {
            "action": action_label,
            "events": events,
            "full_rerun_ms": summary["full_rerun_ms"],
            "recorded_load_ms": summary["recorded_load_ms"],
            "estimated_ui_build_ms": summary["estimated_ui_build_ms"],
            "slowest_event": summary["slowest_event"],
            "slowest_ms": summary["slowest_ms"],
        }

        signature_text = json.dumps(run_signature, sort_keys=True, default=str)
        last_saved_signature = st.session_state.get("last_perf_saved_signature", "")

        if signature_text == last_saved_signature:
            return

        history = st.session_state.setdefault("perf_history", [])

        history.append({
            "run": len(history) + 1,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "action": action_label,
            "full_rerun_ms": summary["full_rerun_ms"],
            "recorded_load_ms": summary["recorded_load_ms"],
            "estimated_ui_build_ms": summary["estimated_ui_build_ms"],
            "slowest_event": summary["slowest_event"],
            "slowest_ms": summary["slowest_ms"],
            "event_count": summary["event_count"],
        })

        if len(history) > 300:
            st.session_state["perf_history"] = history[-300:]

        st.session_state["last_perf_saved_signature"] = signature_text

        # One-shot button labels should not leak into later debug toggles/reruns.
        st.session_state.pop("active_perf_action_label", None)
    except Exception:
        pass


def render_inline_loading_debug() -> None:
    save_current_perf_run()

    st.divider()

    control_col1, control_col2 = st.columns([1, 4])
    with control_col1:
        show_debug = st.checkbox(
            "Show profiler",
            key="show_loading_debug_inline",
            value=False,
        )
    with control_col2:
        st.text_input(
            "Manual action label for next test",
            key="perf_action_label",
            placeholder="Optional label: edit title, change reviewer, load images...",
            label_visibility="collapsed",
        )

    if not show_debug:
        return

    st.subheader("Loading / render debug")

    summary = build_current_perf_summary()
    events = summary["events"]

    metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
    metric_col1.metric("Full rerun", f"{summary['full_rerun_ms'] or 0} ms")
    metric_col2.metric("Recorded load", f"{summary['recorded_load_ms']} ms")
    metric_col3.metric("UI/build estimate", f"{summary['estimated_ui_build_ms'] or 0} ms")
    metric_col4.metric("Events", str(summary["event_count"]))

    if summary["slowest_event"]:
        st.caption(
            f"Slowest recorded load step: {summary['slowest_event']} ({summary['slowest_ms']} ms)"
        )

    with st.expander("Current rerun events", expanded=False):
        if not events:
            st.warning("No loading events recorded for this rerun.")
        else:
            rows = [
                {
                    "step": event.get("step", ""),
                    "ms": event.get("ms", ""),
                    "detail": event.get("detail", ""),
                }
                for event in events
            ]
            st.dataframe(rows, hide_index=True, width="stretch")

    approved_generation_step_rows = list(st.session_state.get("approved_generation_step_rows", []))
    if approved_generation_step_rows:
        with st.expander("Last approved generation step breakdown", expanded=True):
            st.dataframe(approved_generation_step_rows, hide_index=True, width="stretch")

    history = list(st.session_state.get("perf_history", []))
    st.markdown("### Performance history")

    clear_col, download_col = st.columns([1, 3])
    with clear_col:
        if st.button("Clear perf history", key="clear_perf_history_btn", width="stretch"):
            st.session_state["perf_history"] = []
            st.session_state.pop("last_perf_saved_signature", None)
            st.session_state.pop("active_perf_action_label", None)
            st.session_state.pop("perf_action_label", None)
            st.rerun()

    if not history:
        st.caption("No completed runs saved yet.")
        return

    st.dataframe(history[-50:], hide_index=True, width="stretch")

    csv_lines = [
        "run,timestamp,action,full_rerun_ms,recorded_load_ms,estimated_ui_build_ms,slowest_event,slowest_ms,event_count"
    ]

    for row in history:
        values = [
            row.get("run", ""),
            row.get("timestamp", ""),
            str(row.get("action", "")).replace('"', '""'),
            row.get("full_rerun_ms", ""),
            row.get("recorded_load_ms", ""),
            row.get("estimated_ui_build_ms", ""),
            str(row.get("slowest_event", "")).replace('"', '""'),
            row.get("slowest_ms", ""),
            row.get("event_count", ""),
        ]
        csv_lines.append(
            ",".join(
                f'"{value}"' if isinstance(value, str) and "," in value else str(value)
                for value in values
            )
        )

    with download_col:
        st.download_button(
            "Download performance history CSV",
            data="\n".join(csv_lines).encode("utf-8"),
            file_name="amazon_lister_performance_history.csv",
            mime="text/csv",
            key="download_perf_history_csv",
        )

def render_rerun_cause_debug() -> None:
    if not st.session_state.get("show_loading_debug_inline", False):
        return

    changed_keys = list(st.session_state.get("current_rerun_changed_keys", []))

    st.markdown("### Rerun cause tracker")
    st.caption(
        "These are Streamlit session-state keys that changed since the previous completed rerun. "
        "This helps identify which widget/action triggered the loading spinner."
    )

    if not changed_keys:
        st.success("No session-state changes detected from the previous completed rerun.")
        return

    summary = ", ".join(row.get("key", "") for row in changed_keys[:12])
    st.info(f"Likely trigger key(s): {summary}")

    with st.expander("Changed session-state keys", expanded=False):
        st.dataframe(changed_keys, hide_index=True, width="stretch")

SHEET_NAME = "Template"
HEADER_ROW = 3
PARENT_ROW = 4
FIRST_CHILD_ROW = 5

SKU_DECORATION_OPTIONS = ["DEF", "EMB", "PERSO", "PLAIN", "Custom"]


def get_workbook_layout(profile: dict[str, Any]) -> dict[str, Any]:
    schema = profile.get("_schema", {})

    def as_int(key: str, fallback: int) -> int:
        value = profile.get(key, schema.get(key, fallback))
        try:
            return int(value)
        except (TypeError, ValueError):
            return fallback

    return {
        "mode": str(profile.get("workbook_mode", schema.get("workbook_mode", "")) or "").strip(),
        "header_row": as_int("workbook_header_row", HEADER_ROW),
        "parent_row": as_int("workbook_parent_row", PARENT_ROW),
        "first_child_row": as_int("workbook_first_child_row", FIRST_CHILD_ROW),
    }


def load_dropbox_templates_config() -> dict[str, Any]:
    config_path = CONFIG_DIR / "dropbox_templates.json"
    if not config_path.exists():
        return {}
    with config_path.open("r", encoding="utf-8-sig") as f:
        return json.load(f)


def load_stock_references_config() -> dict[str, Any]:
    config_path = CONFIG_DIR / "stock_references.json"
    if not config_path.exists():
        return {}
    try:
        with config_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return {}
    references = data.get("references", data)
    return references if isinstance(references, dict) else {}


def list_template_profiles() -> list[dict[str, Any]]:
    profiles: list[dict[str, Any]] = []
    if not TEMPLATES_DIR.exists():
        return profiles

    stock_references = load_stock_references_config()

    for family_folder in sorted(TEMPLATES_DIR.iterdir()):
        if not family_folder.is_dir():
            continue

        schema_path = family_folder / "schema.json"
        if not schema_path.exists():
            continue

        try:
            with schema_path.open("r", encoding="utf-8") as f:
                schema = json.load(f)
        except Exception:
            continue

        for garment_folder in sorted(family_folder.iterdir()):
            if not garment_folder.is_dir():
                continue

            config_path = garment_folder / "config.json"
            if not config_path.exists():
                continue

            try:
                with config_path.open("r", encoding="utf-8") as f:
                    config = json.load(f)

                config["_folder"] = garment_folder
                config["_slug"] = garment_folder.name
                config["_family_folder"] = family_folder
                config["_family_slug"] = family_folder.name
                config["_schema"] = schema

                # family owns workbook now
                config["template_file"] = schema.get("workbook_file", "")

                stock_reference_key = str(config.get("stock_reference_key", "") or "").strip()
                if stock_reference_key:
                    stock_reference = stock_references.get(stock_reference_key, {})
                    if isinstance(stock_reference, dict):
                        config["_stock_reference"] = dict(stock_reference)

                profiles.append(config)
            except Exception:
                continue

    return profiles


def get_default(profile: dict[str, Any], key: str, fallback: Any = "") -> Any:
    return profile.get(key, fallback)


@st.cache_data(show_spinner=False)
def dropbox_preview_url(path: str) -> str:
    if not path:
        return ""

    try:
        shared = get_or_create_shared_link(path)
        return to_direct_url(shared)
    except Exception as exc:
        raise FileNotFoundError(f"Dropbox preview failed for {path}: {exc}") from exc


def get_cached_dropbox_shared_link(path: str) -> str:
    path = str(path or "").strip()
    if not path:
        return ""

    cache = st.session_state.setdefault("dropbox_shared_link_cache", {})
    if path in cache:
        return str(cache[path])

    shared_link = get_or_create_shared_link(path)
    cache[path] = shared_link
    return shared_link


def render_dropbox_folder_links(
    source_folder_path: str | None,
    dropbox_overview: dict[str, Any],
) -> None:
    st.markdown("**Dropbox folders**")

    folder_rows: list[dict[str, str]] = []

    if source_folder_path:
        folder_rows.append({
            "label": "Listing folder",
            "path": source_folder_path,
        })

    resource_root = str(dropbox_overview.get("resource_root", "") or "").strip()
    garment_resource_root = str(dropbox_overview.get("garment_resource_root", "") or "").strip()

    if garment_resource_root:
        folder_rows.append({
            "label": "Garment resources",
            "path": garment_resource_root,
        })

    if resource_root:
        folder_rows.append({
            "label": "Shared resources root",
            "path": resource_root,
        })

    if not folder_rows:
        st.caption("No Dropbox folder links available.")
        return

    for row in folder_rows:
        label = row["label"]
        path = row["path"]

        try:
            shared_link = get_cached_dropbox_shared_link(path)
            st.markdown(f"- **{label}:** [{path}]({shared_link})")
        except Exception:
            st.markdown(f"- **{label}:** `{path}`")


def build_header_map(ws, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is not None:
            key = str(value).strip()
            if key:
                mapping[key] = col
    return mapping


def copy_row_format(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)

        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def clear_row_values(ws, row_idx: int) -> None:
    for col in range(1, ws.max_column + 1):
        ws.cell(row_idx, col).value = None


def set_field(ws, row_idx: int, header_map: dict[str, int], field: str, value: Any) -> bool:
    col = header_map.get(field)
    if col is None:
        return False

    ws.cell(row_idx, col).value = value
    return True

def write_values_with_debug(
    ws,
    row_idx: int,
    header_map: dict[str, int],
    values: dict[str, Any],
    row_label: str,
) -> None:
    missing_fields: list[str] = []

    for field, value in values.items():
        written = set_field(ws, row_idx, header_map, field, value)
        if not written:
            missing_fields.append(field)

    if missing_fields and st.session_state.get("show_header_debug", False):
        st.warning(f"{row_label}: {len(missing_fields)} field(s) not found in template headers")
        st.code("\n".join(missing_fields), language=None)

def normalize_size(size: str) -> str:
    size_map = {
        "2XL": "XXL",
        "XXL": "XXL",
        "3XL": "3XL",
        "4XL": "4XL",
        "5XL": "5XL",
        "6XL": "6XL",
    }
    return size_map.get(size, size)


def format_year_size(year: int) -> str:
    return f"{year} Year" if year == 1 else f"{year} Years"


def get_profile_size_range_end(profile: dict[str, Any], size: str) -> str:
    aliases_by_dimension = profile.get("saved_variant_value_aliases", {})
    size_aliases = (
        aliases_by_dimension.get("size", {})
        if isinstance(aliases_by_dimension, dict)
        else {}
    )
    if not isinstance(size_aliases, dict):
        return ""

    normalized_size = str(size or "").strip().casefold()
    for source_size, target_size in size_aliases.items():
        if str(target_size or "").strip().casefold() != normalized_size:
            continue
        compact_source = str(source_size or "").replace("/", "-")
        match = re.search(r"(\d+)\s*-\s*(\d+)", compact_source)
        if match:
            return format_year_size(int(match.group(2)))
    return ""


def get_amazon_shirt_size_range(
    size: str,
    profile: dict[str, Any] | None = None,
) -> tuple[str, str]:
    normalized_size = normalize_size(str(size or "").strip())
    if not is_child_size_label(normalized_size):
        return normalized_size, ""

    compact_size = normalized_size.replace("/", "-")
    match = re.search(r"(\d+)\s*-\s*(\d+)", compact_size)
    if match:
        start_year = int(match.group(1))
        end_year = int(match.group(2))
        return format_year_size(start_year), format_year_size(end_year)

    match = re.search(r"(\d+)", normalized_size)
    if match:
        return (
            format_year_size(int(match.group(1))),
            get_profile_size_range_end(profile or {}, normalized_size),
        )

    return normalized_size, ""


def get_row_age_range_description(
    profile: dict[str, Any],
    data: dict[str, Any],
    size_value: str,
) -> str:
    if is_child_size_label(size_value):
        return str(profile.get("child_age_range_description") or "Child")
    return str(profile.get("adult_age_range_description") or data.get("age_range_description", ""))


def get_row_department_name(
    profile: dict[str, Any],
    data: dict[str, Any],
    size_value: str,
) -> str:
    if is_child_size_label(size_value):
        return str(profile.get("child_department_name") or "unisex-child")
    return str(profile.get("adult_department_name") or data.get("department_name", ""))


def infer_model_name_from_supplier_key(supplier_stock_key: str) -> str:
    stock_key = str(supplier_stock_key or "").strip()
    match = re.match(r"^(\d{3,4})", stock_key)
    if match:
        return f"UC{match.group(1)}"
    return ""


def get_product_model_name(profile: dict[str, Any], data: dict[str, Any]) -> str:
    return str(
        profile.get("model_name")
        or data.get("model_name")
        or data.get("style_name")
        or profile.get("style_name")
        or data.get("item_type_name")
        or profile.get("item_type_name")
        or ""
    ).strip()


def get_garment_model_number(profile: dict[str, Any], data: dict[str, Any]) -> str:
    return str(
        data.get("base_parent_sku")
        or profile.get("parent_sku")
        or profile.get("template_key")
        or profile.get("parent_sku")
        or profile.get("_slug")
        or ""
    ).strip()


def get_child_model_number(
    profile: dict[str, Any],
    data: dict[str, Any],
    sku_details: dict[str, str],
) -> str:
    supplier_model = infer_model_name_from_supplier_key(sku_details.get("supplier_stock_key", ""))
    return supplier_model or get_garment_model_number(profile, data)


def is_variant_combo_allowed(profile: dict[str, Any], variant_values: dict[str, str]) -> bool:
    color = variant_values.get("color", "")
    size = variant_values.get("size", "")
    design = variant_values.get("design", "")

    color_size_map = profile.get("color_size_map", {})
    if color and size and color_size_map:
        allowed_sizes = color_size_map.get(color)
        if allowed_sizes is not None and size not in allowed_sizes:
            return False

    design_size_map = profile.get("design_size_map", {})
    if design and size and design_size_map:
        allowed_sizes = design_size_map.get(design)
        if allowed_sizes is not None and size not in allowed_sizes:
            return False

    design_color_map = profile.get("design_color_map", {})
    if design and color and design_color_map:
        allowed_colors = design_color_map.get(design)
        if allowed_colors is not None and color not in allowed_colors:
            return False

    return True


def build_variant_combinations(
    profile: dict[str, Any],
    selected_variants: dict[str, list[str]],
) -> list[dict[str, str]]:
    keys = list(selected_variants.keys())
    if not keys:
        return []

    value_lists = [selected_variants[k] for k in keys]
    combos: list[dict[str, str]] = []

    for values in product(*value_lists):
        combo = dict(zip(keys, values))
        if is_variant_combo_allowed(profile, combo):
            combos.append(combo)

    return combos


def build_variant_price_key(variant_values: dict[str, str]) -> str:
    design_value = str(variant_values.get("design", "") or "").strip()
    size_value = str(variant_values.get("size", "") or "").strip()
    if design_value and size_value:
        return f"{design_value}||{size_value}"
    return size_value or design_value or "default"


def get_variant_price_from_map(
    profile: dict[str, Any],
    size_price_map: dict[str, Any],
    variant_values: dict[str, str],
    fallback: Any = 0,
) -> Any:
    size_price_map = dict(size_price_map or {})
    price_key = build_variant_price_key(variant_values)
    if price_key in size_price_map:
        return size_price_map[price_key]

    size_value = str(variant_values.get("size", "") or "").strip()
    if size_value and size_value in size_price_map:
        return size_price_map[size_value]

    design_value = str(variant_values.get("design", "") or "").strip()
    if design_value and design_value in size_price_map:
        return size_price_map[design_value]

    if "default" in size_price_map:
        return size_price_map["default"]

    if size_value:
        return get_default_price_for_size(
            profile,
            size_value,
            size_price_map,
            design=design_value,
        )

    return fallback


def get_positive_variant_price_from_map(
    profile: dict[str, Any],
    size_price_map: dict[str, Any],
    variant_values: dict[str, str],
) -> float | None:
    try:
        price = float(get_variant_price_from_map(profile, size_price_map, variant_values, fallback=0))
    except (TypeError, ValueError):
        return None
    return price if price > 0 else None


def has_design_size_pricing(profile: dict[str, Any], selected_variants: dict[str, list[str]]) -> bool:
    return bool(
        selected_variants.get("design")
        and selected_variants.get("size")
        and any(str(dim.get("name", "")).strip().lower() == "design" for dim in profile.get("variant_dimensions", []))
    )


def normalize_variant_price_map_for_selected_variants(
    profile: dict[str, Any],
    selected_variants: dict[str, list[str]],
    size_price_map: dict[str, Any],
) -> dict[str, float]:
    raw_price_map = dict(size_price_map or {})
    if not has_design_size_pricing(profile, selected_variants):
        normalized: dict[str, float] = {}
        for key, value in raw_price_map.items():
            try:
                normalized[str(key)] = float(value)
            except (TypeError, ValueError):
                normalized[str(key)] = 0.0
        return normalized

    normalized = {}
    variant_combos = build_variant_combinations(
        profile,
        {
            "design": list(selected_variants.get("design", []) or []),
            "size": list(selected_variants.get("size", []) or []),
        },
    )
    for combo in variant_combos:
        price_key = build_variant_price_key(combo)
        try:
            normalized[price_key] = float(get_variant_price_from_map(profile, raw_price_map, combo, fallback=0))
        except (TypeError, ValueError):
            normalized[price_key] = 0.0
    return normalized


def sort_variant_combinations_by_price(
    profile: dict[str, Any],
    variant_combos: list[dict[str, str]],
    size_price_map: dict[str, float],
) -> list[dict[str, str]]:
    indexed = list(enumerate(variant_combos))

    def sort_key(item: tuple[int, dict[str, str]]) -> tuple[float, int]:
        idx, combo = item
        price = get_variant_price_from_map(profile, size_price_map, combo, fallback=999999)
        try:
            price_value = float(price)
        except (TypeError, ValueError):
            price_value = 999999
        return price_value, idx

    return [combo for _, combo in sorted(indexed, key=sort_key)]


def get_lowest_variant_price(size_price_map: dict[str, Any]) -> float | str:
    prices: list[float] = []
    for raw_price in dict(size_price_map or {}).values():
        try:
            price = float(raw_price)
        except (TypeError, ValueError):
            continue
        if price > 0:
            prices.append(price)

    if not prices:
        return ""

    return min(prices)


def get_selected_colors_for_image_resolution(
    profile: dict[str, Any],
    selected_variants: dict[str, list[str]],
) -> list[str]:
    variant_dimensions = profile.get("variant_dimensions", [])
    if variant_dimensions:
        for dim in variant_dimensions:
            dim_name = dim.get("name", "")
            if dim_name.lower() == "color":
                return list(selected_variants.get(dim_name, []))

    return list(selected_variants.get("color", []))


def build_child_sku(profile: dict[str, Any], parent_sku: str, variant_values: dict[str, str]) -> str:
    return build_child_sku_details(profile, parent_sku, variant_values)["amazon_seller_sku"]


def build_child_sku_length_report(
    profile: dict[str, Any],
    parent_sku: str,
    selected_variants: dict[str, list[str]],
    *,
    sku_decoration_code: str = "",
    sku_listing_code: str = "",
) -> dict[str, Any]:
    effective_profile = apply_sku_context_to_profile(profile, sku_decoration_code, sku_listing_code)
    combos = build_variant_combinations(profile, selected_variants)
    rows: list[dict[str, Any]] = []
    for combo in combos:
        sku = build_child_sku_details(effective_profile, parent_sku, combo)["amazon_seller_sku"]
        rows.append({"sku": sku, "length": len(sku), "variant": dict(combo)})
    rows.sort(key=lambda row: row["length"], reverse=True)
    return {
        "max_length": rows[0]["length"] if rows else 0,
        "longest": rows[:5],
        "oversized": [row for row in rows if row["length"] > MAX_AMAZON_SKU_LENGTH],
        "count": len(rows),
    }


def build_child_item_name(
    base_title: str,
    variant_values: dict[str, str],
    profile: dict[str, Any] | None = None,
) -> str:
    return build_child_title_for_validation(profile or {}, base_title, variant_values)


def get_variant_design_overrides(profile: dict[str, Any], variant_values: dict[str, str]) -> dict[str, Any]:
    design_value = str(variant_values.get("design", "") or "").strip()
    if not design_value:
        return {}

    overrides = profile.get("design_field_overrides", {})
    if not isinstance(overrides, dict):
        return {}

    raw_override = overrides.get(design_value, {})
    return dict(raw_override) if isinstance(raw_override, dict) else {}


def get_variant_size_display_label(profile: dict[str, Any], variant_values: dict[str, str]) -> str:
    size_value = str(variant_values.get("size", "") or "").strip()
    if not size_value:
        return ""

    design_value = str(variant_values.get("design", "") or "").strip()
    prefixes = profile.get("size_display_prefix_by_design", {})
    prefix = str(dict(prefixes or {}).get(design_value, "") or "").strip()
    return f"{prefix} - {size_value}" if prefix else size_value


def get_variant_color_display_label(profile: dict[str, Any], variant_values: dict[str, str]) -> str:
    color_value = str(variant_values.get("color", "") or "").strip()
    if not color_value:
        return ""

    design_value = str(variant_values.get("design", "") or "").strip()
    prefixes = profile.get("color_display_prefix_by_design", {})
    suffixes = profile.get("color_display_suffix_by_design", {})
    prefix = str(dict(prefixes or {}).get(design_value, "") or "").strip()
    suffix = str(dict(suffixes or {}).get(design_value, "") or "").strip()

    if prefix and suffix:
        return f"{prefix} {color_value} - {suffix}"
    if prefix:
        return f"{prefix} {color_value}"
    if suffix:
        return f"{color_value} - {suffix}"
    return color_value


def apply_sku_decoration_to_profile(profile: dict[str, Any], sku_decoration_code: str = "") -> dict[str, Any]:
    sku_decoration_code = str(sku_decoration_code or "").strip()
    if not sku_decoration_code:
        return profile

    effective_profile = dict(profile)
    effective_profile["sku_decoration_code"] = sku_decoration_code
    return effective_profile


def get_default_sku_decoration_code(profile: dict[str, Any], listing_memory: dict[str, Any] | None = None) -> str:
    saved_code = str((listing_memory or {}).get("sku_decoration_code", "") or "").strip()
    return saved_code or resolve_sku_decoration_code(profile)


def get_garment_sku_code(profile: dict[str, Any]) -> str:
    return sanitize_sku(
        str(
            profile.get("parent_sku")
            or profile.get("template_key")
            or profile.get("_slug")
            or "GARMENT"
        )
    ).upper()


def get_saved_generated_sku_listing_code(listing_memory: dict[str, Any] | None = None) -> str:
    listing_memory = listing_memory or {}
    saved_code = str(
        listing_memory.get("generated_sku_listing_code")
        or listing_memory.get("sku_listing_code")
        or ""
    ).strip()
    return sanitize_sku(saved_code).upper()


def get_or_create_generated_sku_listing_code(listing_memory: dict[str, Any] | None = None) -> str:
    saved_code = get_saved_generated_sku_listing_code(listing_memory)
    if saved_code:
        return saved_code

    if "generated_sku_listing_code" not in st.session_state:
        st.session_state["generated_sku_listing_code"] = f"D{generate_unique_sku(5)}"

    return sanitize_sku(str(st.session_state.get("generated_sku_listing_code", ""))).upper()


def build_parent_sku_from_context(
    profile: dict[str, Any],
    sku_decoration_code: str,
    sku_listing_code: str,
) -> str:
    include_template_code = bool(profile.get("include_template_code_in_parent_sku", True))
    parts = [
        sanitize_sku(str(sku_decoration_code or "")).upper(),
        sanitize_sku(str(sku_listing_code or "")).upper(),
    ]
    if include_template_code:
        parts.append(get_garment_sku_code(profile))
    return "-".join(part for part in parts if part)


def apply_sku_context_to_profile(
    profile: dict[str, Any],
    sku_decoration_code: str = "",
    sku_listing_code: str = "",
) -> dict[str, Any]:
    effective_profile = apply_sku_decoration_to_profile(profile, sku_decoration_code)
    sku_listing_code = sanitize_sku(str(sku_listing_code or "")).upper()
    if sku_listing_code:
        effective_profile = dict(effective_profile)
        effective_profile["design_or_listing_code"] = sku_listing_code
        effective_profile["listing_code"] = sku_listing_code
        effective_profile["sku_listing_code"] = sku_listing_code
    return effective_profile


def build_variant_field_values(profile: dict[str, Any], variant_values: dict[str, str]) -> dict[str, Any]:
    values: dict[str, Any] = {}

    if "color" in variant_values:
        values["color_name"] = get_variant_color_display_label(profile, variant_values)

    if "size" in variant_values:
        normalized_size = normalize_size(get_variant_size_display_label(profile, variant_values))
        values["size_name"] = normalized_size
        values["apparel_size"] = normalized_size

    if "design" in variant_values:
        values["style_name"] = variant_values["design"]

    return values


def get_apparel_size_class(size_value: str) -> str:
    return "Age" if is_child_size_label(size_value) else "Alpha"


def apply_apparel_size_fields(
    values: dict[str, Any],
    size_value: str,
    *,
    is_apparel: bool,
    profile: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if not is_apparel:
        return values

    normalized_size = normalize_size(size_value) if size_value else ""
    size_class = get_apparel_size_class(normalized_size)
    shirt_size, shirt_size_to = get_amazon_shirt_size_range(
        normalized_size,
        profile,
    )
    body_type = "" if size_class == "Age" else str(
        values.get("apparel_body_type")
        or values.get("shirt_body_type")
        or ("Regular" if normalized_size else "")
    )
    height_type = "" if size_class == "Age" else str(
        values.get("apparel_height_type")
        or values.get("shirt_height_type")
        or ("Regular" if normalized_size else "")
    )

    size_values = {
        "apparel_size_system": "UK",
        "apparel_size_class": size_class,
        "apparel_size": shirt_size,
        "apparel_size_to": shirt_size_to,
        "size_map": normalized_size,
        "apparel_body_type": body_type,
        "apparel_height_type": height_type,
        "shirt_size_system": "UK",
        "shirt_size_class": size_class,
        "shirt_size": shirt_size,
        "shirt_size_to": shirt_size_to,
        "shirt_body_type": body_type,
        "shirt_height_type": height_type,
    }

    for field, value in size_values.items():
        values[field] = value

    return values


def validate_variant_dimensions(selected_variants: dict[str, list[str]]) -> list[str]:
    errors: list[str] = []

    for dim_name, items in selected_variants.items():
        if not items:
            errors.append(f"At least one option is required for {dim_name}.")

    return errors

def slugify_part(value: str) -> str:
    safe = value.strip().replace(" ", "-").replace("/", "-")
    while "--" in safe:
        safe = safe.replace("--", "-")
    return safe

def sanitize_sku(value: str) -> str:
    safe = value.strip()
    for ch in ['\\', '/', ':', '*', '?', '"', '<', '>', '|', ' ']:
        safe = safe.replace(ch, '-')
    while '--' in safe:
        safe = safe.replace('--', '-')
    return safe.strip('-')


def generate_unique_sku(length: int = 6) -> str:
    alphabet = string.ascii_uppercase + string.digits
    return "".join(random.choices(alphabet, k=length))


def build_final_folder_sku(parent_sku: str, unique_sku: str) -> str:
    return f"{unique_sku}-{sanitize_sku(parent_sku)}"


def build_stage_folder_path(dropbox_cfg: dict[str, Any], staged_folder_name: str) -> str:
    stage_root = dropbox_cfg.get("stage_root", "").rstrip("/")
    return f"{stage_root}/{staged_folder_name}"


def sanitize_stage_folder_name(value: str) -> str:
    cleaned = str(value or "").strip()
    for ch in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
        cleaned = cleaned.replace(ch, "-")
    cleaned = re.sub(r"\s+", " ", cleaned)
    cleaned = re.sub(r"-{2,}", "-", cleaned)
    return cleaned.strip(" .-")


def stage_folder_lookup_key(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def build_stage_folder_name_candidates(staged_folder_name: str) -> list[str]:
    original = str(staged_folder_name or "").strip()
    candidates = [
        original,
        sanitize_stage_folder_name(original),
        sanitize_sku(original),
        original.replace("-", " "),
        original.replace(" ", "-"),
    ]
    seen: set[str] = set()
    unique_candidates: list[str] = []
    for candidate in candidates:
        candidate = str(candidate or "").strip()
        if candidate and candidate not in seen:
            seen.add(candidate)
            unique_candidates.append(candidate)
    return unique_candidates


def resolve_existing_stage_folder_name(
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    known_stage_folder_names: list[str] | None = None,
) -> str:
    folder_name = str(staged_folder_name or "").strip()
    if not folder_name:
        raise ValueError("Select a staged Dropbox folder.")

    candidates = build_stage_folder_name_candidates(folder_name)
    known_stage_folder_names = list(known_stage_folder_names or [])
    known_by_lookup = {
        stage_folder_lookup_key(name): name
        for name in known_stage_folder_names
        if stage_folder_lookup_key(name)
    }

    for candidate in list(candidates):
        exact_known = next((name for name in known_stage_folder_names if name.lower() == candidate.lower()), "")
        if exact_known:
            return exact_known

        lookup_match = known_by_lookup.get(stage_folder_lookup_key(candidate), "")
        if lookup_match:
            return lookup_match

    tried_paths: list[str] = []
    for candidate in candidates:
        candidate_path = build_stage_folder_path(dropbox_cfg, candidate)
        tried_paths.append(candidate_path)
        if path_exists(candidate_path):
            return candidate

    tried_text = ", ".join(f"`{path}`" for path in tried_paths)
    raise FileNotFoundError(f"Staged folder was not found in Dropbox. Tried: {tried_text}")


def build_finished_folder_path(dropbox_cfg: dict[str, Any], final_sku: str) -> str:
    finished_root = dropbox_cfg.get("finished_root", "").rstrip("/")
    return f"{finished_root}/{final_sku}"


def build_ready_folder_path(dropbox_cfg: dict[str, Any], ready_folder_name: str) -> str:
    ready_root = dropbox_cfg.get("ready_root", "").rstrip("/")
    return f"{ready_root}/{ready_folder_name}"


def build_approved_folder_path(dropbox_cfg: dict[str, Any], approved_folder_name: str) -> str:
    approved_root = dropbox_cfg.get("approved_root", "").rstrip("/")
    return f"{approved_root}/{approved_folder_name}"


def restage_finished_dropbox_folder(
    dropbox_cfg: dict[str, Any],
    finished_folder_name: str,
) -> str:
    stage_root = dropbox_cfg.get("stage_root", "").rstrip("/")
    finished_root = dropbox_cfg.get("finished_root", "").rstrip("/")

    if not finished_folder_name:
        raise ValueError("Finished folder name is required.")

    source_path = f"{finished_root}/{finished_folder_name}"

    candidate_name = f"{finished_folder_name}_restaged"
    target_path = f"{stage_root}/{candidate_name}"

    counter = 1
    while path_exists(target_path):
        candidate_name = f"{finished_folder_name}_restaged_{counter}"
        target_path = f"{stage_root}/{candidate_name}"
        counter += 1

    moved_path = move_dropbox_folder(source_path, target_path)
    return moved_path


def move_finished_dropbox_folder_to_approved(
    dropbox_cfg: dict[str, Any],
    finished_folder_name: str,
) -> str:
    approved_root = dropbox_cfg.get("approved_root", "").rstrip("/")
    finished_root = dropbox_cfg.get("finished_root", "").rstrip("/")

    if not finished_folder_name:
        raise ValueError("Finished folder name is required.")

    source_path = f"{finished_root}/{finished_folder_name}"
    candidate_name = finished_folder_name
    target_path = f"{approved_root}/{candidate_name}"

    counter = 1
    while path_exists(target_path):
        candidate_name = f"{finished_folder_name}_approved_{counter}"
        target_path = f"{approved_root}/{candidate_name}"
        counter += 1

    moved_path = move_dropbox_folder(source_path, target_path)
    return moved_path


def reset_restaged_selection_state() -> None:
    st.session_state["last_loaded_listing_memory_folder"] = ""
    st.session_state.pop("finalized_stage_folder", None)
    st.session_state.pop("finalized_finished_folder_path", None)
    st.session_state.pop("finalized_sku", None)
    st.session_state.pop("last_detected_template_folder", None)
    st.session_state.pop("applied_listing_memory_key_v2", None)
    st.session_state.pop("applied_listing_memory_widget_key_v2", None)
    st.session_state.pop("initialized_listing_context_key", None)
    st.session_state.pop("last_loaded_listing_memory_signature", None)


def restage_finished_listing_for_review(
    dropbox_cfg: dict[str, Any],
    profiles: list[dict[str, Any]],
    fallback_profile: dict[str, Any],
    finished_folder_name: str,
    target_state: str = "stage",
) -> dict[str, Any]:
    finished_folder_name = str(finished_folder_name or "").strip()
    target_state = str(target_state or "stage").strip().lower()
    if target_state not in {"stage", "approved"}:
        raise ValueError("target_state must be either 'stage' or 'approved'.")

    old_finished_path = build_finished_folder_path(dropbox_cfg, finished_folder_name)
    result = {
        "old_finished_folder_name": finished_folder_name,
        "old_finished_folder_path": old_finished_path,
        "new_staged_folder_name": "",
        "new_staged_folder_path": "",
        "new_approved_folder_name": "",
        "new_approved_folder_path": "",
        "target_state": target_state,
        "status": "Failed",
        "warning": "",
        "error": "",
    }

    try:
        if target_state == "approved":
            moved_path = move_finished_dropbox_folder_to_approved(
                dropbox_cfg=dropbox_cfg,
                finished_folder_name=finished_folder_name,
            )
            result["new_approved_folder_path"] = moved_path
            result["new_approved_folder_name"] = Path(moved_path).name
        else:
            moved_path = restage_finished_dropbox_folder(
                dropbox_cfg=dropbox_cfg,
                finished_folder_name=finished_folder_name,
            )
            result["new_staged_folder_path"] = moved_path
            result["new_staged_folder_name"] = Path(moved_path).name

        warning_messages: list[str] = []
        listing_memory_path = build_listing_memory_path(moved_path)
        try:
            if not path_exists(listing_memory_path):
                warning_messages.append("listing_inputs.json was missing; created a minimal restage memory file.")
                restaged_listing_memory = {}
            else:
                restaged_listing_memory = load_listing_memory_from_dropbox(moved_path)
                if not restaged_listing_memory:
                    warning_messages.append("listing_inputs.json was empty; saved restage metadata.")
        except Exception as exc:
            warning_messages.append(f"listing_inputs.json could not be loaded: {exc}")
            restaged_listing_memory = {}

        restaged_listing_memory["original_finished_folder_name"] = finished_folder_name
        append_workflow_event(
            restaged_listing_memory,
            action="restore_finished_to_approved" if target_state == "approved" else "restage_finished_listing",
            actor="",
            from_state="finished",
            to_state="approved" if target_state == "approved" else "stage",
            folder_path=moved_path,
            details={
                "original_finished_folder_name": finished_folder_name,
                "old_finished_folder_name": finished_folder_name,
                "old_finished_folder_path": old_finished_path,
            },
        )

        restaged_profile = find_profile_for_listing_memory(profiles, restaged_listing_memory) or fallback_profile
        save_listing_inputs_json_to_dropbox(
            profile=restaged_profile,
            payload=restaged_listing_memory,
            folder_path=moved_path,
        )

        result["status"] = "Success"
        result["warning"] = " ".join(warning_messages)
    except Exception as exc:
        result["error"] = str(exc)

    return result


def mark_finished_generation_ignored(
    dropbox_cfg: dict[str, Any],
    profiles: list[dict[str, Any]],
    fallback_profile: dict[str, Any],
    finished_folder_name: str,
    reason: str = "",
    actor: str = "",
) -> dict[str, Any]:
    finished_folder_name = str(finished_folder_name or "").strip()
    finished_folder_path = build_finished_folder_path(dropbox_cfg, finished_folder_name)
    result = {
        "folder_name": finished_folder_name,
        "finished_folder_path": finished_folder_path,
        "status": "Failed",
        "message": "",
    }

    try:
        if not finished_folder_name:
            raise ValueError("Finished folder name is required.")
        if not path_exists(finished_folder_path):
            raise FileNotFoundError(f"Finished folder not found: {finished_folder_path}")

        listing_memory = load_listing_memory_from_dropbox(finished_folder_path)
        if not listing_memory:
            raise FileNotFoundError(f"listing_inputs.json not found in {finished_folder_path}")

        ignored_at = format_workflow_timestamp()
        payload = dict(listing_memory)
        profile = find_profile_for_listing_memory(profiles, payload) or fallback_profile

        ignore_record = {
            "folder_name": finished_folder_name,
            "finished_folder_path": finished_folder_path,
            "ignored_at": ignored_at,
            "ignored_by": actor,
            "reason": reason,
        }
        ignored_generations = list(payload.get("ignored_generations", []))
        ignored_generations.append(ignore_record)
        payload["ignored_generations"] = ignored_generations[-50:]

        payload["generation_status"] = "ignored"
        payload["ignored_at"] = ignored_at
        payload["ignored_by"] = actor
        payload["ignored_reason"] = reason

        generated_outputs = list(payload.get("generated_outputs", []))
        if generated_outputs:
            latest_output = dict(generated_outputs[-1])
            latest_output["status"] = "ignored"
            latest_output["ignored_at"] = ignored_at
            latest_output["ignored_by"] = actor
            latest_output["ignored_reason"] = reason
            generated_outputs[-1] = latest_output
            payload["generated_outputs"] = generated_outputs

        append_workflow_event(
            payload,
            action="ignore_finished_generation",
            actor=actor,
            from_state="finished",
            to_state="finished",
            folder_path=finished_folder_path,
            details={
                "finished_folder_name": finished_folder_name,
                "reason": reason,
            },
        )

        save_listing_inputs_json_to_dropbox(
            profile=profile,
            payload=payload,
            folder_path=finished_folder_path,
        )

        result["status"] = "Success"
        result["message"] = "Marked generation as ignored."
    except Exception as exc:
        result["message"] = str(exc)

    return result


def format_workflow_timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def save_listing_memory_to_dropbox(
    profile: dict[str, Any],
    payload: dict[str, Any],
    folder_path: str,
) -> str:
    return save_listing_inputs_json_to_dropbox(profile, payload, folder_path)


def save_listing_inputs_json_to_dropbox(
    profile: dict[str, Any],
    payload: dict[str, Any],
    folder_path: str,
) -> str:
    json_path = build_listing_memory_path(folder_path)
    memory_payload = build_listing_memory_payload(profile, payload)
    upload_text_file(
        json_path,
        json.dumps(memory_payload, indent=2, ensure_ascii=False),
    )

    folder_cache_key = str(folder_path or "").rstrip("/")
    if folder_cache_key:
        st.session_state.setdefault("listing_memory_cache", {})[folder_cache_key] = {
            "json_path": json_path,
            "data": json.loads(json.dumps(memory_payload)),
            "missing": False,
        }

    return json_path


def create_staged_listing_task_in_dropbox(
    *,
    profile: dict[str, Any],
    payload: dict[str, Any],
    staged_folder_name: str,
    dropbox_cfg: dict[str, Any],
) -> dict[str, Any]:
    return create_staged_listing_task(
        profile=profile,
        payload=payload,
        staged_folder_name=staged_folder_name,
        stage_root=dropbox_cfg.get("stage_root", ""),
        destination_exists=path_exists_strict,
        create_folder=create_folder_exclusive,
        save_listing_memory=save_listing_inputs_json_to_dropbox,
    )


def load_grouped_christmas_image_manifest_from_dropbox(
    *,
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    profile: dict[str, Any],
) -> dict[str, Any]:
    folder_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)
    return build_christmas_group_image_manifest(
        list_folder_files(folder_path),
        profile,
    )


def save_grouped_christmas_draft_to_dropbox(
    *,
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    profile: dict[str, Any],
    payload: dict[str, Any],
) -> str:
    if not is_grouped_christmas_memory(payload):
        raise ValueError("Grouped Christmas draft memory is required.")
    folder_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)
    saved_source = load_listing_memory_from_dropbox_fresh(folder_path)
    if is_group_submission_locked(saved_source):
        raise ValueError(
            "Grouped publication has already begun. Use Resume grouped submission; "
            "the staged source is locked against draft changes."
        )
    return save_listing_inputs_json_to_dropbox(profile, payload, folder_path)


def load_listing_memory_from_dropbox_fresh(folder_path: str) -> dict[str, Any]:
    content = download_text_file(build_listing_memory_path(folder_path))
    data = json.loads(content)
    if not isinstance(data, dict):
        raise ValueError("listing_inputs.json must contain a JSON object.")
    return data


def resolve_active_staged_listing_memory(
    *,
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    load_stage_memory: Callable[[str], dict[str, Any]],
    destination_exists: Callable[[str], bool],
    load_fresh_memory: Callable[[str], dict[str, Any]],
) -> dict[str, Any]:
    stage_folder_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)
    stage_memory: dict[str, Any] = {}
    stage_load_error: Exception | None = None
    try:
        stage_memory = load_stage_memory(stage_folder_path)
    except Exception as exc:
        stage_load_error = exc
    if stage_memory:
        return {"memory": stage_memory, "location": "stage", "error": ""}

    try:
        stage_exists = destination_exists(stage_folder_path)
    except Exception as exc:
        return {
            "memory": {},
            "location": "stage",
            "error": f"The selected staged folder could not be verified: {exc}",
        }
    if stage_exists:
        if stage_load_error is not None:
            return {
                "memory": {},
                "location": "stage",
                "error": f"Saved staged listing state could not be loaded: {stage_load_error}",
            }
        return {"memory": stage_memory, "location": "stage", "error": ""}

    archive_root = str(dropbox_cfg.get("grouped_archive_root", "") or "").rstrip("/")
    archive_path = f"{archive_root}/{staged_folder_name}" if archive_root else ""
    if archive_path:
        try:
            archive_exists = destination_exists(archive_path)
        except Exception as exc:
            return {
                "memory": {},
                "location": "missing",
                "error": f"The selected staged folder archive state could not be verified: {exc}",
            }
        if archive_exists:
            try:
                archive_memory = load_fresh_memory(archive_path)
            except Exception as exc:
                return {
                    "memory": {},
                    "location": "archive",
                    "error": f"Grouped Christmas archived state could not be loaded: {exc}",
                }
            if is_grouped_christmas_memory(archive_memory):
                return {"memory": archive_memory, "location": "archive", "error": ""}
            return {
                "memory": archive_memory,
                "location": "archive",
                "error": (
                    "Grouped Christmas state could not be loaded. This staged task cannot "
                    "be edited or submitted until its saved grouped state is recovered."
                ),
            }

    return {
        "memory": {},
        "location": "missing",
        "error": (
            "The selected staged folder no longer exists. Refresh Product setup before "
            "editing or submitting this listing."
        ),
    }


def submit_grouped_christmas_to_review(
    *,
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    profile: dict[str, Any],
    draft_payload: dict[str, Any],
    profiles: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    target_profiles = (
        build_christmas_group_target_profiles(profiles)
        if profiles is not None
        else {}
    )
    publication_profile = dict(profile)
    if target_profiles:
        publication_profile["_group_target_profiles"] = target_profiles

    def profile_for_payload(payload: dict[str, Any]) -> dict[str, Any]:
        template_key = str(payload.get("template_key", "") or "").strip()
        for target_profile in target_profiles.values():
            if str(target_profile.get("template_key", "") or "").strip() == template_key:
                return target_profile
        return profile

    source_folder_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)
    if path_exists_strict(source_folder_path):
        saved_source = load_listing_memory_from_dropbox_fresh(source_folder_path)
        if not is_group_submission_locked(saved_source):
            save_grouped_christmas_draft_to_dropbox(
                dropbox_cfg=dropbox_cfg,
                staged_folder_name=staged_folder_name,
                profile=profile,
                payload=draft_payload,
            )

    timestamp = format_workflow_timestamp()
    actor = str(
        draft_payload.get("content_prepared_by", "")
        or draft_payload.get("assets_prepared_by", "")
        or ""
    )

    def save_memory(payload: dict[str, Any], folder_path: str) -> str:
        return save_listing_inputs_json_to_dropbox(
            profile_for_payload(payload),
            payload,
            folder_path,
        )

    def prepare_child_payload(
        payload: dict[str, Any],
        grouped_source_path: str,
        ready_path: str,
    ) -> dict[str, Any]:
        prepared = dict(payload)
        child_profile = profile_for_payload(prepared)
        prepared["prepared_at"] = timestamp
        prepared["review_snapshot"] = build_review_snapshot(
            profile=child_profile,
            payload=prepared,
            dropbox_cfg=dropbox_cfg,
            folder_path=ready_path,
        )
        append_workflow_event(
            prepared,
            action="submit_for_review",
            actor=actor,
            from_state="_stage",
            to_state="ready",
            folder_path=grouped_source_path,
            details={
                "assets_prepared_by": prepared.get("assets_prepared_by", ""),
                "content_prepared_by": prepared.get("content_prepared_by", ""),
                "grouped_member": prepared.get("source_group", {}).get("member_key", ""),
            },
        )
        return prepared

    storage = ChristmasGroupPublicationStorage(
        path_exists=path_exists_strict,
        ensure_folder=create_folder_if_missing,
        create_folder=create_folder_exclusive,
        load_memory=load_listing_memory_from_dropbox_fresh,
        save_memory=save_memory,
        list_files=list_folder_files,
        copy_file=copy_dropbox_file,
        move_folder=move_dropbox_folder,
    )
    result = publish_christmas_group(
        publication_profile,
        source_folder_name=staged_folder_name,
        source_folder_path=source_folder_path,
        preparation_root=dropbox_cfg.get("grouped_preparation_root", ""),
        ready_root=dropbox_cfg.get("ready_root", ""),
        archive_root=dropbox_cfg.get("grouped_archive_root", ""),
        storage=storage,
        prepare_child_payload=prepare_child_payload,
    )
    if (
        result.get("success")
        and result.get("state") == "released"
        and result.get("archive_path")
    ):
        clear_grouped_publication_active_context(
            staged_folder_name=staged_folder_name,
            source_folder_path=source_folder_path,
            archive_path=str(result.get("archive_path", "") or ""),
        )
        set_workflow_flash(
            "success",
            "Submitted 3 Christmas listings for review",
            "Prepared, verified, released, and archived the grouped source safely.",
        )
    return result


def clear_grouped_publication_active_context(
    *,
    staged_folder_name: str,
    source_folder_path: str,
    archive_path: str,
) -> None:
    refresh_cached_folder_names("stage")
    clear_cached_listing_memory(source_folder_path, archive_path)
    clear_runtime_caches()
    clear_grouped_christmas_session_state(st.session_state)

    st.session_state["staged_folder_select"] = None
    st.session_state["active_staged_folder_select"] = ""
    st.session_state["last_loaded_listing_memory_folder"] = ""
    for key in [
        "pending_staged_folder_selection_on_rerun",
        "clear_staged_folder_selection_on_rerun",
        "last_detected_template_folder",
        "applied_listing_memory_key_v2",
        "applied_listing_memory_widget_key_v2",
        "initialized_listing_context_key",
        "last_loaded_listing_memory_signature",
        "image_mappings_loaded_folder",
        "image_mappings_loaded_context",
    ]:
        st.session_state.pop(key, None)

    known_grouped_sources = {
        str(folder_name)
        for folder_name in st.session_state.get("known_grouped_source_folders", [])
        if str(folder_name).casefold() != str(staged_folder_name).casefold()
    }
    if known_grouped_sources:
        st.session_state["known_grouped_source_folders"] = sorted(known_grouped_sources)
    else:
        st.session_state.pop("known_grouped_source_folders", None)


def load_grouped_christmas_test_content() -> str:
    return (BASE_DIR / "samples" / "christmas_grouped_listing_content_test.json").read_text(
        encoding="utf-8"
    )


def load_listing_memory_from_dropbox(folder_path: str) -> dict[str, Any]:
    folder_path = str(folder_path or "").rstrip("/")
    if not folder_path:
        return {}

    json_path = build_listing_memory_path(folder_path)
    cache = st.session_state.setdefault("listing_memory_cache", {})
    cached = cache.get(folder_path)

    if cached and cached.get("json_path") == json_path and not cached.get("missing"):
        started_at = time.perf_counter()
        cached_data = json.loads(json.dumps(cached.get("data", {})))
        record_load_event(
            "Dropbox: cached listing_inputs.json",
            started_at,
            format_folder_detail(folder_path),
        )
        return cached_data

    started_at = time.perf_counter()

    if not path_exists(json_path):
        cache[folder_path] = {
            "json_path": json_path,
            "data": {},
            "missing": True,
        }
        record_load_event(
            "Dropbox: missing listing_inputs.json",
            started_at,
            format_folder_detail(folder_path),
        )
        return {}

    content = download_text_file(json_path)
    data = json.loads(content)

    cache[folder_path] = {
        "json_path": json_path,
        "data": json.loads(json.dumps(data)),
        "missing": False,
    }

    record_load_event(
        "Dropbox: load listing_inputs.json",
        started_at,
        format_folder_detail(folder_path),
    )

    return data

def initialize_listing_context_defaults(profile: dict[str, Any]) -> None:
    normalize_selected_variants_session_state(profile, {}, force_defaults=True)
    st.session_state["parent_main_image_choice"] = "Automatic (recommended)"
    st.session_state["variant_quantity"] = 100
    st.session_state["handling_time_days"] = DEFAULT_HANDLING_TIME_DAYS
    st.session_state["merchant_shipping_group_name"] = DEFAULT_MERCHANT_SHIPPING_GROUP
    sku_decoration_code = get_default_sku_decoration_code(profile)
    st.session_state["sku_decoration_choice"] = sku_decoration_code if sku_decoration_code in SKU_DECORATION_OPTIONS else "Custom"
    st.session_state["custom_sku_decoration_code"] = "" if sku_decoration_code in SKU_DECORATION_OPTIONS else sku_decoration_code
    st.session_state["manual_sku_listing_code"] = ""
    st.session_state["generated_sku_listing_code"] = f"D{generate_unique_sku(5)}"


CONTENT_EDITOR_KEYS = {
    "title": "content_title_input_v3",
    "description": "content_product_description_v3",
    "keywords": "content_generic_keywords_v3",
    "bullets": [
        "content_bullet_1_v3",
        "content_bullet_2_v3",
        "content_bullet_3_v3",
        "content_bullet_4_v3",
        "content_bullet_5_v3",
    ],
}


def hydrate_content_editor_widget_state(listing_memory: dict[str, Any]) -> None:
    bullet_points = listing_memory.get("bullet_points", [])
    bullet_points = (bullet_points + ["", "", "", "", ""])[:5]

    st.session_state[CONTENT_EDITOR_KEYS["title"]] = str(listing_memory.get("title", "") or "")
    st.session_state[CONTENT_EDITOR_KEYS["description"]] = str(listing_memory.get("product_description", "") or "")
    st.session_state[CONTENT_EDITOR_KEYS["keywords"]] = str(listing_memory.get("generic_keywords", "") or "")
    for idx, value in enumerate(bullet_points):
        st.session_state[CONTENT_EDITOR_KEYS["bullets"][idx]] = str(value or "")


def listing_memory_has_content(listing_memory: dict[str, Any]) -> bool:
    if not listing_memory:
        return False

    content_values = [
        listing_memory.get("title", ""),
        listing_memory.get("product_description", ""),
        listing_memory.get("generic_keywords", ""),
        *listing_memory.get("bullet_points", []),
    ]
    return any(str(value or "").strip() for value in content_values)


def listing_content_widget_state_is_missing(
    session_state: dict[str, Any],
    listing_memory: dict[str, Any],
) -> bool:
    widget_keys = [
        "handling_time_days",
        "merchant_shipping_group_name",
        "sku_decoration_choice",
        "manual_sku_listing_code",
        "variant_quantity",
        "content_prepared_by",
    ]
    if not is_grouped_christmas_memory(listing_memory):
        widget_keys.extend([
            CONTENT_EDITOR_KEYS["title"],
            CONTENT_EDITOR_KEYS["description"],
            CONTENT_EDITOR_KEYS["keywords"],
            *CONTENT_EDITOR_KEYS["bullets"],
        ])
    return any(key not in session_state for key in widget_keys)


def selectbox_index_without_state_conflict(
    key: str,
    options: list[Any],
    fallback_value: Any = None,
) -> int | None:
    """Return an index only when Streamlit does not already own this widget key."""
    if key in st.session_state:
        if st.session_state.get(key) in options:
            return None
        st.session_state.pop(key, None)

    if fallback_value in options:
        return options.index(fallback_value)

    return None


def sync_content_editor_to_canonical_state(
    title: str,
    bullets: list[str],
    product_description: str,
    generic_keywords: str,
) -> None:
    st.session_state["title_input"] = title
    for idx, value in enumerate((bullets + ["", "", "", "", ""])[:5], start=1):
        st.session_state[f"bullet_{idx}"] = value
    st.session_state["product_description"] = product_description
    st.session_state["generic_keywords"] = generic_keywords


def apply_listing_memory_to_session(listing_memory: dict[str, Any], profile: dict[str, Any]) -> None:
    st.session_state["title_input"] = listing_memory.get("title", "")
    for field_name in ["assets_prepared_by", "content_prepared_by", "reviewed_by"]:
        if field_name not in st.session_state:
            st.session_state[field_name] = listing_memory.get(field_name, "")
    st.session_state["prepared_at"] = listing_memory.get("prepared_at", "")
    st.session_state["reviewed_at"] = listing_memory.get("reviewed_at", "")

    bullet_points = listing_memory.get("bullet_points", [])
    bullet_points = (bullet_points + ["", "", "", "", ""])[:5]
    for idx, value in enumerate(bullet_points, start=1):
        st.session_state[f"bullet_{idx}"] = value

    st.session_state["product_description"] = listing_memory.get("product_description", "")
    st.session_state["generic_keywords"] = listing_memory.get("generic_keywords", "")
    hydrate_content_editor_widget_state(listing_memory)
    st.session_state["use_same_price_for_all_sizes"] = listing_memory.get("use_same_price_for_all_sizes", False)
    sku_decoration_code = get_default_sku_decoration_code(profile, listing_memory)
    st.session_state["sku_decoration_choice"] = sku_decoration_code if sku_decoration_code in SKU_DECORATION_OPTIONS else "Custom"
    st.session_state["custom_sku_decoration_code"] = "" if sku_decoration_code in SKU_DECORATION_OPTIONS else sku_decoration_code
    st.session_state["manual_sku_listing_code"] = str(listing_memory.get("manual_sku_listing_code", "") or "")
    st.session_state["generated_sku_listing_code"] = (
        get_saved_generated_sku_listing_code(listing_memory)
        or f"D{generate_unique_sku(5)}"
    )
    st.session_state["parent_main_image_choice"] = listing_memory.get(
        "parent_main_image_choice",
        "Automatic (recommended)",
    ) or "Automatic (recommended)"
    st.session_state["variant_quantity"] = normalize_variant_quantity(
        listing_memory.get("quantity", DEFAULT_VARIANT_QUANTITY)
    )
    st.session_state["handling_time_days"] = normalize_handling_time_days(
        listing_memory.get("handling_time_days", DEFAULT_HANDLING_TIME_DAYS)
    )
    st.session_state["merchant_shipping_group_name"] = normalize_merchant_shipping_group(
        listing_memory.get("merchant_shipping_group_name", "")
    )

    saved_prices = normalize_saved_price_map_for_profile(
        profile,
        listing_memory.get("size_price_map", {}),
    )
    for size, price in saved_prices.items():
        st.session_state[f"price_{size}"] = float(price)

    normalize_selected_variants_session_state(profile, listing_memory, force_saved_values=True)

def finalize_staged_dropbox_folder(
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    parent_sku: str,
    reuse_finished_folder_name: str = "",
) -> tuple[str, str]:
    parent_sku = sanitize_sku(parent_sku)
    if not parent_sku:
        raise ValueError("Template parent_sku is missing.")

    finished_root = dropbox_cfg.get("finished_root", "").rstrip("/")
    stage_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)

    create_folder_if_missing(finished_root)

    max_attempts = 20
    final_sku = ""
    final_folder_path = ""

    reuse_finished_folder_name = sanitize_sku(reuse_finished_folder_name)
    if reuse_finished_folder_name:
        final_sku = reuse_finished_folder_name
        final_folder_path = build_finished_folder_path(dropbox_cfg, final_sku)
        moved_path = move_dropbox_folder(stage_path, final_folder_path)
        return final_sku, moved_path

    for _ in range(max_attempts):
        unique_sku = generate_unique_sku()
        final_sku = build_final_folder_sku(parent_sku, unique_sku)
        final_folder_path = build_finished_folder_path(dropbox_cfg, final_sku)

        if not path_exists(final_folder_path):
            moved_path = move_dropbox_folder(stage_path, final_folder_path)
            return final_sku, moved_path

    raise ValueError("Could not generate a unique finished folder SKU after multiple attempts.")


def move_staged_dropbox_folder_to_ready(
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    ready_folder_name: str,
) -> str:
    ready_folder_name = sanitize_sku(ready_folder_name)
    if not ready_folder_name:
        raise ValueError("Ready folder name is required.")

    ready_root = dropbox_cfg.get("ready_root", "").rstrip("/")
    actual_staged_folder_name = resolve_existing_stage_folder_name(dropbox_cfg, staged_folder_name)
    stage_path = build_stage_folder_path(dropbox_cfg, actual_staged_folder_name)

    create_folder_if_missing(ready_root)

    final_ready_folder_name = ready_folder_name
    counter = 1

    while True:
        ready_folder_path = build_ready_folder_path(dropbox_cfg, final_ready_folder_name)
        if path_exists(ready_folder_path):
            final_ready_folder_name = f"{ready_folder_name}-{counter}"
            counter += 1
            continue

        moved_path = move_dropbox_folder(stage_path, ready_folder_path)
        if not moved_path:
            raise RuntimeError("Dropbox returned an empty path after moving the folder to ready.")
        return moved_path


def move_ready_dropbox_folder_to_approved(
    dropbox_cfg: dict[str, Any],
    ready_folder_name: str,
    approved_folder_name: str,
) -> str:
    approved_folder_name = sanitize_sku(approved_folder_name)
    if not approved_folder_name:
        raise ValueError("Approved folder name is required.")

    approved_root = dropbox_cfg.get("approved_root", "").rstrip("/")
    ready_path = build_ready_folder_path(dropbox_cfg, ready_folder_name)

    create_folder_if_missing(approved_root)

    final_approved_folder_name = approved_folder_name
    counter = 1

    while True:
        approved_folder_path = build_approved_folder_path(dropbox_cfg, final_approved_folder_name)
        if path_exists(approved_folder_path):
            final_approved_folder_name = f"{approved_folder_name}-{counter}"
            counter += 1
            continue

        moved_path = move_dropbox_folder(ready_path, approved_folder_path)
        if not moved_path:
            raise RuntimeError("Dropbox returned an empty path after moving the folder to approved.")
        return moved_path


def move_ready_dropbox_folder_to_denied_stage(
    dropbox_cfg: dict[str, Any],
    ready_folder_name: str,
) -> str:
    denied_folder_name = sanitize_sku(f"{ready_folder_name}_denied")
    if not denied_folder_name:
        raise ValueError("Denied folder name is required.")

    stage_root = dropbox_cfg.get("stage_root", "").rstrip("/")
    ready_path = build_ready_folder_path(dropbox_cfg, ready_folder_name)

    create_folder_if_missing(stage_root)

    final_denied_folder_name = denied_folder_name
    counter = 1

    while True:
        denied_stage_folder_path = build_stage_folder_path(dropbox_cfg, final_denied_folder_name)
        if path_exists(denied_stage_folder_path):
            final_denied_folder_name = f"{denied_folder_name}-{counter}"
            counter += 1
            continue

        moved_path = move_dropbox_folder(ready_path, denied_stage_folder_path)
        if not moved_path:
            raise RuntimeError("Dropbox returned an empty path after moving the folder back to staging.")
        return moved_path


def move_approved_dropbox_folder_to_ready(
    dropbox_cfg: dict[str, Any],
    approved_folder_name: str,
) -> str:
    approved_folder_name = sanitize_sku(approved_folder_name)
    if not approved_folder_name:
        raise ValueError("Approved folder name is required.")

    ready_root = dropbox_cfg.get("ready_root", "").rstrip("/")
    approved_path = build_approved_folder_path(dropbox_cfg, approved_folder_name)

    create_folder_if_missing(ready_root)

    final_ready_folder_name = approved_folder_name
    counter = 1

    while True:
        ready_folder_path = build_ready_folder_path(dropbox_cfg, final_ready_folder_name)
        if path_exists(ready_folder_path):
            final_ready_folder_name = f"{approved_folder_name}-{counter}"
            counter += 1
            continue

        moved_path = move_dropbox_folder(approved_path, ready_folder_path)
        if not moved_path:
            raise RuntimeError("Dropbox returned an empty path after moving the folder back to review.")
        return moved_path


def move_approved_dropbox_folder_to_stage(
    dropbox_cfg: dict[str, Any],
    approved_folder_name: str,
) -> str:
    approved_folder_name = sanitize_sku(approved_folder_name)
    if not approved_folder_name:
        raise ValueError("Approved folder name is required.")

    stage_root = dropbox_cfg.get("stage_root", "").rstrip("/")
    approved_path = build_approved_folder_path(dropbox_cfg, approved_folder_name)
    stage_folder_name = sanitize_sku(f"{approved_folder_name}_rejected")

    create_folder_if_missing(stage_root)

    final_stage_folder_name = stage_folder_name
    counter = 1

    while True:
        stage_folder_path = build_stage_folder_path(dropbox_cfg, final_stage_folder_name)
        if path_exists(stage_folder_path):
            final_stage_folder_name = f"{stage_folder_name}-{counter}"
            counter += 1
            continue

        moved_path = move_dropbox_folder(approved_path, stage_folder_path)
        if not moved_path:
            raise RuntimeError("Dropbox returned an empty path after moving the folder back to staging.")
        return moved_path


def return_approved_listing(
    dropbox_cfg: dict[str, Any],
    profiles: list[dict[str, Any]],
    fallback_profile: dict[str, Any],
    approved_folder_name: str,
    target_state: str,
    actor: str = "",
    reason: str = "",
) -> dict[str, Any]:
    approved_folder_name = sanitize_sku(approved_folder_name)
    target_state = str(target_state or "").strip().lower()
    if target_state not in {"ready", "stage"}:
        raise ValueError("target_state must be either 'ready' or 'stage'.")

    source_path = build_approved_folder_path(dropbox_cfg, approved_folder_name)
    result = {
        "folder_name": approved_folder_name,
        "from_state": "approved",
        "to_state": target_state,
        "target_folder_name": "",
        "target_path": "",
        "status": "Failed",
        "message": "",
    }

    try:
        if not approved_folder_name:
            raise ValueError("Approved folder name is required.")
        if not path_exists(source_path):
            raise FileNotFoundError(f"Approved folder not found: {source_path}")

        listing_memory = load_listing_memory_from_dropbox(source_path)
        if not listing_memory:
            listing_memory = {}

        moved_path = (
            move_approved_dropbox_folder_to_ready(dropbox_cfg, approved_folder_name)
            if target_state == "ready"
            else move_approved_dropbox_folder_to_stage(dropbox_cfg, approved_folder_name)
        )

        payload = dict(listing_memory)
        payload["reviewed_by"] = ""
        payload["reviewed_at"] = ""
        if target_state == "stage":
            payload["prepared_at"] = ""

        append_workflow_event(
            payload,
            action="return_approved_to_review" if target_state == "ready" else "return_approved_to_stage",
            actor=actor,
            from_state="approved",
            to_state=target_state,
            folder_path=moved_path,
            details={
                "old_approved_folder_name": approved_folder_name,
                "old_approved_folder_path": source_path,
                "reason": reason,
            },
        )

        profile = find_profile_for_listing_memory(profiles, payload) or fallback_profile
        save_listing_inputs_json_to_dropbox(
            profile=profile,
            payload=payload,
            folder_path=moved_path,
        )

        result.update(
            {
                "target_folder_name": Path(moved_path).name,
                "target_path": moved_path,
                "status": "Success",
                "message": (
                    "Returned to review."
                    if target_state == "ready"
                    else "Returned to staging."
                ),
            }
        )
    except Exception as exc:
        result["message"] = str(exc)

    return result


def finalize_ready_dropbox_folder(
    dropbox_cfg: dict[str, Any],
    ready_folder_name: str,
    parent_sku: str,
    reuse_finished_folder_name: str = "",
) -> tuple[str, str]:
    parent_sku = sanitize_sku(parent_sku)
    if not parent_sku:
        raise ValueError("Template parent_sku is missing.")

    finished_root = dropbox_cfg.get("finished_root", "").rstrip("/")
    ready_path = build_ready_folder_path(dropbox_cfg, ready_folder_name)

    create_folder_if_missing(finished_root)

    max_attempts = 20

    reuse_finished_folder_name = sanitize_sku(reuse_finished_folder_name)
    if reuse_finished_folder_name:
        final_sku = reuse_finished_folder_name
        final_folder_path = build_finished_folder_path(dropbox_cfg, final_sku)
        moved_path = move_dropbox_folder(ready_path, final_folder_path)
        return final_sku, moved_path

    for _ in range(max_attempts):
        unique_sku = generate_unique_sku()
        final_sku = build_final_folder_sku(parent_sku, unique_sku)
        final_folder_path = build_finished_folder_path(dropbox_cfg, final_sku)

        if not path_exists(final_folder_path):
            moved_path = move_dropbox_folder(ready_path, final_folder_path)
            return final_sku, moved_path

    raise ValueError("Could not generate a unique finished folder SKU after multiple attempts.")

def finalize_approved_dropbox_folder(
    dropbox_cfg: dict[str, Any],
    approved_folder_name: str,
    parent_sku: str,
    reuse_finished_folder_name: str = "",
) -> tuple[str, str]:
    parent_sku = sanitize_sku(parent_sku)
    if not parent_sku:
        raise ValueError("Template parent_sku is missing.")

    finished_root = dropbox_cfg.get("finished_root", "").rstrip("/")
    approved_path = build_approved_folder_path(dropbox_cfg, approved_folder_name)

    create_folder_if_missing(finished_root)

    max_attempts = 20

    reuse_finished_folder_name = sanitize_sku(reuse_finished_folder_name)
    if reuse_finished_folder_name:
        final_sku = reuse_finished_folder_name
        final_folder_path = build_finished_folder_path(dropbox_cfg, final_sku)
        moved_path = move_dropbox_folder(approved_path, final_folder_path)
        return final_sku, moved_path

    for _ in range(max_attempts):
        unique_sku = generate_unique_sku()
        final_sku = build_final_folder_sku(parent_sku, unique_sku)
        final_folder_path = build_finished_folder_path(dropbox_cfg, final_sku)

        if not path_exists(final_folder_path):
            moved_path = move_dropbox_folder(approved_path, final_folder_path)
            return final_sku, moved_path

    raise ValueError("Could not generate a unique finished folder SKU after multiple attempts.")


def choose_finished_folder_target(
    dropbox_cfg: dict[str, Any],
    parent_sku: str,
    reuse_finished_folder_name: str = "",
) -> tuple[str, str]:
    parent_sku = sanitize_sku(parent_sku)
    if not parent_sku:
        raise ValueError("Template parent_sku is missing.")

    finished_root = dropbox_cfg.get("finished_root", "").rstrip("/")
    create_folder_if_missing(finished_root)

    reuse_finished_folder_name = sanitize_sku(reuse_finished_folder_name)
    if reuse_finished_folder_name:
        final_sku = reuse_finished_folder_name
        return final_sku, build_finished_folder_path(dropbox_cfg, final_sku)

    for _ in range(20):
        unique_sku = generate_unique_sku()
        final_sku = build_final_folder_sku(parent_sku, unique_sku)
        final_folder_path = build_finished_folder_path(dropbox_cfg, final_sku)
        if not path_exists(final_folder_path):
            return final_sku, final_folder_path

    raise ValueError("Could not generate a unique finished folder SKU after multiple attempts.")


def split_folder_images(folder_path: str) -> tuple[str, list[str]]:
    files = [p for p in list_folder_files(folder_path) if is_image_file(p)]
    files = sorted(files, key=lambda p: Path(p).name.lower())

    parent_main = ""
    other_images: list[str] = []

    for path in files:
        stem = Path(path).stem.lower()

        if stem == "main":
            parent_main = path
        else:
            other_images.append(path)

    if not parent_main and other_images:
        parent_main = other_images[0]
        other_images = other_images[1:]

    return parent_main, other_images

def build_stage_preview_paths(dropbox_cfg: dict[str, Any], staged_folder_name: str) -> list[str]:
    if not staged_folder_name:
        return []

    stage_root = dropbox_cfg.get("stage_root", "").rstrip("/")
    stage_folder_path = f"{stage_root}/{staged_folder_name}"

    try:
        files = [p for p in list_folder_files(stage_folder_path) if is_image_file(p)]
    except Exception:
        return []
    return sorted(files, key=lambda p: Path(p).name.lower())


def build_stage_resource_folder_path(dropbox_cfg: dict[str, Any], staged_folder_name: str) -> str:
    if not staged_folder_name:
        return ""

    stage_root = dropbox_cfg.get("stage_root", "").rstrip("/")
    return f"{stage_root}/{staged_folder_name}/resources" if stage_root else ""


def build_stage_resource_image_paths(dropbox_cfg: dict[str, Any], staged_folder_name: str) -> list[str]:
    resource_folder_path = build_stage_resource_folder_path(dropbox_cfg, staged_folder_name)
    return list_image_paths_in_dropbox_folder(resource_folder_path)


def list_image_paths_in_dropbox_folder(folder_path: str) -> list[str]:
    if not folder_path:
        return []

    try:
        files = [p for p in list_folder_files(folder_path) if is_image_file(p)]
    except Exception:
        return []

    return sorted(files, key=lambda p: Path(p).name.lower())


def padded_list(values: list[str], target_len: int = 8) -> list[str]:
    trimmed = values[:target_len]
    return trimmed + [""] * (target_len - len(trimmed))

def is_image_file(path: str) -> bool:
    suffix = Path(path).suffix.lower()
    return suffix in {".png", ".jpg", ".jpeg", ".webp"}


REVIEW_RESOURCE_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg"}


def validate_review_resource_image(filename: str, content: bytes) -> str:
    safe_filename = Path(str(filename or "").replace("\\", "/")).name.strip()
    suffix = Path(safe_filename).suffix.lower()
    if not safe_filename or safe_filename in {".", ".."}:
        raise ValueError("Choose a PNG or JPG resource image.")
    if suffix not in REVIEW_RESOURCE_IMAGE_EXTENSIONS:
        raise ValueError("Resource images must be PNG or JPG files.")
    if not content:
        raise ValueError("The selected resource image is empty.")
    if suffix == ".png" and not content.startswith(b"\x89PNG\r\n\x1a\n"):
        raise ValueError("The selected file is not a valid PNG image.")
    if suffix in {".jpg", ".jpeg"} and not content.startswith(b"\xff\xd8\xff"):
        raise ValueError("The selected file is not a valid JPG image.")
    return safe_filename


def upload_ready_review_resource_image(
    source_folder_path: str,
    filename: str,
    content: bytes,
) -> str:
    ready_folder_path = str(source_folder_path or "").rstrip("/")
    if not ready_folder_path:
        raise ValueError("The ready listing folder is unavailable.")

    resources_folder_path = f"{ready_folder_path}/resources"
    return upload_resource_images_to_folder(
        resources_folder_path,
        [(filename, content)],
    )[0]


def upload_resource_images_to_folder(
    folder_path: str,
    images: list[tuple[str, bytes]],
) -> list[str]:
    target_folder_path = str(folder_path or "").rstrip("/")
    if not target_folder_path:
        raise ValueError("The resource folder is unavailable.")
    if not images:
        raise ValueError("Choose at least one PNG or JPG resource image.")

    validated_images: list[tuple[str, bytes]] = []
    seen_filenames: set[str] = set()
    for filename, content in images:
        safe_filename = validate_review_resource_image(filename, content)
        filename_key = safe_filename.casefold()
        if filename_key in seen_filenames:
            raise ValueError(f"Duplicate resource image filename: {safe_filename}")
        seen_filenames.add(filename_key)
        validated_images.append((safe_filename, content))

    create_folder_if_missing(target_folder_path)
    return [
        upload_binary_file(f"{target_folder_path}/{safe_filename}", content)
        for safe_filename, content in validated_images
    ]


def is_ignored_zip_member(parts: list[str]) -> bool:
    if not parts:
        return True
    filename = parts[-1]
    return (
        any(part == "__MACOSX" for part in parts)
        or filename in {".DS_Store", "Thumbs.db"}
        or filename.startswith("._")
    )


def inspect_stage_images_zip(zip_bytes: bytes, staged_folder_path: str) -> dict[str, Any]:
    stage_root = staged_folder_path.rstrip("/")
    resources_root = f"{stage_root}/resources"
    plan: dict[str, Any] = {
        "mockups": [],
        "resources": [],
        "skipped": [],
    }

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as archive:
        member_parts: list[list[str]] = []
        for member in archive.infolist():
            if member.is_dir():
                continue
            normalized_name = member.filename.replace("\\", "/").lstrip("/")
            parts = [part for part in normalized_name.split("/") if part and part != "."]
            if parts and not is_ignored_zip_member(parts):
                member_parts.append(parts)

        wrapper_folder = ""
        if member_parts and all(len(parts) > 1 for parts in member_parts):
            first_parts = {parts[0] for parts in member_parts}
            if len(first_parts) == 1 and next(iter(first_parts)).lower() != "resources":
                wrapper_folder = next(iter(first_parts))

        for member in archive.infolist():
            raw_name = member.filename
            normalized_name = raw_name.replace("\\", "/").lstrip("/")
            parts = [part for part in normalized_name.split("/") if part and part != "."]

            if member.is_dir() or is_ignored_zip_member(parts):
                continue
            if wrapper_folder and parts and parts[0] == wrapper_folder:
                parts = parts[1:]

            if not parts or any(part == ".." for part in parts) or (parts and ":" in parts[0]):
                plan["skipped"].append({
                    "file": raw_name,
                    "reason": "Unsafe ZIP path",
                })
                continue

            filename = parts[-1]
            if not is_image_file(filename):
                plan["skipped"].append({
                    "file": raw_name,
                    "reason": "Not a supported image file",
                })
                continue

            if parts[0].lower() == "resources":
                if len(parts) < 2:
                    plan["skipped"].append({
                        "file": raw_name,
                        "reason": "Resource filename missing",
                    })
                    continue
                plan["resources"].append({
                    "source": raw_name,
                    "filename": filename,
                    "destination": f"{resources_root}/{filename}",
                })
                continue

            if len(parts) > 1:
                plan["skipped"].append({
                    "file": raw_name,
                    "reason": "Nested image is not inside resources/",
                })
                continue

            plan["mockups"].append({
                "source": raw_name,
                "filename": filename,
                "destination": f"{stage_root}/{filename}",
            })

    return plan


def count_stage_image_files(folder_path: str) -> int:
    if not folder_path:
        return 0
    try:
        return len([path for path in list_folder_files(folder_path) if is_image_file(path)])
    except Exception:
        return 0


def upload_stage_images_zip(
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    zip_bytes: bytes,
) -> dict[str, Any]:
    staged_folder_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)
    resources_folder_path = f"{staged_folder_path.rstrip('/')}/resources"
    plan = inspect_stage_images_zip(zip_bytes, staged_folder_path)

    before_mockups = count_stage_image_files(staged_folder_path)
    before_resources = count_stage_image_files(resources_folder_path)

    create_folder_if_missing(staged_folder_path)
    create_folder_if_missing(resources_folder_path)

    uploaded: list[dict[str, str]] = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as archive:
        for item in [*plan["mockups"], *plan["resources"]]:
            upload_binary_file(item["destination"], archive.read(item["source"]))
            uploaded.append({
                "source": item["source"],
                "destination": item["destination"],
            })

    after_mockups = count_stage_image_files(staged_folder_path)
    after_resources = count_stage_image_files(resources_folder_path)

    return {
        "staged_folder_path": staged_folder_path,
        "resources_folder_path": resources_folder_path,
        "planned_mockups": len(plan["mockups"]),
        "planned_resources": len(plan["resources"]),
        "skipped": plan["skipped"],
        "uploaded": uploaded,
        "before_mockups": before_mockups,
        "before_resources": before_resources,
        "after_mockups": after_mockups,
        "after_resources": after_resources,
    }


def render_stage_images_zip_upload(
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
) -> None:
    zip_upload_active = st.session_state.get("stage_images_zip_upload") is not None
    with st.expander("Upload staged images from ZIP", expanded=zip_upload_active):
        uploaded_zip = st.file_uploader(
            "Upload ZIP with mockup images in the root and resource images inside resources/",
            type=["zip"],
            key="stage_images_zip_upload",
            help="Root images become staged variant/mockup images. Files inside resources/ become secondary resource images.",
        )

        if not staged_folder_name:
            st.info(
                "Select an existing staged folder, or upload a staged-folder ZIP here. "
                "If no folder is selected, the ZIP filename is used as the new staged folder name."
            )
            zip_folder_guess = ""
            if uploaded_zip:
                zip_folder_guess = sanitize_stage_folder_name(Path(uploaded_zip.name).stem)
                if st.session_state.get("zip_new_staged_folder_source") != uploaded_zip.name:
                    st.session_state["zip_new_staged_folder_name"] = zip_folder_guess
                    st.session_state["zip_new_staged_folder_source"] = uploaded_zip.name
            new_folder_raw = st.text_input(
                "New staged folder name",
                key="zip_new_staged_folder_name",
                placeholder="Example: THMRS T01",
                help="Defaults to the ZIP filename. This creates a folder in Dropbox _stage, selects it, and can upload the ZIP in the same action.",
            )
            new_folder_name = sanitize_stage_folder_name(new_folder_raw)
            if new_folder_raw and new_folder_name != new_folder_raw.strip():
                st.caption(f"Folder will be created as `{new_folder_name}`.")

            if uploaded_zip:
                new_folder_path = build_stage_folder_path(dropbox_cfg, new_folder_name)
                zip_bytes = uploaded_zip.getvalue()
                try:
                    plan = inspect_stage_images_zip(zip_bytes, new_folder_path)
                except zipfile.BadZipFile:
                    st.error("That file is not a valid ZIP archive.")
                    return

                preview_cols = st.columns(3)
                preview_cols[0].metric("ZIP mockup images", len(plan["mockups"]))
                preview_cols[1].metric("ZIP resource images", len(plan["resources"]))
                preview_cols[2].metric("Skipped files", len(plan["skipped"]))

                if plan["mockups"]:
                    with st.expander("Mockup files to upload", expanded=False):
                        st.dataframe(
                            [{"filename": item["filename"], "destination": item["destination"]} for item in plan["mockups"]],
                            width="stretch",
                            hide_index=True,
                        )
                if plan["resources"]:
                    with st.expander("Resource files to upload", expanded=False):
                        st.dataframe(
                            [{"filename": item["filename"], "destination": item["destination"]} for item in plan["resources"]],
                            width="stretch",
                            hide_index=True,
                        )
                if plan["skipped"]:
                    with st.expander("Skipped ZIP files", expanded=False):
                        st.dataframe(plan["skipped"], width="stretch", hide_index=True)

            create_disabled = not bool(new_folder_name)
            if st.button(
                "Create staged folder and upload ZIP" if uploaded_zip else "Create staged folder",
                key="zip_create_staged_folder_btn",
                width="stretch",
                disabled=create_disabled,
            ):
                try:
                    if uploaded_zip:
                        result = upload_stage_images_zip(dropbox_cfg, new_folder_name, uploaded_zip.getvalue())
                        st.session_state["stage_images_zip_upload_result"] = result
                    else:
                        new_folder_path = build_stage_folder_path(dropbox_cfg, new_folder_name)
                        create_folder_if_missing(new_folder_path)
                except Exception as exc:
                    st.error(f"Could not create staged folder/upload ZIP: {exc}")
                    return

                clear_runtime_caches()
                clear_loaded_image_mapping_state(st.session_state)
                refresh_cached_folder_names("stage")
                st.session_state["active_folder_source_mode"] = "Use staged folder"
                st.session_state["active_staged_folder_select"] = new_folder_name
                st.session_state["pending_staged_folder_selection_on_rerun"] = new_folder_name
                set_workflow_flash(
                    "success",
                    f"Created staged folder {new_folder_name}.",
                    (
                        f"Uploaded {len(st.session_state.get('stage_images_zip_upload_result', {}).get('uploaded', []))} image file(s)."
                        if uploaded_zip
                        else "You can now upload a ZIP with mockups in the root and resources inside resources/."
                    ),
                )
                st.rerun()
            return

        staged_folder_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)
        resources_folder_path = f"{staged_folder_path.rstrip('/')}/resources"
        last_upload_result = st.session_state.get("stage_images_zip_upload_result", {})
        if last_upload_result.get("staged_folder_path") == staged_folder_path:
            existing_mockups = last_upload_result.get("after_mockups", 0)
            existing_resources = last_upload_result.get("after_resources", 0)
        else:
            existing_mockups = count_stage_image_files(staged_folder_path)
            existing_resources = count_stage_image_files(resources_folder_path)

        metric_cols = st.columns(4)
        metric_cols[0].metric("Current mockups", existing_mockups)
        metric_cols[1].metric("Current resources", existing_resources)
        metric_cols[2].write(f"Folder: `{staged_folder_name}`")
        metric_cols[3].write("Overwrites same filenames")

        if not uploaded_zip:
            return

        zip_bytes = uploaded_zip.getvalue()
        try:
            plan = inspect_stage_images_zip(zip_bytes, staged_folder_path)
        except zipfile.BadZipFile:
            st.error("That file is not a valid ZIP archive.")
            return

        preview_cols = st.columns(3)
        preview_cols[0].metric("ZIP mockup images", len(plan["mockups"]))
        preview_cols[1].metric("ZIP resource images", len(plan["resources"]))
        preview_cols[2].metric("Skipped files", len(plan["skipped"]))

        if plan["mockups"]:
            with st.expander("Mockup files to upload", expanded=False):
                st.dataframe(
                    [{"filename": item["filename"], "destination": item["destination"]} for item in plan["mockups"]],
                    width="stretch",
                    hide_index=True,
                )

        if plan["resources"]:
            with st.expander("Resource files to upload", expanded=False):
                st.dataframe(
                    [{"filename": item["filename"], "destination": item["destination"]} for item in plan["resources"]],
                    width="stretch",
                    hide_index=True,
                )

        if plan["skipped"]:
            with st.expander("Skipped ZIP files", expanded=False):
                st.dataframe(plan["skipped"], width="stretch", hide_index=True)

        can_upload = bool(plan["mockups"] or plan["resources"])
        if st.button(
            "Upload ZIP images to staged folder",
            key="upload_stage_images_zip_button",
            width="stretch",
            disabled=not can_upload,
        ):
            try:
                result = upload_stage_images_zip(dropbox_cfg, staged_folder_name, zip_bytes)
            except Exception as exc:
                st.error(f"ZIP upload failed: {exc}")
                return

            clear_loaded_image_mapping_state(st.session_state)
            st.session_state["stage_images_zip_upload_result"] = result
            set_workflow_flash(
                "success",
                f"Uploaded {len(result['uploaded'])} image file(s) to {staged_folder_name}.",
                (
                    f"Mockups: {result['before_mockups']} -> {result['after_mockups']}. "
                    f"Resources: {result['before_resources']} -> {result['after_resources']}."
                ),
            )
            st.rerun()

        result = st.session_state.get("stage_images_zip_upload_result")
        if result and result.get("staged_folder_path") == staged_folder_path:
            st.success(
                "Last ZIP upload: "
                f"{len(result.get('uploaded', []))} uploaded, "
                f"{len(result.get('skipped', []))} skipped. "
                f"Mockups {result.get('before_mockups', 0)} -> {result.get('after_mockups', 0)}, "
                f"resources {result.get('before_resources', 0)} -> {result.get('after_resources', 0)}."
            )


def build_design_color_preview_paths(
    profile: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    selected_variants: dict[str, list[str]],
    staged_folder_name: str,
) -> list[dict[str, str]]:
    template_key = profile.get("template_key", "")
    templates_map = dropbox_cfg.get("templates", {})
    template_block = templates_map.get(template_key, {})

    stage_root = dropbox_cfg.get("stage_root", "").rstrip("/")
    stage_folder_path = f"{stage_root}/{staged_folder_name}" if staged_folder_name else ""
    combo_map = template_block.get("design_color_image_map", {})

    selected_colors = selected_variants.get("color", [])
    selected_designs = selected_variants.get("design", [])

    rows: list[dict[str, str]] = []

    for color in selected_colors:
        design_map = combo_map.get(color, {})
        for design in selected_designs:
            filename = design_map.get(design, "")
            path = f"{stage_folder_path}/{filename}" if stage_folder_path and filename else ""
            rows.append({
                "color": color,
                "design": design,
                "path": path,
            })

    return rows


def render_design_color_grid(
    entries: list[dict[str, Any]],
    cols_per_row: int = 5,
    image_width: int = 150,
) -> None:
    st.markdown("**Design/colour image mapping**")

    if not entries:
        st.caption("No design/colour combinations configured.")
        return

    cols = st.columns(cols_per_row)
    for idx, entry in enumerate(entries):
        with cols[idx % cols_per_row]:
            label = entry.get("label", "")
            st.caption(label)

            path = entry.get("path", "")
            if not path:
                st.warning("Missing")
                continue

            if entry.get("exists") and entry.get("direct_url"):
                st.image(entry["direct_url"], width=image_width)
            else:
                st.warning("Not found")
                st.code(path, language=None)

def render_variant_combinations_preview(
    profile: dict[str, Any],
    parent_sku: str,
    selected_variants: dict[str, list[str]],
    base_title: str = "",
    sku_decoration_code: str = "",
    sku_listing_code: str = "",
) -> None:
    profile = apply_sku_context_to_profile(profile, sku_decoration_code, sku_listing_code)
    combos = build_variant_combinations(profile, selected_variants)

    st.markdown("**Selected variant combinations**")

    if not combos:
        st.caption("No combinations selected.")
        return

    rows = []
    for idx, combo in enumerate(combos, start=1):
        row = {"#": idx}
        row.update(combo)
        row["child_title"] = build_child_item_name(base_title, combo, profile)
        sku_details = build_child_sku_details(profile, parent_sku, combo)
        row["child_sku"] = sku_details["amazon_seller_sku"]
        if has_stock_reference(profile):
            row["supplier"] = sku_details.get("supplier", "")
            row["supplier_stock_key"] = sku_details.get("supplier_stock_key", "")
            row["supplier_stock_key_status"] = sku_details.get("supplier_stock_key_status", "")
            row["supplier_stock_key_reason"] = sku_details.get("supplier_stock_key_reason", "")
        rows.append(row)

    st.dataframe(rows, width="stretch", hide_index=True)

def get_profile_color_options(profile: dict[str, Any]) -> list[str]:
    colors = list(profile.get("colors", []))
    if colors:
        return colors

    color_size_map = profile.get("color_size_map", {})
    if color_size_map:
        return list(color_size_map.keys())

    color_sku_map = profile.get("color_sku_map", {})
    if color_sku_map:
        return list(color_sku_map.keys())

    return []


def normalize_saved_selected_variants(saved_selected_variants: dict[str, Any]) -> dict[str, list[str]]:
    normalized: dict[str, list[str]] = {}
    for key, value in dict(saved_selected_variants or {}).items():
        normalized_key = str(key).strip().lower()
        if isinstance(value, list):
            normalized[normalized_key] = list(value)
        elif value is None:
            normalized[normalized_key] = []
        else:
            normalized[normalized_key] = [value]
    return normalized


def get_saved_variant_values(
    saved_variants_normalized: dict[str, list[str]],
    dim_name: str,
) -> list[str]:
    normalized_name = str(dim_name).strip().lower()
    aliases = [normalized_name]

    if normalized_name in {"color", "colour", "colors", "colours"}:
        aliases.extend(["color", "colour", "colors", "colours"])
    elif normalized_name in {"size", "sizes"}:
        aliases.extend(["size", "sizes"])
    elif normalized_name in {"design", "style", "styles"}:
        aliases.extend(["design", "style", "styles"])

    for alias in aliases:
        candidate = saved_variants_normalized.get(alias)
        if candidate is not None:
            return list(candidate)

    return []


def normalize_multiselect_values(
    current_values: list[str] | None,
    valid_options: list[str],
    fallback_values: list[str] | None,
    allow_empty: bool = True,
) -> tuple[list[str], bool]:
    valid_options = list(valid_options)
    current_missing = current_values is None
    current_list = list(current_values or [])
    fallback_list = list(fallback_values or [])

    valid_current = [value for value in current_list if value in valid_options]
    valid_fallback = [value for value in fallback_list if value in valid_options]

    if not valid_fallback:
        valid_fallback = list(valid_options)

    should_reset = (
        current_missing
        or not valid_current
        or len(valid_current) != len(current_list)
    )

    if allow_empty and not current_missing and not current_list:
        return [], False

    if should_reset:
        return valid_fallback, True

    return valid_current, False


def expand_full_selection_after_option_change(
    session_key: str,
    options_state_key: str,
    current_values: list[str] | None,
    valid_options: list[str],
) -> tuple[list[str] | None, bool]:
    previous_options = st.session_state.get(options_state_key)
    st.session_state[options_state_key] = list(valid_options)

    if current_values is None or not isinstance(previous_options, list):
        return current_values, False

    current_list = list(current_values)
    previous_list = [value for value in previous_options if value in valid_options]
    if not previous_list or len(previous_list) == len(valid_options):
        return current_values, False

    was_all_previous_options = (
        len(current_list) == len(previous_list)
        and set(current_list) == set(previous_list)
    )
    if not was_all_previous_options:
        return current_values, False

    expanded_values = list(valid_options)
    st.session_state[session_key] = expanded_values
    return expanded_values, True


def normalize_selected_variants_session_state(
    profile: dict[str, Any],
    listing_memory: dict[str, Any],
    force_saved_values: bool = False,
    force_defaults: bool = False,
) -> dict[str, list[str]]:
    saved_variants_normalized = normalize_saved_selected_variants(
        listing_memory.get("selected_variants", {})
    )
    saved_variants_normalized = normalize_saved_variant_values_for_profile(
        profile,
        saved_variants_normalized,
    )
    variant_dimensions = profile.get("variant_dimensions", [])

    if variant_dimensions:
        normalized_variants: dict[str, list[str]] = {}
        for dim in variant_dimensions:
            dim_name = str(dim.get("name", "")).strip()
            dim_options = list(dim.get("options", []))
            widget_key = f"variant_{dim_name}"
            saved_values = [
                value for value in get_saved_variant_values(saved_variants_normalized, dim_name)
                if value in dim_options
            ]
            fallback_values = list(dim_options) if force_defaults else (saved_values or list(dim_options))
            current_values = (
                []
                if force_saved_values or force_defaults
                else st.session_state.get(widget_key) if widget_key in st.session_state else None
            )
            if not force_saved_values and not force_defaults:
                current_values, _ = expand_full_selection_after_option_change(
                    widget_key,
                    f"{widget_key}_available_options",
                    current_values,
                    dim_options,
                )
            normalized_values, should_set = normalize_multiselect_values(
                current_values,
                dim_options,
                fallback_values,
                allow_empty=not (force_saved_values or force_defaults),
            )
            if should_set or widget_key not in st.session_state:
                st.session_state[widget_key] = list(normalized_values)
            normalized_variants[dim_name] = list(st.session_state.get(widget_key, normalized_values))

        return normalized_variants

    color_options = get_profile_color_options(profile)
    saved_colors = [
        color for color in get_saved_variant_values(saved_variants_normalized, "color")
        if color in color_options
    ]
    color_fallback = list(color_options) if force_defaults else (saved_colors or list(color_options))
    current_colors = (
        []
        if force_saved_values or force_defaults
        else st.session_state.get("selected_colours") if "selected_colours" in st.session_state else None
    )
    if not force_saved_values and not force_defaults:
        current_colors, _ = expand_full_selection_after_option_change(
            "selected_colours",
            "selected_colours_available_options",
            current_colors,
            color_options,
        )
    normalized_colors, should_set_colors = normalize_multiselect_values(
        current_colors,
        color_options,
        color_fallback,
        allow_empty=not (force_saved_values or force_defaults),
    )
    if should_set_colors or "selected_colours" not in st.session_state:
        st.session_state["selected_colours"] = list(normalized_colors)
    selected_colors = list(st.session_state.get("selected_colours", normalized_colors))

    available_sizes = get_available_sizes_for_selected_colors(profile, selected_colors)
    saved_sizes = [
        size for size in get_saved_variant_values(saved_variants_normalized, "size")
        if size in available_sizes
    ]
    size_fallback = list(available_sizes) if force_defaults else (saved_sizes or list(available_sizes))
    current_sizes = (
        []
        if force_saved_values or force_defaults
        else st.session_state.get("selected_sizes") if "selected_sizes" in st.session_state else None
    )
    normalized_sizes, should_set_sizes = normalize_multiselect_values(
        current_sizes,
        available_sizes,
        size_fallback,
        allow_empty=not (force_saved_values or force_defaults),
    )
    if should_set_sizes or "selected_sizes" not in st.session_state:
        st.session_state["selected_sizes"] = list(normalized_sizes)
    selected_sizes = list(st.session_state.get("selected_sizes", normalized_sizes))

    return {
        "color": selected_colors,
        "size": selected_sizes,
    }


def get_available_sizes_for_selected_colors(
    profile: dict[str, Any],
    selected_colors: list[str],
) -> list[str]:
    all_sizes = profile.get("sizes", [])
    color_size_map = profile.get("color_size_map", {})

    if not color_size_map or not selected_colors:
        return all_sizes

    allowed: list[str] = []
    seen: set[str] = set()

    for color in selected_colors:
        for size in color_size_map.get(color, []):
            if size in all_sizes and size not in seen:
                allowed.append(size)
                seen.add(size)

    return allowed

def build_dropbox_overview(
    profile: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    include_garment_resource_images: bool = True,
) -> dict[str, Any]:
    if not dropbox_cfg:
        return {}

    template_key = profile.get("template_key", "")
    templates_map = dropbox_cfg.get("templates", {})
    template_block = templates_map.get(template_key, {})

    resource_root = dropbox_cfg.get("resource_root", "").rstrip("/")
    variant_folder = template_block.get("variant_folder", template_key)

    shared_resource_images = [
        f"{resource_root}/{name}"
        for name in dropbox_cfg.get("general_resource_images", [])
    ]

    garment_resource_root = f"{resource_root}/{variant_folder}" if resource_root and variant_folder else ""
    garment_resource_images: list[str] = []
    garment_resource_warning = ""
    resource_groups_folder = str(template_block.get("resource_groups_folder", "")).strip().strip("/")
    garment_resource_group_root = (
        f"{resource_root}/{resource_groups_folder}"
        if resource_root and resource_groups_folder
        else ""
    )
    garment_resource_group_root_images: list[str] = []
    garment_resource_group_root_warning = ""
    garment_resource_groups: list[dict[str, Any]] = []

    if garment_resource_root and include_garment_resource_images:
        try:
            garment_resource_images = [
                p for p in list_folder_files(garment_resource_root)
                if is_image_file(p)
            ]
            garment_resource_images = sorted(
                garment_resource_images,
                key=lambda p: Path(p).name.lower(),
            )
            if not garment_resource_images:
                garment_resource_warning = (
                    f"No garment support images found in {garment_resource_root}."
                )
        except Exception as exc:
            garment_resource_warning = f"Garment support images unavailable: {exc}"
    elif garment_resource_root:
        garment_resource_warning = "Garment support images not loaded yet."

    if garment_resource_group_root and include_garment_resource_images:
        try:
            garment_resource_group_root_images = sorted(
                [
                    path
                    for path in list_folder_files(garment_resource_group_root)
                    if is_image_file(path)
                ],
                key=lambda path: Path(path).name.lower(),
            )
            if not garment_resource_group_root_images:
                garment_resource_group_root_warning = (
                    f"No shared resource images found in {garment_resource_group_root}."
                )
        except Exception as exc:
            garment_resource_group_root_warning = f"Shared resource images unavailable: {exc}"
    elif garment_resource_group_root:
        garment_resource_group_root_warning = "Shared resource images not loaded yet."

    for resource_group in template_block.get("resource_groups", []):
        group_key = str(resource_group.get("key", "")).strip()
        group_label = str(resource_group.get("label", "")).strip()
        group_folder = str(resource_group.get("folder", "")).strip().strip("/")
        group_path = (
            f"{garment_resource_group_root}/{group_folder}"
            if garment_resource_group_root and group_folder
            else ""
        )
        group_images: list[str] = []
        group_warning = ""
        if group_path and include_garment_resource_images:
            try:
                group_images = sorted(
                    [path for path in list_folder_files(group_path) if is_image_file(path)],
                    key=lambda path: Path(path).name.lower(),
                )
                if not group_images:
                    group_warning = f"No resource images found in {group_path}."
            except Exception as exc:
                group_warning = f"{group_label or group_key} resource images unavailable: {exc}"
        elif group_path:
            group_warning = "Resource images not loaded yet."

        garment_resource_groups.append({
            "key": group_key,
            "label": group_label,
            "folder": group_folder,
            "path": group_path,
            "images": group_images,
            "warning": group_warning,
        })

    return {
        "resource_root": resource_root,
        "template_key": template_key,
        "variant_folder": variant_folder,
        "garment_resource_root": garment_resource_root,
        "garment_resource_images": garment_resource_images,
        "garment_resource_warning": garment_resource_warning,
        "garment_resource_group_root": garment_resource_group_root,
        "garment_resource_group_root_images": garment_resource_group_root_images,
        "garment_resource_group_root_warning": garment_resource_group_root_warning,
        "garment_resource_groups": garment_resource_groups,
        "shared_resource_images": shared_resource_images,
        "main_image_map": template_block.get("main_image_map", {}),
        "design_color_image_map": template_block.get("design_color_image_map", {}),
        "color_sku_map": profile.get("color_sku_map", {}),
    }


def normalize_image_filename_part(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "-", str(value or "").strip()).strip("-")


def normalize_image_match_key(value: str) -> str:
    stem = re.sub(r"\.[A-Za-z0-9]+$", "", str(value or "").strip())
    return normalize_image_filename_part(stem).lower()


def tokenize_image_match_value(value: str) -> list[str]:
    stem = re.sub(r"\.[A-Za-z0-9]+$", "", str(value or "").strip())
    return re.findall(r"[a-z0-9]+", stem.lower())


def contains_token_sequence(tokens: list[str], sequence: list[str]) -> bool:
    if not sequence:
        return False

    sequence_len = len(sequence)
    for idx in range(0, len(tokens) - sequence_len + 1):
        if tokens[idx:idx + sequence_len] == sequence:
            return True
    return False


def infer_design_color_image_path_map_from_paths(
    profile: dict[str, Any],
    image_paths: list[str],
    selected_colors: list[str],
    selected_designs: list[str] | None = None,
) -> dict[str, dict[str, str]]:
    design_sku_map = {
        str(design): str(code)
        for design, code in dict(profile.get("design_sku_map", {}) or {}).items()
        if str(design or "").strip() and str(code or "").strip()
    }
    if not design_sku_map or not image_paths:
        return {}

    selected_design_set = set(selected_designs or [])
    candidate_designs = [
        design
        for design in design_sku_map
        if not selected_design_set or design in selected_design_set
    ]
    if not candidate_designs:
        return {}

    color_tokens = {
        color: tokenize_image_match_value(color)
        for color in selected_colors
        if str(color or "").strip()
    }
    design_code_tokens = {
        design: tokenize_image_match_value(code)
        for design, code in design_sku_map.items()
        if design in candidate_designs
    }

    inferred: dict[str, dict[str, str]] = {}
    for path in image_paths:
        filename_tokens = tokenize_image_match_value(Path(path).name)
        if not filename_tokens:
            continue

        matched_designs = [
            design
            for design, code_tokens in design_code_tokens.items()
            if contains_token_sequence(filename_tokens, code_tokens)
        ]
        if not matched_designs:
            continue

        matched_colors = [
            color
            for color, tokens in color_tokens.items()
            if contains_token_sequence(filename_tokens, tokens)
        ]
        if not matched_colors:
            continue

        for color in matched_colors:
            for design in matched_designs:
                inferred.setdefault(color, {}).setdefault(design, path)

    return inferred


def infer_design_color_image_url_map_from_folder(
    profile: dict[str, Any],
    folder_path: str,
    selected_colors: list[str],
    selected_designs: list[str] | None = None,
) -> dict[str, dict[str, str]]:
    if not folder_path:
        return {}

    try:
        image_paths = [path for path in list_folder_files(folder_path) if is_image_file(path)]
    except Exception:
        return {}

    path_map = infer_design_color_image_path_map_from_paths(
        profile=profile,
        image_paths=image_paths,
        selected_colors=selected_colors,
        selected_designs=selected_designs,
    )
    url_map: dict[str, dict[str, str]] = {}
    for color, design_map in path_map.items():
        for design, path in design_map.items():
            try:
                image_url = dropbox_preview_url(path)
            except Exception:
                image_url = ""
            if image_url:
                url_map.setdefault(color, {})[design] = image_url

    return url_map


def build_color_image_filename_candidates(
    template_key: str,
    color: str,
    configured_filename: str = "",
    color_code: str = "",
) -> list[str]:
    color_part = normalize_image_filename_part(color)
    template_part = normalize_image_filename_part(template_key)
    code_part = normalize_image_filename_part(color_code)
    configured_filename = str(configured_filename or "").strip()
    extension = Path(configured_filename).suffix if configured_filename else ".jpg"
    candidates: list[str] = []

    for filename in [
        configured_filename,
        f"{template_part}{code_part}{extension}" if template_part and code_part else "",
        f"{template_part}-{color_part}{extension}" if template_part and color_part else "",
        f"{color_part.replace('-', ' ')}.jpg" if color_part else "",
        f"{color_part.replace('-', ' ')}.png" if color_part else "",
        f"{color}.jpg" if color else "",
        f"{color}.png" if color else "",
        f"{color_part}.jpg" if color_part else "",
        f"{color_part}.png" if color_part else "",
    ]:
        if filename and filename not in candidates:
            candidates.append(filename)

    return candidates


def resolve_existing_color_image_path_from_paths(
    image_paths: list[str],
    template_key: str,
    color: str,
    configured_filename: str = "",
    color_code: str = "",
) -> tuple[str, list[str]]:
    candidates = build_color_image_filename_candidates(
        template_key,
        color,
        configured_filename,
        color_code,
    )
    if not image_paths:
        return "", candidates

    candidate_lookup = {filename.lower(): filename for filename in candidates}
    candidate_stem_lookup = {
        normalize_image_match_key(filename): filename
        for filename in candidates
        if normalize_image_match_key(filename)
    }

    for path in image_paths:
        filename = Path(path).name
        suffix = Path(filename).suffix.lower()
        if filename.lower() in candidate_lookup:
            return path, candidates
        if suffix in {".jpg", ".jpeg", ".png", ".webp"}:
            filename_key = normalize_image_match_key(filename)
            if filename_key in candidate_stem_lookup:
                return path, candidates

    return "", candidates


def resolve_existing_color_image_path(
    folder_path: str,
    template_key: str,
    color: str,
    configured_filename: str = "",
    color_code: str = "",
) -> tuple[str, list[str]]:
    candidates = build_color_image_filename_candidates(
        template_key,
        color,
        configured_filename,
        color_code,
    )
    for filename in candidates:
        path = f"{folder_path}/{filename}"
        if path_exists(path):
            return path, candidates

    try:
        candidate_lookup = {filename.lower(): filename for filename in candidates}
        candidate_stem_lookup = {
            normalize_image_match_key(filename): filename
            for filename in candidates
            if normalize_image_match_key(filename)
        }
        for path in list_folder_files(folder_path):
            filename = Path(path).name
            if filename.lower() in candidate_lookup:
                return path, candidates
            if Path(filename).suffix.lower() in {".jpg", ".jpeg", ".png"}:
                filename_key = normalize_image_match_key(filename)
                if filename_key in candidate_stem_lookup:
                    return path, candidates
    except Exception:
        pass

    return "", candidates


def resolve_child_variant_image_url(
    variant_values: dict[str, str],
    color_image_map: dict[str, str],
    design_color_image_url_map: dict[str, dict[str, str]] | None = None,
) -> str:
    design_color_image_url_map = design_color_image_url_map or {}

    color_value = variant_values.get("color", "")
    design_value = variant_values.get("design", "")

    if color_value and design_value:
        image_url = (
            design_color_image_url_map
            .get(color_value, {})
            .get(design_value, "")
        )
        if image_url:
            return image_url

    if color_value:
        return color_image_map.get(color_value, "")

    return ""


def build_parent_main_image_options(
    profile: dict[str, Any],
    selected_variants: dict[str, list[str]],
    color_image_map: dict[str, str],
    design_color_image_url_map: dict[str, dict[str, str]] | None = None,
) -> list[tuple[str, str]]:
    options: list[tuple[str, str]] = []
    seen_urls: set[str] = set()
    variant_combos = build_variant_combinations(profile, selected_variants)

    for combo in variant_combos:
        image_url = resolve_child_variant_image_url(
            variant_values=combo,
            color_image_map=color_image_map,
            design_color_image_url_map=design_color_image_url_map,
        )
        if not image_url or image_url in seen_urls:
            continue

        label = " / ".join([v for v in combo.values() if v]) or "Unnamed variant"
        options.append((label, image_url))
        seen_urls.add(image_url)

    return options


def resolve_selected_parent_main_image_url(
    parent_main_options: list[tuple[str, str]],
    selected_parent_main_image_label: str = "",
    selected_parent_main_image_url: str = "",
) -> str:
    selected_label = str(selected_parent_main_image_label or "").strip()
    if selected_label and selected_label != "Automatic (recommended)":
        for label, url in parent_main_options:
            if label == selected_label and url:
                return url

    if selected_parent_main_image_url:
        return selected_parent_main_image_url

    return parent_main_options[0][1] if parent_main_options else ""


def build_dropbox_overview_cache_key(
    profile: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    include_garment_resource_images: bool = True,
) -> str:
    template_key = profile.get("template_key", "")
    cache_parts = {
        "template_key": template_key,
        "template_cfg": dropbox_cfg.get("templates", {}).get(template_key, {}),
        "general_resource_images": dropbox_cfg.get("general_resource_images", []),
        "resource_root": dropbox_cfg.get("resource_root", ""),
        "include_garment_resource_images": include_garment_resource_images,
    }
    return json.dumps(cache_parts, sort_keys=True)


def get_cached_dropbox_overview(
    profile: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    include_garment_resource_images: bool = True,
) -> dict[str, Any]:
    cache_key = build_dropbox_overview_cache_key(
        profile,
        dropbox_cfg,
        include_garment_resource_images=include_garment_resource_images,
    )
    cache = st.session_state.get("dropbox_overview_cache", {})

    if cache.get("key") == cache_key:
        return cache.get("data", {})

    data = build_dropbox_overview(
        profile,
        dropbox_cfg,
        include_garment_resource_images=include_garment_resource_images,
    )
    st.session_state["dropbox_overview_cache"] = {
        "key": cache_key,
        "data": data,
    }
    return data


def clear_runtime_caches() -> None:
    for key in [
        "listing_memory_cache",
        "dropbox_overview_cache",
        "preview_image_cache",
        "preview_image_mapping_cache",
        "resolved_image_bundle_cache",
        "ready_queue_items_cache",
        "approved_queue_items_cache",
        "load_image_mappings_now",
        "scan_mapped_colours_now",
        "current_run_image_resolution_debug",
    ]:
        st.session_state.pop(key, None)

    for key in list(st.session_state.keys()):
        if key.endswith("_load_image_review") or key.endswith("_run_full_quality"):
            st.session_state.pop(key, None)


def clear_resource_image_caches() -> None:
    for key in [
        "dropbox_overview_cache",
        "preview_image_cache",
        "preview_image_mapping_cache",
        "resolved_image_bundle_cache",
    ]:
        st.session_state.pop(key, None)


def clear_loaded_image_mapping_state(session_state: Any) -> None:
    for key in [
        "load_image_mappings_now",
        "image_mappings_loaded_folder",
        "image_mappings_loaded_context",
        "preview_image_cache",
        "preview_image_mapping_cache",
        "resolved_image_bundle_cache",
    ]:
        session_state.pop(key, None)


def set_workflow_flash(level: str, message: str, detail: str = "") -> None:
    st.session_state["workflow_flash"] = {
        "level": level,
        "message": message,
        "detail": detail,
    }


def render_workflow_flash() -> None:
    flash = st.session_state.pop("workflow_flash", None)
    if not flash:
        return

    level = flash.get("level", "info")
    message = flash.get("message", "")
    detail = flash.get("detail", "")

    if message:
        if level == "success":
            st.success(message)
        elif level == "warning":
            st.warning(message)
        elif level == "error":
            st.error(message)
        else:
            st.info(message)

    if detail:
        st.info(detail)


def build_safe_error_context() -> dict[str, str]:
    keys = [
        "workflow_active_tab",
        "active_folder_source_mode",
        "active_staged_folder_select",
        "active_template_family_select",
        "active_listing_template_select",
        "folder_source_mode",
        "staged_folder_select",
        "template_family_select",
        "listing_template_select",
        "image_mappings_loaded_folder",
        "image_mappings_loaded_context",
    ]
    return {
        key: normalize_debug_state_value(st.session_state.get(key, ""))
        for key in keys
        if key in st.session_state
    }


def record_app_error(exc: Exception) -> dict[str, Any]:
    error_row = {
        "timestamp": format_workflow_timestamp(),
        "type": type(exc).__name__,
        "message": str(exc),
        "context": build_safe_error_context(),
        "traceback": traceback.format_exc(),
    }
    history = list(st.session_state.get("app_error_history", []))
    history.append(error_row)
    st.session_state["app_error_history"] = history[-10:]
    return error_row


def render_app_error_report(error_row: dict[str, Any]) -> None:
    st.error("The app hit an error, but the context was captured so we can debug it.")
    st.write(f"Error: `{error_row.get('type', '')}`")
    if error_row.get("message"):
        st.code(str(error_row.get("message", "")), language=None)

    with st.expander("Captured app context", expanded=True):
        context = dict(error_row.get("context", {}) or {})
        if context:
            st.json(context)
        else:
            st.caption("No context keys were available.")

    with st.expander("Traceback", expanded=False):
        st.code(str(error_row.get("traceback", "")), language="python")

    history = list(st.session_state.get("app_error_history", []))
    if history:
        with st.expander("Recent app errors in this session", expanded=False):
            st.dataframe(
                [
                    {
                        "timestamp": row.get("timestamp", ""),
                        "type": row.get("type", ""),
                        "message": row.get("message", ""),
                    }
                    for row in history
                ],
                hide_index=True,
                width="stretch",
            )


def build_preview_image_cache_key(
    profile: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    selected_variants: dict[str, list[str]],
    include_mappings: bool = False,
    resolve_preview_urls: bool = False,
) -> str:
    template_key = profile.get("template_key", "")
    cache_parts = {
        "template_key": template_key,
        "staged_folder_name": staged_folder_name,
        "selected_colors": get_selected_colors_for_image_resolution(profile, selected_variants),
        "selected_designs": selected_variants.get("design", []),
        "template_cfg": dropbox_cfg.get("templates", {}).get(template_key, {}),
        "general_resource_images": dropbox_cfg.get("general_resource_images", []),
        "include_mappings": include_mappings,
        "resolve_preview_urls": resolve_preview_urls,
    }
    return json.dumps(cache_parts, sort_keys=True)


def build_preview_image_mapping_cache_key(
    profile: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    resolve_preview_urls: bool = False,
) -> str:
    template_key = profile.get("template_key", "")
    cache_parts = {
        "template_key": template_key,
        "template_slug": profile.get("_slug", ""),
        "staged_folder_name": staged_folder_name,
        "template_cfg": dropbox_cfg.get("templates", {}).get(template_key, {}),
        "general_resource_images": dropbox_cfg.get("general_resource_images", []),
        "resource_root": dropbox_cfg.get("resource_root", ""),
        "resolve_preview_urls": resolve_preview_urls,
    }
    return json.dumps(cache_parts, sort_keys=True)


def build_image_mapping_variants_for_cache(
    profile: dict[str, Any],
    selected_variants: dict[str, list[str]],
) -> dict[str, list[str]]:
    # Heavy Dropbox image mapping should be cached at folder/template level,
    # not invalidated by normal listing-content variant edits.
    mapping_variants: dict[str, list[str]] = {}

    colors = get_profile_color_options(profile)
    if colors:
        mapping_variants["color"] = list(colors)
    elif selected_variants.get("color"):
        mapping_variants["color"] = list(selected_variants.get("color", []))

    for dim in profile.get("variant_dimensions", []):
        dim_name = str(dim.get("name", "")).strip().lower()
        if dim_name == "design":
            options = list(dim.get("options", []))
            if options:
                mapping_variants["design"] = options
            elif selected_variants.get("design"):
                mapping_variants["design"] = list(selected_variants.get("design", []))

    return mapping_variants or dict(selected_variants)


def filter_preview_image_maps_for_selected_variants(
    profile: dict[str, Any],
    selected_variants: dict[str, list[str]],
    full_color_image_map: dict[str, str],
    full_design_color_image_url_map: dict[str, dict[str, str]],
) -> tuple[dict[str, str], dict[str, dict[str, str]]]:
    selected_colors = get_selected_colors_for_image_resolution(profile, selected_variants)
    selected_designs = list(selected_variants.get("design", []))

    if selected_colors:
        color_image_map = {
            color: image_url
            for color, image_url in full_color_image_map.items()
            if color in selected_colors
        }
    else:
        color_image_map = dict(full_color_image_map)

    if selected_colors or selected_designs:
        design_color_image_url_map: dict[str, dict[str, str]] = {}
        for color, design_map in full_design_color_image_url_map.items():
            if selected_colors and color not in selected_colors:
                continue

            filtered_design_map = {
                design: image_url
                for design, image_url in dict(design_map).items()
                if not selected_designs or design in selected_designs
            }

            if filtered_design_map:
                design_color_image_url_map[color] = filtered_design_map
    else:
        design_color_image_url_map = {
            color: dict(design_map)
            for color, design_map in full_design_color_image_url_map.items()
        }

    return color_image_map, design_color_image_url_map


def get_mapped_color_options(
    valid_colors: list[str],
    color_image_map: dict[str, str],
    design_color_image_url_map: dict[str, dict[str, str]],
) -> list[str]:
    valid_color_set = set(valid_colors)
    mapped_colors: list[str] = []

    for color in list(color_image_map.keys()) + list(design_color_image_url_map.keys()):
        if color in valid_color_set and color not in mapped_colors:
            mapped_colors.append(color)

    return mapped_colors


def get_color_widget_keys(profile: dict[str, Any]) -> list[str]:
    variant_dimensions = profile.get("variant_dimensions", [])
    if variant_dimensions:
        return [
            f"variant_{str(dim.get('name', '')).strip()}"
            for dim in variant_dimensions
            if str(dim.get("name", "")).strip().lower() in {"color", "colour"}
        ]
    return ["selected_colours"]


def apply_mapped_colors_to_widget_once(
    widget_key: str,
    mapped_colors: list[str],
    context_key: str,
    allow_replace_existing: bool = False,
) -> bool:
    if not mapped_colors:
        return False

    applied_key = f"{widget_key}_auto_mapped_context"
    if st.session_state.get(applied_key) == context_key:
        return False

    if widget_key in st.session_state and not allow_replace_existing:
        return False

    st.session_state[widget_key] = list(mapped_colors)
    st.session_state[applied_key] = context_key
    return True


def build_lenient_image_maps(
    profile: dict[str, Any],
    selected_variants: dict[str, list[str]],
    selected_colors: list[str],
    dropbox_overview: dict[str, Any],
    folder_path: str,
    image_paths: list[str] | None = None,
    resolve_urls: bool = True,
) -> tuple[dict[str, str], dict[str, dict[str, str]]]:
    template_key = str(dropbox_overview.get("template_key", "") or "")
    main_image_map = dropbox_overview.get("main_image_map", {})
    design_color_image_map = dropbox_overview.get("design_color_image_map", {})
    color_sku_map = dropbox_overview.get("color_sku_map", {})
    image_paths = list(image_paths or [])
    if not image_paths and folder_path:
        try:
            image_paths = [path for path in list_folder_files(folder_path) if is_image_file(path)]
        except Exception:
            image_paths = []

    color_image_map: dict[str, str] = {}
    for color in selected_colors:
        filename = main_image_map.get(color, "")
        path, _ = resolve_existing_color_image_path_from_paths(
            image_paths,
            template_key,
            color,
            filename,
            str(color_sku_map.get(color, "") or ""),
        )
        if not path:
            continue
        if resolve_urls:
            try:
                resolved_image = dropbox_preview_url(path)
            except Exception:
                resolved_image = ""
        else:
            resolved_image = path
        if resolved_image:
            color_image_map[color] = resolved_image

    design_color_image_url_map: dict[str, dict[str, str]] = {}
    image_path_set = set(image_paths)
    for color in selected_colors:
        for design, filename in dict(design_color_image_map.get(color, {})).items():
            if not filename:
                continue
            path = f"{folder_path}/{filename}"
            if image_paths and path not in image_path_set:
                continue
            if not image_paths and not path_exists(path):
                continue
            if resolve_urls:
                try:
                    resolved_image = dropbox_preview_url(path)
                except Exception:
                    resolved_image = ""
            else:
                resolved_image = path
            if resolved_image:
                design_color_image_url_map.setdefault(color, {})[design] = resolved_image

    inferred_design_color_image_url_map = infer_design_color_image_path_map_from_paths(
        profile=profile,
        image_paths=image_paths,
        selected_colors=selected_colors,
        selected_designs=list(selected_variants.get("design", [])),
    )
    for color, design_map in inferred_design_color_image_url_map.items():
        for design, path in design_map.items():
            if resolve_urls:
                try:
                    resolved_image = dropbox_preview_url(path)
                except Exception:
                    resolved_image = ""
            else:
                resolved_image = path
            if resolved_image:
                design_color_image_url_map.setdefault(color, {}).setdefault(design, resolved_image)

    return color_image_map, design_color_image_url_map


def get_cached_preview_image_data(
    profile: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    selected_variants: dict[str, list[str]],
    dropbox_overview: dict[str, Any],
    include_mappings: bool = False,
    resolve_preview_urls: bool = False,
) -> dict[str, Any]:
    cache_key = build_preview_image_cache_key(
        profile,
        dropbox_cfg,
        staged_folder_name,
        selected_variants,
        include_mappings,
        resolve_preview_urls,
    )
    cache = st.session_state.get("preview_image_cache", {})

    if cache.get("key") == cache_key:
        return cache.get("data", {})

    staged_preview_paths = build_stage_preview_paths(dropbox_cfg, staged_folder_name) if staged_folder_name else []
    staged_resource_paths = build_stage_resource_image_paths(dropbox_cfg, staged_folder_name) if staged_folder_name else []
    design_color_preview_rows = build_design_color_preview_paths(
        profile=profile,
        dropbox_cfg=dropbox_cfg,
        selected_variants=selected_variants,
        staged_folder_name=staged_folder_name or "",
    )

    preview_color_image_map: dict[str, str] = {}
    preview_design_color_image_url_map: dict[str, dict[str, str]] = {}
    full_color_image_map: dict[str, str] = {}
    full_design_color_image_url_map: dict[str, dict[str, str]] = {}
    parent_main_image_options: list[tuple[str, str]] = []

    def resolve_display_entries(items: list[tuple[str, str]]) -> list[dict[str, Any]]:
        entries: list[dict[str, Any]] = []
        for label, path in items:
            if not path:
                entries.append({"label": label, "path": path, "exists": False, "direct_url": ""})
                continue
            if not resolve_preview_urls:
                entries.append({"label": label, "path": path, "exists": True, "direct_url": ""})
                continue
            try:
                result = resolve_one(path, label)
                entries.append({
                    "label": label,
                    "path": path,
                    "exists": result.get("exists", False),
                    "direct_url": result.get("direct_url", ""),
                })
            except Exception:
                entries.append({"label": label, "path": path, "exists": False, "direct_url": ""})
        return entries

    if staged_folder_name and include_mappings:
        try:
            preview_stage_folder_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)
            mapping_cache_key = build_preview_image_mapping_cache_key(
                profile,
                dropbox_cfg,
                staged_folder_name,
                resolve_preview_urls=resolve_preview_urls,
            )
            mapping_cache = st.session_state.get("preview_image_mapping_cache", {})

            if mapping_cache.get("key") == mapping_cache_key:
                mapping_data = dict(mapping_cache.get("data", {}))
                full_color_image_map = dict(mapping_data.get("color_image_map", {}))
                full_design_color_image_url_map = {
                    color: dict(design_map)
                    for color, design_map in dict(mapping_data.get("design_color_image_url_map", {})).items()
                }
            else:
                mapping_variants = build_image_mapping_variants_for_cache(profile, selected_variants)
                mapping_colors = get_selected_colors_for_image_resolution(profile, mapping_variants)

                full_color_image_map, full_design_color_image_url_map = build_lenient_image_maps(
                    profile,
                    mapping_variants,
                    mapping_colors,
                    dropbox_overview,
                    preview_stage_folder_path,
                    image_paths=staged_preview_paths,
                    resolve_urls=resolve_preview_urls,
                )

                st.session_state["preview_image_mapping_cache"] = {
                    "key": mapping_cache_key,
                    "data": {
                        "color_image_map": dict(full_color_image_map),
                        "design_color_image_url_map": {
                            color: dict(design_map)
                            for color, design_map in dict(full_design_color_image_url_map).items()
                        },
                    },
                }

            preview_color_image_map, preview_design_color_image_url_map = filter_preview_image_maps_for_selected_variants(
                profile,
                selected_variants,
                full_color_image_map,
                full_design_color_image_url_map,
            )
            if resolve_preview_urls:
                parent_main_image_options = build_parent_main_image_options(
                    profile=profile,
                    selected_variants=selected_variants,
                    color_image_map=preview_color_image_map,
                    design_color_image_url_map=preview_design_color_image_url_map,
                )
        except Exception:
            preview_color_image_map = {}
            preview_design_color_image_url_map = {}
            full_color_image_map = {}
            full_design_color_image_url_map = {}
            parent_main_image_options = []

    staged_preview_entries = resolve_display_entries([(Path(path).name, path) for path in staged_preview_paths])
    staged_resource_entries = resolve_display_entries([(Path(path).name, path) for path in staged_resource_paths])
    garment_resource_entries = resolve_display_entries([(Path(path).name, path) for path in dropbox_overview.get("garment_resource_images", [])])
    garment_resource_group_root_entries = resolve_display_entries([
        (Path(path).name, path)
        for path in dropbox_overview.get("garment_resource_group_root_images", [])
    ])
    garment_resource_group_entries = [
        {
            **resource_group,
            "entries": resolve_display_entries([
                (Path(path).name, path)
                for path in resource_group.get("images", [])
            ]),
        }
        for resource_group in dropbox_overview.get("garment_resource_groups", [])
    ]
    global_resource_entries = resolve_display_entries([(Path(path).name, path) for path in dropbox_overview.get("shared_resource_images", [])])

    stage_folder_path_for_preview = build_stage_folder_path(dropbox_cfg, staged_folder_name) if staged_folder_name else ""
    color_preview_source = get_profile_color_options(profile) or get_selected_colors_for_image_resolution(profile, selected_variants)
    staged_variant_entries: list[dict[str, Any]] = []
    fallback_variant_entries: list[tuple[str, str]] = []
    main_image_map = dropbox_overview.get("main_image_map", {})
    color_sku_map = dropbox_overview.get("color_sku_map", {})
    for color in color_preview_source:
        resolved_url = preview_color_image_map.get(color, "")
        if resolved_url:
            staged_variant_entries.append({
                "label": color,
                "path": resolved_url,
                "exists": True,
                "direct_url": resolved_url if resolve_preview_urls else "",
            })
            continue

        if stage_folder_path_for_preview and include_mappings:
            path, _ = resolve_existing_color_image_path_from_paths(
                staged_preview_paths,
                str(dropbox_overview.get("template_key", "") or profile.get("template_key", "") or ""),
                color,
                str(main_image_map.get(color, "") or ""),
                str(color_sku_map.get(color, "") or ""),
            )
            if path:
                try:
                    resolved_url = dropbox_preview_url(path) if resolve_preview_urls else path
                except Exception:
                    resolved_url = ""
            if resolved_url:
                staged_variant_entries.append({
                    "label": color,
                    "path": resolved_url,
                    "exists": True,
                    "direct_url": resolved_url if resolve_preview_urls else "",
                })
                continue

        fallback_variant_entries.append((
            color,
            f"{stage_folder_path_for_preview}/{main_image_map.get(color, '')}"
            if stage_folder_path_for_preview and main_image_map.get(color, "")
            else "",
        ))
    staged_variant_entries.extend(resolve_display_entries(fallback_variant_entries))
    design_color_preview_entries = resolve_display_entries([
        (f"{row['color']} / {row['design']}", row.get("path", ""))
        for row in design_color_preview_rows
    ])

    data = {
        "staged_preview_paths": staged_preview_paths,
        "staged_preview_entries": staged_preview_entries,
        "staged_resource_paths": staged_resource_paths,
        "staged_resource_entries": staged_resource_entries,
        "design_color_preview_rows": design_color_preview_rows,
        "design_color_preview_entries": design_color_preview_entries,
        "color_image_map": preview_color_image_map,
        "design_color_image_url_map": preview_design_color_image_url_map,
        "full_color_image_map": full_color_image_map,
        "full_design_color_image_url_map": full_design_color_image_url_map,
        "parent_main_image_options": parent_main_image_options,
        "garment_resource_entries": garment_resource_entries,
        "garment_resource_group_root_entries": garment_resource_group_root_entries,
        "garment_resource_group_entries": garment_resource_group_entries,
        "global_resource_entries": global_resource_entries,
        "staged_variant_entries": staged_variant_entries,
    }
    st.session_state["preview_image_cache"] = {"key": cache_key, "data": data}
    return data

def build_resolved_image_bundle_cache_key(
    profile: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    selected_variants: dict[str, list[str]],
    selected_parent_main_image_label: str = "",
    selected_parent_main_image_url: str = "",
    use_resource_fallback_images: bool = False,
) -> str:
    template_key = profile.get("template_key", "")
    cache_parts = {
        "template_key": template_key,
        "staged_folder_name": staged_folder_name,
        "selected_colors": get_selected_colors_for_image_resolution(profile, selected_variants),
        "selected_designs": selected_variants.get("design", []),
        "selected_parent_main_image_label": selected_parent_main_image_label,
        "selected_parent_main_image_url": selected_parent_main_image_url,
        "template_cfg": dropbox_cfg.get("templates", {}).get(template_key, {}),
        "general_resource_images": dropbox_cfg.get("general_resource_images", []),
        "resource_root": dropbox_cfg.get("resource_root", ""),
        "use_resource_fallback_images": use_resource_fallback_images,
    }
    return json.dumps(cache_parts, sort_keys=True)


def get_cached_resolved_image_bundle(
    profile: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    staged_folder_name: str,
    selected_variants: dict[str, list[str]],
    dropbox_overview: dict[str, Any],
    selected_parent_main_image_label: str = "",
    selected_parent_main_image_url: str = "",
    use_resource_fallback_images: bool = False,
) -> dict[str, Any]:
    if not staged_folder_name:
        return {
            "parent_main_image_url": "",
            "other_images": [],
            "color_image_map": {},
            "design_color_image_url_map": {},
        }

    cache_key = build_resolved_image_bundle_cache_key(
        profile,
        dropbox_cfg,
        staged_folder_name,
        selected_variants,
        selected_parent_main_image_label,
        selected_parent_main_image_url,
        use_resource_fallback_images,
    )
    cache = st.session_state.get("resolved_image_bundle_cache", {})

    if cache.get("key") == cache_key:
        return cache.get("data", {})

    stage_folder_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)
    parent_main_image_url, other_images, color_image_map, design_color_image_url_map = resolve_folder_image_urls(
        profile,
        selected_variants,
        get_selected_colors_for_image_resolution(profile, selected_variants),
        dropbox_overview,
        stage_folder_path,
        selected_parent_main_image_label=selected_parent_main_image_label,
        selected_parent_main_image_url=selected_parent_main_image_url,
        use_resource_fallback_images=use_resource_fallback_images,
    )

    data = {
        "parent_main_image_url": parent_main_image_url,
        "other_images": other_images,
        "color_image_map": color_image_map,
        "design_color_image_url_map": design_color_image_url_map,
    }
    st.session_state["resolved_image_bundle_cache"] = {
        "key": cache_key,
        "data": data,
    }
    return data


def resolve_folder_image_urls(
    profile: dict[str, Any],
    selected_variants: dict[str, list[str]],
    selected_colors: list[str],
    dropbox_overview: dict[str, Any],
    folder_path: str,
    selected_parent_main_image_label: str = "",
    selected_parent_main_image_url: str = "",
    use_resource_fallback_images: bool = False,
) -> tuple[str, list[str], dict[str, str], dict[str, dict[str, str]]]:
    template_key = str(dropbox_overview.get("template_key", "") or profile.get("template_key", "") or "")
    main_image_map = dropbox_overview.get("main_image_map", {})
    design_color_image_map = dropbox_overview.get("design_color_image_map", {})
    color_sku_map = dropbox_overview.get("color_sku_map", profile.get("color_sku_map", {}))
    garment_resource_images = dropbox_overview.get("garment_resource_images", [])
    shared_resource_images = dropbox_overview.get("shared_resource_images", [])
    staged_resource_images = list_image_paths_in_dropbox_folder(f"{folder_path.rstrip('/')}/resources")
    variant_combos = build_variant_combinations(profile, selected_variants)
    inferred_design_color_image_url_map = infer_design_color_image_url_map_from_folder(
        profile=profile,
        folder_path=folder_path,
        selected_colors=selected_colors,
        selected_designs=list(selected_variants.get("design", [])),
    )
    has_design_image_source = bool(design_color_image_map) or bool(inferred_design_color_image_url_map)

    color_image_map: dict[str, str] = {}
    for color in selected_colors:
        filename = main_image_map.get(color, "")
        path, candidates = resolve_existing_color_image_path(
            folder_path,
            template_key,
            color,
            filename,
            str(color_sku_map.get(color, "") or ""),
        )
        if not path:
            if has_design_image_source:
                continue
            expected = ", ".join(candidates[:4])
            raise ValueError(f"Missing staged mapped image for colour '{color}'. Expected one of: {expected}")
        url = dropbox_preview_url(path)
        if url:
            color_image_map[color] = url

    design_color_image_url_map: dict[str, dict[str, str]] = {}
    for color in selected_colors:
        design_map = design_color_image_map.get(color, {})
        design_color_image_url_map[color] = {}
        for design, filename in design_map.items():
            if not filename:
                continue
            path = f"{folder_path}/{filename}"
            if not path_exists(path):
                continue
            url = dropbox_preview_url(path)
            if url:
                design_color_image_url_map[color][design] = url

    for color, design_map in inferred_design_color_image_url_map.items():
        for design, image_url in design_map.items():
            design_color_image_url_map.setdefault(color, {}).setdefault(design, image_url)

    parent_main_image_url = ""
    missing_variant_labels: list[str] = []

    parent_main_options = build_parent_main_image_options(
        profile=profile,
        selected_variants=selected_variants,
        color_image_map=color_image_map,
        design_color_image_url_map=design_color_image_url_map,
    )

    for combo in variant_combos:
        image_url = resolve_child_variant_image_url(
            variant_values=combo,
            color_image_map=color_image_map,
            design_color_image_url_map=design_color_image_url_map,
        )
        if not image_url:
            label = " / ".join([v for v in combo.values() if v]) or "Unnamed variant"
            missing_variant_labels.append(label)

    if missing_variant_labels:
        raise ValueError(
            "Missing staged mapped image for variant(s): " + ", ".join(missing_variant_labels)
        )

    parent_main_image_url = resolve_selected_parent_main_image_url(
        parent_main_options,
        selected_parent_main_image_label=selected_parent_main_image_label,
        selected_parent_main_image_url=selected_parent_main_image_url,
    )

    if not parent_main_image_url:
        raise ValueError("No staged mapped image exists for the selected variants.")

    staged_resource_urls: list[str] = []
    for path in staged_resource_images:
        try:
            url = dropbox_preview_url(path)
        except Exception:
            continue
        if url:
            staged_resource_urls.append(url)

    fallback_resource_urls: list[str] = []
    if use_resource_fallback_images and not staged_resource_urls:
        for path in list(garment_resource_images) + list(shared_resource_images):
            try:
                url = dropbox_preview_url(path)
            except Exception:
                continue
            if url:
                fallback_resource_urls.append(url)

    other_images = list(dict.fromkeys(staged_resource_urls or fallback_resource_urls))

    return (
        parent_main_image_url,
        other_images,
        color_image_map,
        design_color_image_url_map,
    )

def render_path_grid(
    title: str,
    entries: list[dict[str, Any]],
    cols_per_row: int = 5,
    image_width: int = 150,
) -> None:
    st.markdown(f"**{title}**")
    if not entries:
        st.caption("No files configured.")
        return

    cols = st.columns(cols_per_row)
    for idx, entry in enumerate(entries):
        with cols[idx % cols_per_row]:
            path = entry.get("path", "")
            st.caption(entry.get("label", Path(path).name if path else ""))
            if entry.get("exists") and entry.get("direct_url"):
                st.image(entry["direct_url"], width=image_width)
            else:
                st.warning("Not found")
                st.code(path, language=None)


def render_color_grid(
    entries: list[dict[str, Any]],
    cols_per_row: int = 5,
    image_width: int = 150,
) -> None:
    st.markdown("**Colour image mapping**")
    if not entries:
        st.caption("No colours configured.")
        return

    cols = st.columns(cols_per_row)
    for idx, entry in enumerate(entries):
        with cols[idx % cols_per_row]:
            label = entry.get("label", "")
            path = entry.get("path", "")
            st.caption(label)
            if not path:
                st.warning("Missing")
                continue

            if entry.get("exists") and entry.get("direct_url"):
                st.image(entry["direct_url"], width=image_width)
            else:
                st.warning("Not found")
                st.code(path, language=None)


def render_active_product_context(
    active_staged_folder_name: str,
    active_template_label: str,
    selected_parent_main_label: str,
    preview_parent_main_image_url: str,
    preview_color_image_map: dict[str, str],
    preview_design_color_image_url_map: dict[str, dict[str, str]],
    preview_other_images: list[str],
    image_mapping_status: str = "not_loaded",
    image_mapping_detail: str = "",
) -> None:
    st.subheader("Active product context")
    col1, col2 = st.columns(2)
    with col1:
        st.write(f"Staged folder: `{active_staged_folder_name or '-'}`")
        st.write(f"Template: `{active_template_label or '-'}`")
        st.write(f"Parent main image choice: `{selected_parent_main_label or 'Automatic (recommended)'}`")
    with col2:
        if image_mapping_status == "loaded":
            parent_status = "Resolved" if preview_parent_main_image_url else "Loaded but unresolved"
            support_count = len(preview_other_images)
            child_count = len(preview_color_image_map) + sum(
                len(design_map) for design_map in preview_design_color_image_url_map.values()
            )
            st.write("Image mappings: `Loaded`")
            st.write(f"Parent main image: `{parent_status}`")
            st.write(f"Child image mappings: `{child_count}`")
            st.write(f"Support images: `{support_count}`")
        elif image_mapping_status == "error":
            st.write("Image mappings: `Missing/errors`")
            st.caption(image_mapping_detail or "Image mappings could not be resolved.")
        else:
            st.write("Image mappings: `Not loaded yet`")
            st.caption(image_mapping_detail or "Load image mappings to resolve parent, child, and support image URLs.")


def trim_search_terms(value: str, max_bytes: int = 249) -> str:
    value = (value or "").strip()
    if not value:
        return ""

    # Split on commas first, since Amazon search terms are usually entered that way sasd
    terms = [term.strip() for term in value.split(",") if term.strip()]

    result_terms: list[str] = []
    current = ""

    for term in terms:
        candidate = term if not current else f"{current}, {term}"

        if len(candidate.encode("utf-8")) <= max_bytes:
            current = candidate
            result_terms.append(term)
        else:
            break

    return current.rstrip(" ,;")




def build_size_price_inputs(
    sizes: list[str],
    saved_prices: dict[str, float] | None = None,
    profile: dict[str, Any] | None = None,
    selected_variants: dict[str, list[str]] | None = None,
) -> dict[str, float]:
    selected_variants = selected_variants or {}
    if profile and has_design_size_pricing(profile, selected_variants):
        return build_design_size_price_inputs(profile, selected_variants, saved_prices)

    st.markdown("**Price by size**")

    # An empty selected-size list can mean either:
    # 1. this is a true One Size template, or
    # 2. the template supports sizes but none have been selected yet.
    #
    # Detect a real size dimension before applying the One Size fallback.
    selected_variant_keys = {
        str(key).strip().lower()
        for key in selected_variants
    }

    variation_theme = str(
        profile.get("variation_theme")
        or profile.get("variation-theme")
        or profile.get("variationTheme")
        or ""
    ).strip().lower()

    has_size_dimension = (
        "size" in selected_variant_keys
        or "sizes" in selected_variant_keys
        or "size" in variation_theme
    )

    if not sizes:
        if has_size_dimension:
            st.caption("Select at least one size to configure pricing.")
            return {}

        # True One Size templates, such as RL100, still need a price.
        sizes = ["One Size"]

    saved_prices = dict(saved_prices or {})
    profile = profile or {}

    existing_values = [
        saved_prices.get(size)
        for size in sizes
        if size in saved_prices
    ]
    unique_existing_values = {
        float(value)
        for value in existing_values
        if value is not None
    }
    default_same_price = (
        len(unique_existing_values) == 1
        and len(existing_values) == len(sizes)
    )

    available_clusters: dict[str, list[str]] = {}

    for size in sizes:
        cluster_label = get_design_size_price_cluster_label("", size)
        available_clusters.setdefault(cluster_label, []).append(size)

    has_multiple_clusters = len(available_clusters) >= 2

    pricing_modes = ["Use one price for all"]

    if has_multiple_clusters:
        pricing_modes.append("Use one price per cluster")

    pricing_modes.append("Manual price by size")

    legacy_mode_names = {
        "One price for all": "Use one price for all",
        "One price per cluster": "Use one price per cluster",
        "Manual by size": "Manual price by size",
    }

    if "size_pricing_mode" not in st.session_state:
        st.session_state["size_pricing_mode"] = (
            "Use one price for all"
            if st.session_state.get(
                "use_same_price_for_all_sizes",
                default_same_price,
            )
            else "Manual price by size"
        )

    current_mode = st.session_state.get("size_pricing_mode")
    current_mode = legacy_mode_names.get(current_mode, current_mode)

    if (
        current_mode == "Use one price per cluster"
        and not has_multiple_clusters
    ):
        current_mode = "Use one price for all"

    if current_mode not in pricing_modes:
        current_mode = "Manual price by size"

    st.session_state["size_pricing_mode"] = current_mode

    pricing_mode = st.radio(
        "Pricing mode",
        pricing_modes,
        horizontal=True,
        key="size_pricing_mode",
    )

    use_same_price = pricing_mode == "Use one price for all"
    st.session_state["use_same_price_for_all_sizes"] = use_same_price

    size_price_map: dict[str, float] = {}

    if use_same_price:
        fallback_price = get_default_price_for_size(
            profile,
            sizes[0],
            saved_prices,
        )

        if default_same_price:
            fallback_price = float(saved_prices.get(sizes[0], fallback_price))

        if "shared_price_all_sizes" not in st.session_state:
            st.session_state["shared_price_all_sizes"] = float(fallback_price)

        shared_price = st.number_input(
            "Price for all sizes",
            min_value=0.0,
            step=0.50,
            key="shared_price_all_sizes",
        )

        for size in sizes:
            size_price_map[size] = shared_price

        return size_price_map

    if pricing_mode == "Use one price per cluster":
        cols_per_row = 3
        cols = st.columns(cols_per_row)

        for idx, (cluster_label, cluster_sizes) in enumerate(
            available_clusters.items()
        ):
            saved_cluster_price = None

            for size in cluster_sizes:
                try:
                    candidate = float(saved_prices.get(size, 0))
                except (TypeError, ValueError):
                    candidate = 0

                if candidate > 0:
                    saved_cluster_price = candidate
                    break

            fallback_price = (
                saved_cluster_price
                if saved_cluster_price is not None
                else get_default_price_for_size(
                    profile,
                    cluster_sizes[0],
                    saved_prices,
                )
            )

            cluster_key = (
                "size_cluster_price_"
                f"{sanitize_sku(cluster_label)}"
            )

            with cols[idx % cols_per_row]:
                if cluster_key not in st.session_state:
                    st.session_state[cluster_key] = float(fallback_price)

                cluster_price = st.number_input(
                    f"{cluster_label} price",
                    min_value=0.0,
                    step=0.50,
                    key=cluster_key,
                )

                st.caption(", ".join(cluster_sizes))

            for size in cluster_sizes:
                size_price_map[size] = cluster_price

        return size_price_map

    if len(sizes) == 1:
        size = sizes[0]
        widget_key = f"price_{size}"

        if widget_key not in st.session_state:
            st.session_state[widget_key] = get_default_price_for_size(
                profile,
                size,
                saved_prices,
            )

        size_price_map[size] = st.number_input(
            f"{size} price",
            min_value=0.0,
            step=0.50,
            key=widget_key,
        )

        return size_price_map

    cols_per_row = 4
    cols = st.columns(cols_per_row)

    for idx, size in enumerate(sizes):
        widget_key = f"price_{size}"

        with cols[idx % cols_per_row]:
            if widget_key not in st.session_state:
                st.session_state[widget_key] = get_default_price_for_size(
                    profile,
                    size,
                    saved_prices,
                )

            size_price_map[size] = st.number_input(
                f"{size} price",
                min_value=0.0,
                step=0.50,
                key=widget_key,
            )

    return size_price_map

def build_design_size_price_inputs(
    profile: dict[str, Any],
    selected_variants: dict[str, list[str]],
    saved_prices: dict[str, float] | None = None,
) -> dict[str, float]:
    st.markdown("**Price by garment and size**")
    selected_designs = list(selected_variants.get("design", []) or [])
    selected_sizes = list(selected_variants.get("size", []) or [])
    saved_prices = dict(saved_prices or {})

    if not selected_designs or not selected_sizes:
        st.caption("Select at least one garment and one size to price variants.")
        return {}

    variant_combos = build_variant_combinations(profile, {
        "design": selected_designs,
        "size": selected_sizes,
    })
    if not variant_combos:
        st.caption("No valid garment/size combinations selected.")
        return {}

    default_same_price = False
    existing_values = [
        saved_prices.get(build_variant_price_key(combo))
        for combo in variant_combos
        if build_variant_price_key(combo) in saved_prices
    ]
    unique_existing_values = {v for v in existing_values if v is not None}
    if len(unique_existing_values) == 1 and len(existing_values) == len(variant_combos):
        default_same_price = True

    pricing_modes = [
        "Use one price for all",
        "Use one price per cluster",
        "Manual price by garment/size",
    ]
    if "design_size_pricing_mode" not in st.session_state:
        st.session_state["design_size_pricing_mode"] = (
            "Use one price for all"
            if st.session_state.get("use_same_price_for_all_sizes", default_same_price)
            else "Manual price by garment/size"
        )
    legacy_pricing_modes = {
        "One price for all": "Use one price for all",
        "One price per cluster": "Use one price per cluster",
        "Manual by garment/size": "Manual price by garment/size",
    }
    current_pricing_mode = st.session_state.get("design_size_pricing_mode")
    st.session_state["design_size_pricing_mode"] = legacy_pricing_modes.get(
        current_pricing_mode,
        current_pricing_mode if current_pricing_mode in pricing_modes else "Manual price by garment/size",
    )

    pricing_mode = st.radio(
        "Pricing mode",
        pricing_modes,
        horizontal=True,
        key="design_size_pricing_mode",
    )
    use_same_price = pricing_mode == "Use one price for all"
    st.session_state["use_same_price_for_all_sizes"] = use_same_price

    size_price_map: dict[str, float] = {}
    if use_same_price:
        first_combo = variant_combos[0]
        first_key = build_variant_price_key(first_combo)
        fallback_price = (
            get_positive_variant_price_from_map(profile, saved_prices, first_combo)
            or float(get_default_price_for_size(
                profile,
                first_combo.get("size", ""),
                {},
                design=first_combo.get("design", ""),
            ))
        )
        if "shared_price_all_sizes" not in st.session_state:
            st.session_state["shared_price_all_sizes"] = float(fallback_price)

        shared_price = st.number_input(
            "Price for all garment/size combinations",
            min_value=0.0,
            step=0.50,
            key="shared_price_all_sizes",
        )
        for combo in variant_combos:
            size_price_map[build_variant_price_key(combo)] = shared_price
        return size_price_map

    design_size_map = profile.get("design_size_map", {})
    if pricing_mode == "Use one price per cluster":
        for design in selected_designs:
            valid_design_sizes = [
                size
                for size in selected_sizes
                if not design_size_map.get(design) or size in design_size_map.get(design, [])
            ]
            if not valid_design_sizes:
                continue

            clusters: dict[str, list[str]] = {}
            for size in valid_design_sizes:
                cluster_label = get_design_size_price_cluster_label(design, size)
                clusters.setdefault(cluster_label, []).append(size)

            with st.expander(f"{design} cluster prices", expanded=True):
                cols_per_row = 3
                cols = st.columns(cols_per_row)
                for idx, (cluster_label, cluster_sizes) in enumerate(clusters.items()):
                    first_size = cluster_sizes[0]
                    first_combo = {"design": design, "size": first_size}
                    cluster_key = f"cluster_price_{sanitize_sku(design)}_{sanitize_sku(cluster_label)}"
                    fallback_price = (
                        get_positive_variant_price_from_map(profile, saved_prices, first_combo)
                        or float(get_default_price_for_size(
                            profile,
                            first_size,
                            {},
                            design=design,
                        ))
                    )
                    with cols[idx % cols_per_row]:
                        if cluster_key not in st.session_state:
                            st.session_state[cluster_key] = fallback_price
                        cluster_price = st.number_input(
                            f"{cluster_label} price",
                            min_value=0.0,
                            step=0.50,
                            key=cluster_key,
                        )
                        st.caption(", ".join(cluster_sizes))

                    for size in cluster_sizes:
                        size_price_map[build_variant_price_key({"design": design, "size": size})] = cluster_price

        return size_price_map

    for design in selected_designs:
        valid_design_sizes = [
            size
            for size in selected_sizes
            if not design_size_map.get(design) or size in design_size_map.get(design, [])
        ]
        if not valid_design_sizes:
            continue

        with st.expander(f"{design} prices", expanded=True):
            cols_per_row = 4
            cols = st.columns(cols_per_row)
            for idx, size in enumerate(valid_design_sizes):
                combo = {"design": design, "size": size}
                price_key = build_variant_price_key(combo)
                widget_key = f"price_{sanitize_sku(price_key)}"
                with cols[idx % cols_per_row]:
                    if widget_key not in st.session_state:
                        default_price = (
                            get_positive_variant_price_from_map(profile, saved_prices, combo)
                            or float(get_default_price_for_size(
                                profile,
                                size,
                                {},
                                design=design,
                            ))
                        )
                        st.session_state[widget_key] = float(
                            default_price
                        )
                    size_price_map[price_key] = st.number_input(
                        f"{size} price",
                        min_value=0.0,
                        step=0.50,
                        key=widget_key,
                    )

    return size_price_map


def get_design_size_price_cluster_label(design: str, size: str) -> str:
    design_normalized = str(design or "").strip().lower()
    size_normalized = str(size or "").strip().upper()
    if "kid" in design_normalized or is_child_size_label(size):
        return "Kids"
    if size_normalized in {"3XL", "4XL", "5XL", "6XL", "7XL", "8XL"}:
        return "Adult 3XL+"
    return "Adult XS-2XL"


def is_child_size_label(size: str) -> bool:
    normalized = str(size or "").strip().lower()
    return (
        "year" in normalized
        or "yrs" in normalized
        or normalized.startswith("child ")
        or bool(re.fullmatch(r"\d+\s*yrs?", normalized))
    )


def has_mixed_adult_child_sizes(sizes: list[str]) -> bool:
    return any(is_child_size_label(size) for size in sizes) and any(
        not is_child_size_label(size) for size in sizes
    )


def get_configured_price(
    profile: dict[str, Any],
    size: str,
    design: str = "",
) -> float | None:
    exact_map = profile.get("default_size_price_map", {})
    if isinstance(exact_map, dict):
        exact_keys = [build_variant_price_key({"design": design, "size": size})]
        if size not in exact_keys:
            exact_keys.append(size)
        for exact_key in exact_keys:
            if exact_key in exact_map:
                try:
                    return float(exact_map[exact_key])
                except (TypeError, ValueError):
                    return None

    child_price = profile.get("child_default_price", profile.get("kids_default_price"))
    adult_price = profile.get("adult_default_price")

    try:
        if is_child_size_label(size) and child_price is not None:
            return float(child_price)
        if not is_child_size_label(size) and adult_price is not None:
            return float(adult_price)
    except (TypeError, ValueError):
        return None

    return None


def get_default_price_for_size(
    profile: dict[str, Any],
    size: str,
    saved_prices: dict[str, float] | None = None,
    design: str = "",
) -> float:
    saved_prices = saved_prices or {}
    saved_keys = [build_variant_price_key({"design": design, "size": size})]
    if size not in saved_keys:
        saved_keys.append(size)
    for saved_key in saved_keys:
        if saved_key in saved_prices:
            try:
                return float(saved_prices[saved_key])
            except (TypeError, ValueError):
                pass

    configured_price = get_configured_price(profile, size, design)
    if configured_price is not None:
        return configured_price

    profile_sizes = list(profile.get("sizes", []) or [])
    if has_mixed_adult_child_sizes(profile_sizes):
        return 8.99 if is_child_size_label(size) else 12.99

    return 29.99

def resolve_template_path(profile: dict[str, Any]) -> Path:
    family_folder = profile.get("_family_folder")
    profile_folder = profile.get("_folder")
    template_file = profile.get("template_file", "")

    if family_folder:
        return (Path(family_folder) / template_file).resolve()
    if profile_folder:
        return (Path(profile_folder) / template_file).resolve()
    return (BASE_DIR / "templates" / template_file).resolve()


def sku_contains_code(value: str, code: str) -> bool:
    value_parts = [
        part.lower()
        for part in sanitize_sku(str(value or "")).split("-")
        if part
    ]
    code = sanitize_sku(str(code or "")).lower()
    return bool(code and code in value_parts)


def build_output_workbook_name(profile: dict[str, Any], parent_sku: str) -> str:
    parent_sku = sanitize_sku(str(parent_sku or "listing"))
    profile_slug = sanitize_sku(str(profile.get("_slug", "") or profile.get("template_key", "")))
    template_key = sanitize_sku(str(profile.get("template_key", "") or ""))

    if profile_slug and not sku_contains_code(parent_sku, profile_slug):
        return f"{parent_sku}_{profile_slug}_amazon_listing.xlsm"
    if template_key and not sku_contains_code(parent_sku, template_key):
        return f"{parent_sku}_{template_key}_amazon_listing.xlsm"
    return f"{parent_sku}_amazon_listing.xlsm"


def write_parent_row(ws, header_map: dict[str, int], data: dict[str, Any], parent_row: int = PARENT_ROW) -> None:
    clear_row_values(ws, parent_row)
    other_images = padded_list(data.get("other_images", []), 14)

    variation_theme = data.get("variation_theme", "")
    product_category = data.get("product_category", "apparel")
    is_apparel = product_category == "apparel"
    update_delete_value = "PartialUpdate" if str(data.get("original_finished_folder_name", "") or "").strip() else ""
    has_size = "Size" in variation_theme
    parent_starting_price = ""
    if data.get("write_parent_starting_price", False):
        parent_starting_price = get_lowest_variant_price(data.get("size_price_map", {}))

    values = {
        "item_sku": data["parent_sku"],
        "update_delete": update_delete_value,
        "parent_sku": "",
        "item_name": data["title"],
        "brand_name": data["brand_name"],
        "manufacturer": data["manufacturer"],
        "model_name": get_product_model_name({}, data),
        "model": get_garment_model_number({}, data),
        "part_number": data["parent_sku"],
        "product_description": data["product_description"],
        "generic_keywords": data["generic_keywords"],
        "bullet_point1": data["bullet_points"][0],
        "bullet_point2": data["bullet_points"][1],
        "bullet_point3": data["bullet_points"][2],
        "bullet_point4": data["bullet_points"][3],
        "bullet_point5": data["bullet_points"][4],
        "recommended_browse_nodes": data["recommended_browse_nodes"],

        "parent_child": "parent",
        "relationship_type": "",
        "variation_theme": variation_theme,

        "department_name": data["department_name"],
        "feed_product_type": data["feed_product_type"],
        "target_gender": data["target_gender"],
        "age_range_description": data["age_range_description"],

        "outer_material_type": data["material_type"],
        "material_type1": data["material_type"],
        "fabric_type": data["material_type"],

        "style_name": data["style_name"],
        "care_instructions": data["care_instructions"],
        "collar_style": data.get("collar_style", "Crew Neck"),
        "neck_style": data.get("neck_style", data.get("collar_style", "Crew Neck")),
        "theme": data["theme"],

        "color_name": "",
        "size_name": "",

        "apparel_size_system": "",
        "apparel_size_class": "",
        "apparel_size": "",
        "apparel_body_type": "",
        "apparel_height_type": "",
        "shirt_size_system": "",
        "shirt_size_class": "",
        "shirt_size": "",
        "shirt_size_to": "",
        "shirt_body_type": "",
        "shirt_height_type": "",

        "item_type_name": data["item_type_name"],
        "country_of_origin": data.get("country_of_origin", "United Kingdom"),
        "condition_type": data["condition_type"],

        "fulfillment_availability#1.fulfillment_channel_code": "",
        "fulfillment_availability#1.quantity": "",
        "fulfillment_availability#1.lead_time_to_ship_max_days": "",
        "purchasable_offer[marketplace_id=A1F83G8C2ARO7P]#1.our_price#1.schedule#1.value_with_tax": parent_starting_price,
        "list_price_with_tax": parent_starting_price,

        "main_image_url": data.get("parent_main_image_url", ""),

        "other_image_url1": other_images[0],
        "other_image_url2": other_images[1],
        "other_image_url3": other_images[2],
        "other_image_url4": other_images[3],
        "other_image_url5": other_images[4],
        "other_image_url6": other_images[5],
        "other_image_url7": other_images[6],
        "other_image_url8": other_images[7],

        "other_image_url_ps01": other_images[8],
        "other_image_url_ps02": other_images[9],
        "other_image_url_ps03": other_images[10],
        "other_image_url_ps04": other_images[11],
        "other_image_url_ps05": other_images[12],
        "other_image_url_ps06": other_images[13],
    }

    field_aliases = data.get("field_aliases", {})

    dynamic_profile_fields = data.get("dynamic_profile_fields", {})
    values.update(dynamic_profile_fields)

    extra_parent_fields = data.get("extra_parent_fields", {})
    values = prepare_row_values(values, field_aliases, extra_parent_fields)
    values["item_sku"] = data["parent_sku"]
    values["update_delete"] = update_delete_value
    values["parent_sku"] = ""
    values["part_number"] = data["parent_sku"]
    values["model_name"] = get_product_model_name({}, data)
    values["model"] = get_garment_model_number({}, data)
    values["parent_child"] = "parent"
    values["relationship_type"] = ""
    values["variation_theme"] = variation_theme
    for parent_size_field in [
        "apparel_size_system",
        "apparel_size_class",
        "apparel_size",
        "apparel_body_type",
        "apparel_height_type",
        "shirt_size_system",
        "shirt_size_class",
        "shirt_size",
        "shirt_size_to",
        "shirt_body_type",
        "shirt_height_type",
    ]:
        values[parent_size_field] = ""

    if "dangerous_goods_regulation" in header_map:
        values["dangerous_goods_regulation"] = "Not Applicable"

    if "search_terms" in header_map:
        values["search_terms"] = trim_search_terms(data["generic_keywords"])

    write_values_with_debug(ws, parent_row, header_map, values, "Parent row")

def write_child_rows(
    ws,
    header_map: dict[str, int],
    profile: dict[str, Any],
    data: dict[str, Any],
    first_child_row: int = FIRST_CHILD_ROW,
) -> int:
    row_idx = first_child_row
    template_row = first_child_row
    variants_written = 0
    other_images = padded_list(data.get("other_images", []), 14)
    sku_profile = apply_sku_context_to_profile(
        profile,
        data.get("sku_decoration_code", ""),
        data.get("sku_listing_code", ""),
    )

    variation_theme = data.get("variation_theme", "")
    product_category = data.get("product_category", "apparel")
    is_apparel = product_category == "apparel"
    update_delete_value = "PartialUpdate" if str(data.get("original_finished_folder_name", "") or "").strip() else ""

    selected_variants = data.get("selected_variants", {})
    variant_combos = build_variant_combinations(profile, selected_variants)
    variant_combos = sort_variant_combinations_by_price(profile, variant_combos, data.get("size_price_map", {}))

    for variant_values in variant_combos:
        if row_idx != template_row and st.session_state.get("copy_row_styles", True):
            copy_row_format(ws, template_row, row_idx)
        clear_row_values(ws, row_idx)

        variant_field_values = build_variant_field_values(profile, variant_values)
        sku_details = build_child_sku_details(sku_profile, data["parent_sku"], variant_values)
        item_sku = sku_details["amazon_seller_sku"]

        size_value = variant_values.get("size", "")
        normalized_size = normalize_size(size_value) if size_value else ""
        display_size = get_variant_size_display_label(profile, variant_values)
        normalized_display_size = normalize_size(display_size) if display_size else ""
        design_value = variant_values.get("design", "")
        color_value = variant_values.get("color", "")
        display_color = get_variant_color_display_label(profile, variant_values)
        child_item_name = build_child_item_name(data["title"], variant_values, profile)

        price = get_variant_price_from_map(profile, data.get("size_price_map", {}), variant_values)

        image_url = resolve_child_variant_image_url(
            variant_values=variant_values,
            color_image_map=data.get("color_image_map", {}),
            design_color_image_url_map=data.get("design_color_image_url_map", {}),
        )

        values = {
            "item_sku": item_sku,
            "update_delete": update_delete_value,
            "parent_sku": data["parent_sku"],
            "item_name": child_item_name,
            "brand_name": data["brand_name"],
            "manufacturer": data["manufacturer"],
            "model_name": get_product_model_name(profile, data),
            "model": get_child_model_number(profile, data, sku_details),
            "part_number": item_sku,
            "product_description": data["product_description"],
            "generic_keywords": data["generic_keywords"],

            "bullet_point1": data["bullet_points"][0],
            "bullet_point2": data["bullet_points"][1],
            "bullet_point3": data["bullet_points"][2],
            "bullet_point4": data["bullet_points"][3],
            "bullet_point5": data["bullet_points"][4],

            "recommended_browse_nodes": data["recommended_browse_nodes"],
            "condition_type": data["condition_type"],

            "parent_child": "child",
            "relationship_type": "variation",
            "variation_theme": variation_theme,

            "color_name": display_color,
            "size_name": normalized_display_size,
            "apparel_size": normalized_display_size if is_apparel else "",

            "department_name": get_row_department_name(profile, data, normalized_size),
            "feed_product_type": data["feed_product_type"],
            "target_gender": data["target_gender"],
            "age_range_description": get_row_age_range_description(profile, data, normalized_size),

            "outer_material_type": data["material_type"],
            "material_type1": data["material_type"],
            "fabric_type": data["material_type"],

            "style_name": design_value or data["style_name"],
            "care_instructions": data["care_instructions"],
            "collar_style": data.get("collar_style", "Crew Neck"),
            "neck_style": data.get("neck_style", data.get("collar_style", "Crew Neck")),
            "theme": data["theme"],

            "apparel_size_system": "UK" if normalized_size and is_apparel else "",
            "apparel_size_class": "Alpha" if normalized_size and is_apparel else "",
            "apparel_body_type": "Regular" if is_apparel else "",
            "apparel_height_type": "",

            "item_type_name": data["item_type_name"],
            "country_of_origin": data.get("country_of_origin", "United Kingdom"),

            "fulfillment_availability#1.fulfillment_channel_code": "DEFAULT",
            "fulfillment_availability#1.quantity": data["quantity"],
            "fulfillment_availability#1.lead_time_to_ship_max_days": normalize_handling_time_days(
                data.get("handling_time_days", DEFAULT_HANDLING_TIME_DAYS)
            ),
            "merchant_shipping_group_name": normalize_merchant_shipping_group(
                data.get("merchant_shipping_group_name", "")
            ),
            "purchasable_offer[marketplace_id=A1F83G8C2ARO7P]#1.our_price#1.schedule#1.value_with_tax": price,
            "list_price_with_tax": price,

            "main_image_url": image_url,

            "other_image_url1": other_images[0],
            "other_image_url2": other_images[1],
            "other_image_url3": other_images[2],
            "other_image_url4": other_images[3],
            "other_image_url5": other_images[4],
            "other_image_url6": other_images[5],
            "other_image_url7": other_images[6],
            "other_image_url8": other_images[7],

            "other_image_url_ps01": other_images[8],
            "other_image_url_ps02": other_images[9],
            "other_image_url_ps03": other_images[10],
            "other_image_url_ps04": other_images[11],
            "other_image_url_ps05": other_images[12],
            "other_image_url_ps06": other_images[13],
        }

        dynamic_profile_fields = data.get("dynamic_profile_fields", {})
        values.update(dynamic_profile_fields)

        values.update(variant_field_values)
        values.update(get_variant_design_overrides(profile, variant_values))

        field_aliases = data.get("field_aliases", {})
        extra_child_fields = merge_extra_child_fields_for_variant(
            data.get("extra_child_fields", {}),
            data.get("extra_child_fields_by_size", {}),
            variant_values,
        )
        values = prepare_row_values(values, field_aliases, extra_child_fields)
        values = apply_apparel_size_fields(
            values,
            normalized_size,
            is_apparel=is_apparel,
            profile=profile,
        )
        values = expand_field_aliases(values, field_aliases)
        values["size_name"] = normalized_display_size
        values["item_sku"] = item_sku
        values["update_delete"] = update_delete_value
        values["parent_sku"] = data["parent_sku"]
        values["item_name"] = child_item_name
        values["part_number"] = item_sku
        values["model_name"] = get_product_model_name(profile, data)
        values["model"] = get_child_model_number(profile, data, sku_details)
        values["color_name"] = display_color
        values["department_name"] = get_row_department_name(profile, data, normalized_size)
        values["age_range_description"] = get_row_age_range_description(profile, data, normalized_size)
        values["parent_child"] = "child"
        values["relationship_type"] = "variation"
        values["variation_theme"] = variation_theme

        if st.session_state.get("show_header_debug", False) and row_idx == FIRST_CHILD_ROW:
            st.write("First child size values snapshot")
            st.json({
                "apparel_size_system": values.get("apparel_size_system"),
                "apparel_size_class": values.get("apparel_size_class"),
                "apparel_body_type": values.get("apparel_body_type"),
                "apparel_height_type": values.get("apparel_height_type"),
                "field_aliases": field_aliases,
            })

        if "dangerous_goods_regulation" in header_map:
            values["dangerous_goods_regulation"] = "Not Applicable"

        if "search_terms" in header_map:
            values["search_terms"] = trim_search_terms(data["generic_keywords"])

        label_bits = [v for v in [color_value, size_value, design_value] if v]
        row_label = " / ".join(label_bits) if label_bits else f"row {row_idx}"

        write_values_with_debug(
            ws,
            row_idx,
            header_map,
            values,
            f"Child row {row_idx} ({row_label})",
        )

        row_idx += 1
        variants_written += 1

    return variants_written

def get_extra_parent_fields(profile: dict[str, Any]) -> dict[str, Any]:
    return profile.get("extra_parent_fields", profile.get("extra_fields", {}))

def get_extra_child_fields(profile: dict[str, Any]) -> dict[str, Any]:
    return profile.get("extra_child_fields", profile.get("extra_fields", {}))


def get_extra_child_fields_by_size(profile: dict[str, Any]) -> dict[str, dict[str, Any]]:
    raw_map = profile.get("extra_child_fields_by_size", {})
    if not isinstance(raw_map, dict):
        return {}
    return {
        str(size): dict(fields)
        for size, fields in raw_map.items()
        if isinstance(fields, dict)
    }


def merge_extra_child_fields_for_variant(
    base_fields: dict[str, Any],
    size_fields_map: dict[str, dict[str, Any]],
    variant_values: dict[str, str],
) -> dict[str, Any]:
    merged = dict(base_fields or {})
    size_value = str(variant_values.get("size", "") or "")
    if size_value and size_value in size_fields_map:
        merged.update(size_fields_map[size_value])
    return merged

DEFAULT_FIELD_ALIASES = {
    "apparel_size_system": [
        "shirt_size_system",
        "tops_size_system",
        "outerwear_size_system",
    ],
    "apparel_size_class": [
        "shirt_size_class",
        "tops_size_class",
        "outerwear_size_class",
    ],
    "apparel_size": [
        "shirt_size",
        "tops_size_value",
        "outerwear_size_value",
    ],
    "apparel_body_type": [
        "shirt_body_type",
        "tops_body_type",
        "outerwear_body_type",
    ],
    "apparel_height_type": [
        "shirt_height_type",
        "tops_height_type",
        "outerwear_height_type",
    ],
}


def get_field_aliases(profile: dict[str, Any]) -> dict[str, list[str]]:
    aliases = {field: list(values) for field, values in DEFAULT_FIELD_ALIASES.items()}
    for source_field, configured_aliases in profile.get("field_aliases", {}).items():
        merged_aliases = aliases.setdefault(source_field, [])
        for alias in configured_aliases:
            if alias not in merged_aliases:
                merged_aliases.append(alias)
    return aliases

def debug_find_headers(header_map: dict[str, int], patterns: list[str]) -> None:
    st.write("Header matches:")
    for pattern in patterns:
        matches = [key for key in header_map.keys() if pattern.lower() in key.lower()]
        st.write(f"{pattern}: {matches}")

def expand_field_aliases(values: dict[str, Any], field_aliases: dict[str, list[str]]) -> dict[str, Any]:
    expanded = dict(values)

    for source_field, aliases in field_aliases.items():
        if source_field in expanded:
            for alias in aliases:
                expanded[alias] = expanded[source_field]

    return expanded

def prepare_row_values(
    values: dict[str, Any],
    field_aliases: dict[str, list[str]],
    extra_fields: dict[str, Any],
) -> dict[str, Any]:
    prepared = dict(values)
    prepared.update(extra_fields)
    return expand_field_aliases(prepared, field_aliases)


def validate_profile_schema(profile: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    required_fields = get_required_profile_fields(profile)

    for field in required_fields:
        value = profile.get(field)
        if isinstance(value, str):
            if not value.strip():
                errors.append(f"Template config missing required field: {field}")
        elif value in (None, "", [], {}):
            errors.append(f"Template config missing required field: {field}")

    return errors

def get_schema(profile: dict[str, Any]) -> dict[str, Any]:
    return profile.get("_schema", {})

def get_allowed_dynamic_fields(profile: dict[str, Any]) -> list[str]:
    return get_schema(profile).get("allowed_dynamic_fields", [])

def get_required_profile_fields(profile: dict[str, Any]) -> list[str]:
    return get_schema(profile).get("required_profile_fields", [])

def get_required_workbook_headers(profile: dict[str, Any]) -> list[str]:
    return get_schema(profile).get("required_workbook_headers", [])

def validate_template_file(profile: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    template_file = profile.get("template_file", "")

    if not template_file:
        return ["Template file is missing in config."]

    template_path = resolve_template_path(profile)

    if not template_path.exists():
        return [f"Template file not found: {template_path}"]

    try:
        wb = load_workbook(template_path, keep_vba=True, read_only=True)
    except Exception as exc:
        return [f"Template workbook could not be opened: {exc}"]

    try:
        if SHEET_NAME not in wb.sheetnames:
            return [f"Sheet '{SHEET_NAME}' not found in template workbook."]

        ws = wb[SHEET_NAME]
        layout = get_workbook_layout(profile)
        header_map = build_header_map(ws, layout["header_row"])

        required_headers = get_required_workbook_headers(profile) or [
            "item_sku",
            "item_name",
            "brand_name",
            "manufacturer",
            "product_description",
            "bullet_point1",
            "bullet_point2",
            "bullet_point3",
            "bullet_point4",
            "bullet_point5",
            "generic_keywords",
            "recommended_browse_nodes",
            "parent_child",
            "relationship_type",
            "variation_theme",
            "condition_type",
            "main_image_url",
        ]

        missing_headers = [header for header in required_headers if header not in header_map]
        if missing_headers:
            errors.append("Template is missing required headers: " + ", ".join(missing_headers))
    finally:
        wb.close()

    return errors




def write_listing_rows_to_workbook(
    ws,
    header_map: dict[str, int],
    profile: dict[str, Any],
    payload: dict[str, Any],
    parent_row: int,
    child_start_row: int,
    parent_template_row: int | None = None,
    child_template_row: int | None = None,
) -> dict[str, int]:
    parent_template_row = parent_template_row or parent_row
    child_template_row = child_template_row or child_start_row

    if parent_row != parent_template_row and st.session_state.get("copy_row_styles", True):
        copy_row_format(ws, parent_template_row, parent_row)

    write_parent_row(ws, header_map, payload, parent_row=parent_row)

    if child_start_row != child_template_row and st.session_state.get("copy_row_styles", True):
        copy_row_format(ws, child_template_row, child_start_row)

    variants_written = write_child_rows(
        ws,
        header_map,
        profile,
        payload,
        first_child_row=child_start_row,
    )

    if variants_written <= 0:
        raise ValueError(f"No child variants were generated for parent SKU {payload.get('parent_sku', '')}.")

    next_row = child_start_row + variants_written
    return {
        "parent_row": parent_row,
        "child_start_row": child_start_row,
        "variants_written": variants_written,
        "next_row": next_row,
    }


def build_workbook(profile: dict[str, Any], payload: dict[str, Any]) -> tuple[Path, dict[str, float]]:

    template_path = resolve_template_path(profile)

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    stock_ready_report = validate_stock_ready_payload(profile, payload)
    if stock_ready_report.get("errors"):
        raise ValueError("Stock-ready SKU validation failed: " + "; ".join(stock_ready_report["errors"]))

    t0 = time.perf_counter()
    wb = load_workbook(template_path, keep_vba=True)
    t1 = time.perf_counter()

    ws = wb[SHEET_NAME]
    layout = get_workbook_layout(profile)
    header_map = build_header_map(ws, layout["header_row"])


    dynamic_profile_fields = get_dynamic_profile_fields(profile, header_map)

    if st.session_state.get("show_header_debug", False):
        allowed_fields = set(get_allowed_dynamic_fields(profile))
        missing_dynamic_headers = [
            key for key in allowed_fields
            if profile.get(key) not in (None, "", [], {}) and key not in header_map
        ]
        if missing_dynamic_headers:
            st.write("Allowed dynamic fields missing from workbook headers")
            st.json(missing_dynamic_headers)

    payload = dict(payload)
    payload["dynamic_profile_fields"] = dynamic_profile_fields

    if st.session_state.get("show_header_debug", False):
        st.write("Dynamic profile fields matched to workbook headers")
        st.json(dynamic_profile_fields)

    debug_size_headers(header_map)

    if st.session_state.get("show_header_debug", False):
        debug_find_headers(
            header_map,
            [
                "body",
                "height",
                "fulfillment",
                "quantity",
                "lead_time",
                "ship",
                "price",
                "value_with_tax",
            ],
        )

    if layout["mode"] == "offer_only":
        variants_written = write_child_rows(
            ws,
            header_map,
            profile,
            payload,
            first_child_row=layout["first_child_row"],
        )
    else:
        row_result = write_listing_rows_to_workbook(
            ws,
            header_map,
            profile,
            payload,
            parent_row=layout["parent_row"],
            child_start_row=layout["first_child_row"],
            parent_template_row=layout["parent_row"],
            child_template_row=layout["first_child_row"],
        )
        variants_written = row_result["variants_written"]
    t2 = time.perf_counter()
    t3 = t2

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_name = build_output_workbook_name(profile, payload["parent_sku"])
    output_path = OUTPUT_DIR / output_name
    wb.save(output_path)
    wb.close()
    t4 = time.perf_counter()

    if variants_written == 0:
        raise ValueError("No child variants were generated.")

    timings = {
        "load_workbook": t1 - t0,
        "write_listing_rows": t2 - t1,
        "write_parent_row": t2 - t1,
        "write_child_rows": t3 - t2,
        "save_workbook": t4 - t3,
        "total_build": t4 - t0,
    }

    return output_path, timings





def build_combined_output_workbook_name(profile: dict[str, Any], payloads: list[dict[str, Any]]) -> str:
    family_slug = sanitize_sku(str(profile.get("_family_slug", "") or profile.get("template_family", "") or "template"))
    workbook_stem = sanitize_sku(Path(str(profile.get("template_file", "") or "workbook")).stem)
    template_keys = sorted({
        sanitize_sku(str(payload.get("template_key", "") or payload.get("template_slug", "") or "listing"))
        for payload in payloads
    })
    template_part = "-".join([key for key in template_keys if key][:4]) or workbook_stem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{family_slug}_{workbook_stem}_{template_part}_combined_{len(payloads)}_listings_{timestamp}.xlsm"


def get_combined_workbook_group_identity(profile: dict[str, Any]) -> tuple[Any, ...]:
    layout = get_workbook_layout(profile)
    return (
        str(profile.get("_family_slug", "") or "").strip().lower(),
        str(resolve_template_path(profile)).strip().lower(),
        str(layout.get("mode", "") or "").strip(),
        int(layout.get("header_row", HEADER_ROW)),
        int(layout.get("parent_row", PARENT_ROW)),
        int(layout.get("first_child_row", FIRST_CHILD_ROW)),
    )


def build_combined_workbook_group_label(items: list[dict[str, Any]]) -> str:
    template_keys = sorted({
        str(item.get("profile", {}).get("template_key", "") or item.get("profile", {}).get("_slug", "") or "").strip()
        for item in items
        if item and item.get("profile")
    })
    return " + ".join([key for key in template_keys if key]) or "combined"


def get_generation_template_identity(profile: dict[str, Any]) -> tuple[Any, ...]:
    layout = get_workbook_layout(profile)
    return (
        str(profile.get("template_key", "") or "").strip(),
        str(profile.get("_slug", "") or "").strip(),
        str(profile.get("template_file", "") or "").strip(),
        str(layout.get("mode", "") or "").strip(),
        int(layout.get("header_row", HEADER_ROW)),
        int(layout.get("parent_row", PARENT_ROW)),
        int(layout.get("first_child_row", FIRST_CHILD_ROW)),
    )


def build_combined_workbook(
    profile: dict[str, Any],
    payloads: list[dict[str, Any]],
    payload_profiles: list[dict[str, Any]] | None = None,
) -> tuple[Path, dict[str, float]]:
    if len(payloads) < 2:
        raise ValueError("Select at least two approved listings to build one combined workbook.")

    payload_profiles = payload_profiles or [profile for _ in payloads]
    if len(payload_profiles) != len(payloads):
        raise ValueError("Internal error: combined workbook payload/profile count mismatch.")

    layout = get_workbook_layout(profile)
    if layout["mode"] == "offer_only":
        raise ValueError("Combined workbook generation is only for normal listing templates, not offer-only templates.")

    template_path = resolve_template_path(profile)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    for payload, payload_profile in zip(payloads, payload_profiles):
        quality_profile = apply_sku_context_to_profile(
            payload_profile,
            payload.get("sku_decoration_code", ""),
            payload.get("sku_listing_code", ""),
        )
        quality_report = validate_listing_quality(quality_profile, payload)
        blockers = list(quality_report.get("blockers", []))
        if blockers:
            raise ValueError(
                f"Score check failed for {payload.get('parent_sku', 'listing')}: "
                + "; ".join(blockers)
            )

        stock_ready_report = validate_stock_ready_payload(quality_profile, payload)
        if stock_ready_report.get("errors"):
            raise ValueError(
                f"Stock/SKU check failed for {payload.get('parent_sku', 'listing')}: "
                + "; ".join(stock_ready_report["errors"])
            )

    parent_skus = [str(payload.get("parent_sku", "") or "") for payload in payloads]
    duplicate_parent_skus = sorted([sku for sku, count in Counter(parent_skus).items() if sku and count > 1])
    if duplicate_parent_skus:
        raise ValueError("Duplicate parent SKUs in combined workbook: " + ", ".join(duplicate_parent_skus[:10]))

    child_skus: list[str] = []
    for payload, payload_profile in zip(payloads, payload_profiles):
        sku_profile = apply_sku_context_to_profile(
            payload_profile,
            payload.get("sku_decoration_code", ""),
            payload.get("sku_listing_code", ""),
        )
        for combo in build_variant_combinations(payload_profile, payload.get("selected_variants", {})):
            child_skus.append(build_child_sku_details(sku_profile, payload["parent_sku"], combo)["amazon_seller_sku"])

    duplicate_child_skus = sorted([sku for sku, count in Counter(child_skus).items() if sku and count > 1])
    if duplicate_child_skus:
        raise ValueError("Duplicate child SKUs in combined workbook: " + ", ".join(duplicate_child_skus[:10]))

    t0 = time.perf_counter()
    wb = load_workbook(template_path, keep_vba=True)
    t1 = time.perf_counter()

    ws = wb[SHEET_NAME]
    header_map = build_header_map(ws, layout["header_row"])
    current_parent_row = int(layout["parent_row"])
    parent_template_row = int(layout["parent_row"])
    child_template_row = int(layout["first_child_row"])
    total_children = 0

    for payload, payload_profile in zip(payloads, payload_profiles):
        payload = dict(payload)
        dynamic_profile_fields = get_dynamic_profile_fields(payload_profile, header_map)
        payload["dynamic_profile_fields"] = dynamic_profile_fields

        row_result = write_listing_rows_to_workbook(
            ws,
            header_map,
            payload_profile,
            payload,
            parent_row=current_parent_row,
            child_start_row=current_parent_row + 1,
            parent_template_row=parent_template_row,
            child_template_row=child_template_row,
        )
        total_children += row_result["variants_written"]
        current_parent_row = row_result["next_row"]

    t2 = time.perf_counter()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_name = build_combined_output_workbook_name(profile, payloads)
    output_path = OUTPUT_DIR / output_name
    wb.save(output_path)
    wb.close()
    t3 = time.perf_counter()

    return output_path, {
        "load_workbook": t1 - t0,
        "write_combined_rows": t2 - t1,
        "save_workbook": t3 - t2,
        "total_build": t3 - t0,
        "combined_listing_count": float(len(payloads)),
        "combined_child_count": float(total_children),
    }


def generate_approved_listings_combined(
    selected_items: list[dict[str, Any]],
    dropbox_cfg: dict[str, Any],
) -> dict[str, Any]:
    valid_items = [
        item for item in selected_items
        if item and item.get("profile") and item.get("listing_memory") and not item.get("load_error")
    ]
    if len(valid_items) < 2:
        raise ValueError("Select at least two approved listings for one combined workbook.")

    profile = valid_items[0]["profile"]

    identity = get_combined_workbook_group_identity(profile)
    mismatched = [
        item["folder_name"]
        for item in valid_items
        if get_combined_workbook_group_identity(item["profile"]) != identity
    ]
    if mismatched:
        raise ValueError("Selected listings must use the same workbook group: " + ", ".join(mismatched[:10]))

    prepared_payloads: list[dict[str, Any]] = []
    prepared_profiles: list[dict[str, Any]] = []

    for item in valid_items:
        folder_name = item["folder_name"]
        try:
            profile_i = item["profile"]
            listing_memory = item["listing_memory"]

            title = str(listing_memory.get("title", ""))
            bullets = (list(listing_memory.get("bullet_points", [])) + ["", "", "", "", ""])[:5]
            product_description = str(listing_memory.get("product_description", ""))
            generic_keywords = str(listing_memory.get("generic_keywords", ""))
            selected_variants = dict(listing_memory.get("selected_variants", {}))
            size_price_map = {
                str(size): float(price)
                for size, price in dict(listing_memory.get("size_price_map", {})).items()
            }

            prep = prepare_generation_payload(
                profile=profile_i,
                title=title,
                bullets=bullets,
                product_description=product_description,
                generic_keywords=generic_keywords,
                selected_variants=selected_variants,
                size_price_map=size_price_map,
                sku_decoration_code=get_default_sku_decoration_code(profile_i, listing_memory),
                sku_listing_code=str(listing_memory.get("sku_listing_code", "") or get_saved_generated_sku_listing_code(listing_memory)),
                manual_sku_listing_code=str(listing_memory.get("manual_sku_listing_code", "") or ""),
                generated_sku_listing_code=get_saved_generated_sku_listing_code(listing_memory),
                quantity=normalize_variant_quantity(listing_memory.get("quantity", DEFAULT_VARIANT_QUANTITY)),
                staged_folder_name=folder_name,
                handling_time_days=normalize_handling_time_days(listing_memory.get("handling_time_days", DEFAULT_HANDLING_TIME_DAYS)),
                merchant_shipping_group_name=normalize_merchant_shipping_group(listing_memory.get("merchant_shipping_group_name", "")),
                parent_main_image_choice=str(listing_memory.get("parent_main_image_choice", "") or ""),
                parent_main_image_url=str(listing_memory.get("parent_main_image_url", "") or ""),
                parent_sku_override=get_listing_generation_parent_sku_override(
                    listing_memory,
                    profile_i,
                ),
            )
            if prep.get("errors"):
                raise ValueError("; ".join(prep["errors"]))

            payload = dict(prep["payload"])
            selected_variants = dict(payload.get("selected_variants", {}))
            payload["assets_prepared_by"] = listing_memory.get("assets_prepared_by", "")
            payload["content_prepared_by"] = listing_memory.get("content_prepared_by", "")
            payload["reviewed_by"] = listing_memory.get("reviewed_by", "")
            payload["prepared_at"] = listing_memory.get("prepared_at", "")
            payload["reviewed_at"] = listing_memory.get("reviewed_at", "") or format_workflow_timestamp()
            if isinstance(listing_memory.get("workflow_events"), list):
                payload["workflow_events"] = list(listing_memory.get("workflow_events", []))
            preserve_grouped_child_generation_context(listing_memory, payload)

            approved_folder_path = build_approved_folder_path(dropbox_cfg, folder_name)
            dropbox_overview = get_cached_dropbox_overview(profile_i, dropbox_cfg)
            selected_parent_main_image_label = str(payload.get("parent_main_image_choice", "") or "")
            selected_parent_main_image_url = str(
                payload.get("selected_parent_main_image_url", "")
                or payload.get("parent_main_image_url", "")
                or ""
            )

            parent_main_image_url, other_images, color_image_map, design_color_image_url_map = resolve_folder_image_urls(
                profile_i,
                payload["selected_variants"],
                payload["colors"],
                dropbox_overview,
                approved_folder_path,
                selected_parent_main_image_label=selected_parent_main_image_label,
                selected_parent_main_image_url=selected_parent_main_image_url,
            )

            payload["parent_main_image_url"] = parent_main_image_url
            payload["other_images"] = other_images
            payload["color_image_map"] = color_image_map
            payload["design_color_image_url_map"] = design_color_image_url_map

            prepared_payloads.append(payload)
            prepared_profiles.append(profile_i)
        except Exception as exc:
            raise ValueError(f"{folder_name}: {exc}") from exc

    output_path, timings = build_combined_workbook(profile, prepared_payloads, prepared_profiles)
    if not output_path.exists() or output_path.stat().st_size <= 0:
        raise ValueError(f"Combined workbook was not created correctly: {output_path.name}")

    group_label = build_combined_workbook_group_label(valid_items)
    return {
        "folder_name": f"Combined workbook - {group_label}",
        "status": "Success",
        "message": f"Generated one grouped workbook for {len(prepared_payloads)} compatible listing(s): {output_path.name}",
        "output_path": str(output_path),
        "output_name": output_path.name,
        "timings": timings,
    }


def validate_payload(payload: dict[str, Any]) -> list[str]:
    errors: list[str] = []

    required_fields = [
        "parent_sku",
        "title",
        "brand_name",
        "manufacturer",
        "recommended_browse_nodes",
        "feed_product_type",
        "item_type_name",
        "material_type",
        "condition_type",
        "variation_theme",
        "product_category",
    ]

    for field in required_fields:
        value = payload.get(field, "")
        if isinstance(value, str):
            if not value.strip():
                errors.append(f"{field} is required.")
        elif value in (None, "", []):
            errors.append(f"{field} is required.")

    allowed_variation_themes = {"SizeColor","Colour & Style", ""}
    if payload.get("variation_theme", "") not in allowed_variation_themes:
        errors.append("variation_theme must be one of: SizeColor, Colour & Style, or empty.")

    allowed_product_categories = {"apparel", "accessory"}
    if payload.get("product_category", "") not in allowed_product_categories:
        errors.append("product_category must be either 'apparel' or 'accessory'.")

    return errors

def validate_variants(
    selected_variants: dict[str, list[str]],
    size_price_map: dict[str, float],
    quantity: int,
    profile: dict[str, Any] | None = None,
) -> list[str]:
    errors: list[str] = []
    profile = profile or {}

    for dim_name, values in selected_variants.items():
        if not values:
            errors.append(f"At least one option is required for {dim_name}.")

    if quantity <= 0:
        errors.append("Quantity must be at least 1.")

    variant_combos = build_variant_combinations(profile, selected_variants) if profile else []
    if variant_combos:
        invalid_labels = []
        for combo in variant_combos:
            price = get_variant_price_from_map(profile, size_price_map, combo, fallback=0)
            try:
                valid_price = float(price) > 0
            except (TypeError, ValueError):
                valid_price = False
            if not valid_price:
                invalid_labels.append(" / ".join([str(v) for v in combo.values() if v]) or "Unnamed variant")
        if invalid_labels:
            errors.append(f"Invalid or missing prices for variant(s): {', '.join(invalid_labels[:10])}")
    elif selected_variants.get("size", []):
        for size in selected_variants.get("size", []):
            if size_price_map.get(size, 0) <= 0:
                errors.append(f"Invalid price for size {size}.")
    else:
        if not size_price_map:
            errors.append("At least one price is required.")
        elif all(price <= 0 for price in size_price_map.values()):
            errors.append("At least one valid price is required.")

    return errors

def validate_parent_child_structure(payload: dict[str, Any]) -> list[str]:
    errors: list[str] = []

    if not payload.get("parent_sku", "").strip():
        errors.append("parent_sku is required.")

    allowed_variation_themes = {"SizeColor","Colour & Style", ""}
    if payload.get("variation_theme", "") not in allowed_variation_themes:
        errors.append("variation_theme must be one of: SizeColor, Colour & Style, or empty.")

    allowed_product_categories = {"apparel", "accessory"}
    if payload.get("product_category", "") not in allowed_product_categories:
        errors.append("product_category must be either 'apparel' or 'accessory'.")

    return errors


def validate_stock_ready_payload(profile: dict[str, Any], payload: dict[str, Any]) -> dict[str, Any]:
    profile = apply_sku_context_to_profile(
        profile,
        payload.get("sku_decoration_code", ""),
        payload.get("sku_listing_code", ""),
    )
    selected_variants = dict(payload.get("selected_variants", {}))
    parent_sku = str(payload.get("parent_sku", "") or "").strip()
    variant_combos = build_variant_combinations(profile, selected_variants)
    return validate_stock_ready_skus(profile, parent_sku or "PARENT", variant_combos)


def validate_variant_image_count(
    profile: dict[str, Any],
    payload: dict[str, Any],
    min_total: int = 4,
    max_total: int = 10,
) -> list[str]:
    selected_variants = dict(payload.get("selected_variants", {}))
    variant_combos = build_variant_combinations(profile, selected_variants)
    color_image_map = dict(payload.get("color_image_map", {}) or {})
    design_color_image_url_map = dict(payload.get("design_color_image_url_map", {}) or {})
    secondary_images = [
        image_url
        for image_url in list(payload.get("other_images", []) or [])
        if str(image_url or "").strip()
    ]

    errors: list[str] = []
    for variant_values in variant_combos:
        main_image_url = resolve_child_variant_image_url(
            variant_values=variant_values,
            color_image_map=color_image_map,
            design_color_image_url_map=design_color_image_url_map,
        )
        total_images = (1 if main_image_url else 0) + len(secondary_images)
        if total_images < min_total or total_images > max_total:
            label = " / ".join([str(v) for v in variant_values.values() if v]) or "Unnamed variant"
            errors.append(
                f"Variant '{label}' has {total_images} image(s); {min_total}-{max_total} are allowed before review."
            )

    return errors


def validate_child_title_lengths(
    profile: dict[str, Any],
    payload: dict[str, Any],
    max_chars: int = 200,
) -> list[str]:
    title = str(payload.get("title", "") or "").strip()
    selected_variants = dict(payload.get("selected_variants", {}) or {})
    oversized_titles = find_oversized_child_titles(
        profile,
        title,
        selected_variants,
        max_chars,
    )

    if not oversized_titles:
        return []

    sample = "; ".join(
        f"{child_title[:120]}{'...' if len(child_title) > 120 else ''} ({length} chars)"
        for child_title, length in oversized_titles[:3]
    )
    return [
        f"Amazon titles must be {max_chars} characters or fewer after variant prefixes. "
        f"Shorten the base title so colour/size/garment can be added safely. Too long: {sample}"
    ]


def build_preflight_report(
    profile: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    dropbox_overview: dict[str, Any],
    staged_folder_name: str,
    title: str,
    bullets: list[str],
    product_description: str,
    generic_keywords: str,
    selected_variants: dict[str, list[str]],
    size_price_map: dict[str, float],
    quantity: int,
    resolved_parent_main_image_url: str = "",
    resolved_other_images: list[str] | None = None,
    resolved_color_image_map: dict[str, str] | None = None,
    resolved_design_color_image_url_map: dict[str, dict[str, str]] | None = None,
    sku_decoration_code: str = "",
    sku_listing_code: str = "",
    allow_image_resolution_fallback: bool = True,
    use_resource_fallback_images: bool = False,
    has_staged_resource_images: bool = False,
) -> dict[str, Any]:
    preview_parent_sku = build_parent_sku_from_context(profile, sku_decoration_code, sku_listing_code)
    sku_profile = apply_sku_context_to_profile(profile, sku_decoration_code, sku_listing_code)
    preview_selected_colors = get_selected_colors_for_image_resolution(profile, selected_variants)
    preview_selected_sizes = selected_variants.get("size", [])
    profile_schema_errors = validate_profile_schema(profile)


    preview_payload = {
        "parent_sku": preview_parent_sku,
        "title": title.strip(),
        "brand_name": GLOBAL_BRAND_NAME,
        "manufacturer": profile.get("manufacturer", ""),
        "recommended_browse_nodes": profile.get("recommended_browse_nodes", ""),
        "size_price_map": size_price_map,
        "sku_decoration_code": sku_decoration_code,
        "sku_listing_code": sku_listing_code,
        "write_parent_starting_price": profile.get("write_parent_starting_price", False),
        "quantity": quantity,
        "department_name": profile.get("department_name", ""),
        "target_gender": profile.get("target_gender", ""),
        "age_range_description": profile.get("age_range_description", ""),
        "feed_product_type": profile.get("feed_product_type", ""),
        "variation_theme": profile.get("variation_theme", "SizeColor"),
        "product_category": profile.get("product_category", "apparel"),
        "condition_type": profile.get("condition_type", "New"),
        "item_type_name": profile.get("item_type_name", ""),
        "country_of_origin": profile.get("country_of_origin", "United Kingdom"),
        "material_type": profile.get("material_type", ""),
        "style_name": profile.get("style_name", ""),
        "care_instructions": profile.get("care_instructions", ""),
        "theme": profile.get("theme", ""),
        "field_aliases": get_field_aliases(profile),
        "extra_parent_fields": get_extra_parent_fields(profile),
        "extra_child_fields": get_extra_child_fields(profile),
        "extra_child_fields_by_size": get_extra_child_fields_by_size(profile),
        "parent_main_image_url": "",
        "product_description": product_description.strip(),
        "generic_keywords": generic_keywords.strip(),
        "bullet_points": [bullet.strip() for bullet in bullets],
        "selected_variants": selected_variants,
        "colors": preview_selected_colors,
        "sizes": preview_selected_sizes,
        "other_images": list(resolved_other_images or []),
        "color_image_map": dict(resolved_color_image_map or {}),
        "design_color_image_url_map": dict(resolved_design_color_image_url_map or {}),
        "dynamic_profile_fields": {},
    }
    if resolved_parent_main_image_url:
        preview_payload["parent_main_image_url"] = resolved_parent_main_image_url

    needs_main_or_variant_images = (
        not preview_payload.get("parent_main_image_url")
        or (
            not preview_payload.get("color_image_map")
            and not preview_payload.get("design_color_image_url_map")
        )
    )
    needs_secondary_images = (
        not preview_payload.get("other_images")
        and (has_staged_resource_images or use_resource_fallback_images)
    )
    if (
        allow_image_resolution_fallback
        and (needs_main_or_variant_images or needs_secondary_images)
        and staged_folder_name
        and preview_selected_colors
    ):
        try:
            stage_folder_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)
            (
                preview_payload["parent_main_image_url"],
                preview_payload["other_images"],
                preview_payload["color_image_map"],
                preview_payload["design_color_image_url_map"],
            ) = resolve_folder_image_urls(
                profile,
                selected_variants,
                preview_selected_colors,
                dropbox_overview,
                stage_folder_path,
                use_resource_fallback_images=use_resource_fallback_images,
            )
        except Exception:
            pass

    try:
        template_path = resolve_template_path(profile)
        if template_path.exists():
            wb = load_workbook(template_path, keep_vba=True, read_only=True)
            try:
                if SHEET_NAME in wb.sheetnames:
                    ws = wb[SHEET_NAME]
                    header_map = build_header_map(ws, HEADER_ROW)
                    preview_payload["dynamic_profile_fields"] = get_dynamic_profile_fields(profile, header_map)
                else:
                    preview_payload["dynamic_profile_fields"] = {}
            finally:
                wb.close()
        else:
            preview_payload["dynamic_profile_fields"] = {}
    except Exception:
        preview_payload["dynamic_profile_fields"] = {}

    preview_payload_errors = validate_payload(preview_payload)
    preview_variant_errors = validate_variants(
        selected_variants,
        size_price_map,
        quantity,
        profile=sku_profile,
    )
    preview_structure_errors = validate_parent_child_structure(preview_payload)

    template_errors = validate_template_file(profile)
    stock_ready_report = validate_stock_ready_payload(sku_profile, preview_payload)
    variant_image_count_errors = validate_variant_image_count(sku_profile, preview_payload)
    child_title_length_errors = validate_child_title_lengths(sku_profile, preview_payload)

    all_preview_errors = [
        *profile_schema_errors,
        *preview_payload_errors,
        *preview_variant_errors,
        *preview_structure_errors,
        *template_errors,
        *stock_ready_report.get("errors", []),
        *variant_image_count_errors,
        *child_title_length_errors,
    ]

    sku_profile = apply_sku_context_to_profile(
        profile,
        preview_payload.get("sku_decoration_code", ""),
        preview_payload.get("sku_listing_code", ""),
    )
    quality_report = validate_listing_quality(sku_profile, preview_payload)
    stock_ready_warnings = list(stock_ready_report.get("warnings", []))
    if stock_ready_warnings:
        quality_report["warnings"] = list(quality_report.get("warnings", [])) + stock_ready_warnings
    quality_report["stock_ready"] = {
        "missing_supplier_stock_key_count": stock_ready_report.get("missing_supplier_stock_key_count", 0),
        "duplicate_skus": stock_ready_report.get("duplicate_skus", []),
        "strict_stock_ready": is_strict_stock_ready(profile),
        "stock_reference_key": profile.get("stock_reference_key", ""),
    }

    return {
        "preview_payload": preview_payload,
        "all_preview_errors": all_preview_errors,
        "quality_report": quality_report,
        "stock_ready_report": stock_ready_report,
    }


def prepare_generation_payload(
    profile: dict[str, Any],
    title: str,
    bullets: list[str],
    product_description: str,
    generic_keywords: str,
    selected_variants: dict[str, list[str]],
    size_price_map: dict[str, float],
    sku_decoration_code: str,
    sku_listing_code: str,
    manual_sku_listing_code: str,
    generated_sku_listing_code: str,
    quantity: int,
    staged_folder_name: str,
    handling_time_days: int = DEFAULT_HANDLING_TIME_DAYS,
    merchant_shipping_group_name: str = "",
    parent_main_image_choice: str = "",
    parent_main_image_url: str = "",
    parent_sku_override: str = "",
) -> dict[str, Any]:
    selected_variants = normalize_saved_variant_values_for_profile(
        profile,
        selected_variants,
    )
    size_price_map = normalize_saved_price_map_for_profile(profile, size_price_map)
    base_parent_sku = str(get_default(profile, "parent_sku", "")).strip()
    manual_sku_listing_code = sanitize_sku(str(manual_sku_listing_code or "")).upper()
    generated_sku_listing_code = sanitize_sku(str(generated_sku_listing_code or "")).upper()
    sku_listing_code = sanitize_sku(str(sku_listing_code or manual_sku_listing_code or generated_sku_listing_code or "")).upper()
    if not sku_listing_code:
        generated_sku_listing_code = f"D{generate_unique_sku(5)}"
        sku_listing_code = generated_sku_listing_code
    elif not manual_sku_listing_code and not generated_sku_listing_code:
        generated_sku_listing_code = sku_listing_code
    derived_parent_sku = build_parent_sku_from_context(
        profile,
        sku_decoration_code,
        sku_listing_code,
    )
    parent_sku_override = sanitize_sku(
        str(parent_sku_override or "")
    ).upper()
    parent_sku = parent_sku_override or derived_parent_sku
    size_price_map = normalize_variant_price_map_for_selected_variants(
        profile,
        selected_variants,
        size_price_map,
    )

    payload = {
        "parent_sku": parent_sku,
        "parent_sku_override": parent_sku_override,
        "base_parent_sku": base_parent_sku,
        "model_name": profile.get("model_name", parent_sku),
        "title": title.strip(),
        "brand_name": GLOBAL_BRAND_NAME,
        "manufacturer": profile.get("manufacturer", ""),
        "recommended_browse_nodes": profile.get("recommended_browse_nodes", ""),
        "size_price_map": size_price_map,
        "sku_decoration_code": sku_decoration_code,
        "manual_sku_listing_code": manual_sku_listing_code,
        "generated_sku_listing_code": generated_sku_listing_code,
        "sku_listing_code": sku_listing_code,
        "write_parent_starting_price": profile.get("write_parent_starting_price", False),
        "use_same_price_for_all_sizes": st.session_state.get("use_same_price_for_all_sizes", False),
        "price_input_mode": st.session_state.get("design_size_pricing_mode", ""),
        "quantity": quantity,
        "handling_time_days": normalize_handling_time_days(handling_time_days),
        "merchant_shipping_group_name": normalize_merchant_shipping_group(merchant_shipping_group_name),
        "department_name": profile.get("department_name", ""),
        "target_gender": profile.get("target_gender", ""),
        "age_range_description": profile.get("age_range_description", ""),
        "feed_product_type": profile.get("feed_product_type", ""),
        "variation_theme": profile.get("variation_theme", "SizeColor"),
        "product_category": profile.get("product_category", "apparel"),
        "condition_type": profile.get("condition_type", "New"),
        "item_type_name": profile.get("item_type_name", ""),
        "country_of_origin": profile.get("country_of_origin", "United Kingdom"),
        "material_type": profile.get("material_type", ""),
        "style_name": profile.get("style_name", ""),
        "care_instructions": profile.get("care_instructions", ""),
        "collar_style": profile.get("collar_style", "Crew Neck"),
        "neck_style": profile.get("neck_style", profile.get("collar_style", "Crew Neck")),
        "theme": profile.get("theme", ""),
        "field_aliases": get_field_aliases(profile),
        "extra_parent_fields": get_extra_parent_fields(profile),
        "extra_child_fields": get_extra_child_fields(profile),
        "parent_main_image_url": "",
        "parent_main_image_choice": parent_main_image_choice,
        "selected_parent_main_image_url": parent_main_image_url,
        "product_description": product_description.strip(),
        "generic_keywords": generic_keywords.strip(),
        "bullet_points": [bullet.strip() for bullet in bullets],
        "selected_variants": selected_variants,
        "colors": selected_variants.get("color", []),
        "sizes": selected_variants.get("size", []),
        "other_images": [],
        "color_image_map": {},
        "design_color_image_url_map": {},
        "dynamic_profile_fields": {},
    }

    errors: list[str] = []

    description_chars = len(payload["product_description"])
    if description_chars < 1000:
        errors.append("Description must be at least 1000 characters.")
    if description_chars > 2000:
        errors.append("Description must be under 2000 characters.")

    if not payload["title"]:
        errors.append("Title is required.")

    if any(not bullet for bullet in payload["bullet_points"]):
        errors.append("All five bullet points are required.")

    if not staged_folder_name:
        errors.append("Select a staged Dropbox folder.")

    if not parent_sku:
        errors.append("This template is missing parent_sku in its config.")

    if not sku_decoration_code.strip():
        errors.append("SKU decoration code is required.")

    if not sku_listing_code.strip():
        errors.append("SKU listing/design code is required.")

    sku_profile = apply_sku_context_to_profile(
        profile,
        payload.get("sku_decoration_code", ""),
        payload.get("sku_listing_code", ""),
    )

    errors.extend(validate_variant_dimensions(selected_variants))
    errors.extend(validate_payload(payload))
    errors.extend(validate_variants(selected_variants, size_price_map, quantity, profile=sku_profile))
    errors.extend(validate_parent_child_structure(payload))
    errors.extend(validate_template_file(profile))
    errors.extend(validate_child_title_lengths(sku_profile, payload))

    quality_report = validate_listing_quality(sku_profile, payload)
    stock_ready_report = validate_stock_ready_payload(profile, payload)
    errors.extend(stock_ready_report.get("errors", []))
    stock_ready_warnings = list(stock_ready_report.get("warnings", []))
    if stock_ready_warnings:
        quality_report["warnings"] = list(quality_report.get("warnings", [])) + stock_ready_warnings
    quality_report["stock_ready"] = {
        "missing_supplier_stock_key_count": stock_ready_report.get("missing_supplier_stock_key_count", 0),
        "duplicate_skus": stock_ready_report.get("duplicate_skus", []),
        "strict_stock_ready": is_strict_stock_ready(profile),
        "stock_reference_key": profile.get("stock_reference_key", ""),
    }

    return {
        "payload": payload,
        "errors": errors,
        "quality_report": quality_report,
        "stock_ready_report": stock_ready_report,
    }


def get_dynamic_profile_fields(
    profile: dict[str, Any],
    header_map: dict[str, int],
) -> dict[str, Any]:
    allowed_fields = set(get_allowed_dynamic_fields(profile))
    dynamic: dict[str, Any] = {}

    for key in allowed_fields:
        if key not in header_map:
            continue

        value = profile.get(key)
        if isinstance(value, (list, dict)) or value in (None, ""):
            continue

        dynamic[key] = value

    return dynamic

def render_preflight_dashboard(
    quality_report: dict[str, Any],
    all_preview_errors: list[str],
) -> None:
    blockers = quality_report.get("blockers", [])
    warnings = quality_report.get("warnings", [])
    breakdown = quality_report.get("breakdown", {})
    search_terms_bytes = quality_report.get("search_terms_bytes", 0)

    template_ok = not any("Template" in err or "Sheet" in err for err in all_preview_errors)
    variants_ok = breakdown.get("variant_integrity", 0) > 0 and not any(
        "price" in err.lower() or "variant" in err.lower() or "parent_sku" in err.lower()
        for err in all_preview_errors + blockers
    )
    images_ok = breakdown.get("image_integrity", 0) > 0 and not any(
        "image" in err.lower() for err in blockers
    )

    copy_status = "Pass"
    if blockers:
        copy_status = "Fail"
    elif warnings:
        copy_status = "Warn"

    template_status = "Pass" if template_ok else "Fail"
    variants_status = "Pass" if variants_ok else "Fail"

    if any("image" in item.lower() for item in blockers):
        images_status = "Fail"
    elif any("image" in item.lower() for item in warnings):
        images_status = "Warn"
    else:
        images_status = "Pass"

    ready_to_generate = not all_preview_errors and not blockers

    st.subheader("Preflight dashboard")

    col1, col2, col3, col4, col5, col6 = st.columns(6)
    col1.metric("Template", template_status)
    col2.metric("Copy", copy_status)
    col3.metric("Variants", variants_status)
    col4.metric("Images", images_status)
    col5.metric("Search terms", f"{search_terms_bytes}/249")
    col6.metric("Ready", "Yes" if ready_to_generate else "No")

    top_fixes: list[str] = []
    top_fixes.extend(all_preview_errors[:3])
    top_fixes.extend(blockers[:3])

    for warning in warnings:
        if len(top_fixes) >= 6:
            break
        top_fixes.append(warning)

    if top_fixes:
        st.markdown("**Top fixes**")
        for item in top_fixes[:6]:
            st.write(f"- {item}")
    else:
        st.success("Everything looks ready.")

def render_listing_score_result(
    quality_report: dict[str, Any],
    all_preview_errors: list[str],
) -> None:
    st.subheader("Listing score result")
    st.metric("Internal quality score", f"{quality_report['score']}/100")

    with st.expander("Quality details", expanded=True):
        st.write("Breakdown:")
        st.json(quality_report["breakdown"])

        st.write("Copy metrics:")
        st.write(f"- Title characters: {quality_report.get('title_chars', 0)}")
        st.write(f"- Description characters: {quality_report.get('description_chars', 0)}")
        st.write(f"- Search terms bytes: {quality_report.get('search_terms_bytes', 0)}/249")

        bullet_char_counts = quality_report.get("bullet_char_counts", [])
        if bullet_char_counts:
            st.write("- Bullet character counts:")
            for idx, count in enumerate(bullet_char_counts, start=1):
                st.write(f"  - Bullet {idx}: {count}")

        if all_preview_errors:
            st.error("Validation errors:")
            for item in all_preview_errors:
                st.write(f"- {item}")

        if quality_report["blockers"]:
            st.error("Quality blockers:")
            for item in quality_report["blockers"]:
                st.write(f"- {item}")
        else:
            st.success("No quality blockers found.")

        if quality_report["warnings"]:
            st.warning("Warnings:")
            for item in quality_report["warnings"]:
                st.write(f"- {item}")
        else:
            st.info("No warnings.")


def find_template_matches_for_staged_folder(
    staged_folder_name: str,
    profiles: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    folder_name = (staged_folder_name or "").strip()
    if not folder_name:
        return []

    folder_upper = folder_name.upper()

    def bounded_match(code: str) -> bool:
        code = (code or "").strip().upper()
        if not code:
            return False
        pattern = rf"(?<![A-Z0-9]){re.escape(code)}(?![A-Z0-9])"
        return bool(re.search(pattern, folder_upper))

    matches: list[tuple[int, dict[str, Any]]] = []
    seen_slugs: set[str] = set()

    if bounded_match("T01") or bounded_match("T02"):
        for profile in profiles:
            if (
                str(profile.get("_family_slug", "")).strip().upper() == "SHIRT"
                and str(profile.get("template_key", "")).strip().upper() == "GENERIC_SHIRTS"
            ):
                matches.append((5, profile))
                seen_slugs.add(profile.get("_slug", ""))
                break

    if bounded_match("H01") or bounded_match("H02"):
        for profile in profiles:
            if (
                str(profile.get("_family_slug", "")).strip().upper() == "HOODIE"
                and str(profile.get("template_key", "")).strip().upper() == "GENERIC_HOODIES"
            ):
                matches.append((5, profile))
                seen_slugs.add(profile.get("_slug", ""))
                break

    if bounded_match("S01") or bounded_match("S02"):
        for profile in profiles:
            if (
                str(profile.get("_family_slug", "")).strip().upper() == "HOODIE"
                and str(profile.get("template_key", "")).strip().upper() == "GENERIC_SWEATSHIRTS"
            ):
                matches.append((5, profile))
                seen_slugs.add(profile.get("_slug", ""))
                break

    for profile in profiles:
        template_key = str(profile.get("template_key", "")).strip()
        parent_sku = str(profile.get("parent_sku", "")).strip()

        score = 0
        if bounded_match(template_key):
            score = 2
        elif bounded_match(parent_sku):
            score = 1

        if score <= 0:
            continue

        slug = profile.get("_slug", "")
        if slug in seen_slugs:
            continue

        seen_slugs.add(slug)
        matches.append((score, profile))

    matches.sort(key=lambda item: (-item[0], item[1].get("_family_slug", ""), item[1].get("label", item[1].get("_slug", ""))))
    return [profile for _, profile in matches]


def scan_staged_folder_readiness(
    staged_folder_name: str,
    profiles: list[dict[str, Any]],
    dropbox_cfg: dict[str, Any],
) -> dict[str, Any]:
    matches = find_template_matches_for_staged_folder(staged_folder_name, profiles)

    result = {
        "folder_name": staged_folder_name,
        "detected_template": "",
        "detection_status": "",
        "staged_image_readiness": "",
        "garment_support_readiness": "",
        "overall_status": "",
        "reason": "",
    }

    if not matches:
        result.update({
            "detection_status": "No match",
            "staged_image_readiness": "Unknown",
            "garment_support_readiness": "Unknown",
            "overall_status": "Blocked",
            "reason": "No template match found from the staged folder name.",
        })
        return result

    if len(matches) > 1:
        match_labels = ", ".join(match.get("label", match.get("_slug", "")) for match in matches)
        result.update({
            "detected_template": match_labels,
            "detection_status": "Ambiguous",
            "staged_image_readiness": "Unknown",
            "garment_support_readiness": "Unknown",
            "overall_status": "Blocked",
            "reason": "Multiple template matches found; confirm the template manually.",
        })
        return result

    profile = matches[0]
    result["detected_template"] = profile.get("label", profile.get("_slug", ""))
    result["detection_status"] = "Single match"

    stage_folder_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)
    dropbox_overview = build_dropbox_overview(profile, dropbox_cfg)
    main_image_map = dropbox_overview.get("main_image_map", {})
    color_sku_map = dropbox_overview.get("color_sku_map", profile.get("color_sku_map", {}))
    template_key = str(profile.get("template_key", "") or "")

    missing_files: list[str] = []
    for color in get_profile_color_options(profile):
        filename = main_image_map.get(color, "")
        staged_path, candidates = resolve_existing_color_image_path(
            stage_folder_path,
            template_key,
            color,
            filename,
            str(color_sku_map.get(color, "") or ""),
        )
        if not staged_path:
            missing_files.append(candidates[0] if candidates else color)

    if missing_files:
        result.update({
            "staged_image_readiness": "Missing required mapped images",
            "garment_support_readiness": "Present" if dropbox_overview.get("garment_resource_images") else "Missing or unavailable",
            "overall_status": "Blocked",
            "reason": f"Missing staged mapped images: {', '.join(missing_files[:3])}" + ("..." if len(missing_files) > 3 else ""),
        })
        return result

    staged_resource_images = build_stage_resource_image_paths(dropbox_cfg, staged_folder_name)
    garment_support_images = dropbox_overview.get("garment_resource_images", [])
    garment_warning = dropbox_overview.get("garment_resource_warning", "")

    if staged_resource_images or garment_support_images:
        result.update({
            "staged_image_readiness": "Ready",
            "garment_support_readiness": "Ready",
            "overall_status": "Ready",
            "reason": "Template detected and required staged/support images are present.",
        })
        return result

    result.update({
        "staged_image_readiness": "Ready",
        "garment_support_readiness": "Missing or unavailable",
        "overall_status": "Warning",
        "reason": garment_warning or "Required staged mapped images exist, but garment support images are missing.",
    })
    return result


def find_profile_for_listing_memory(
    profiles: list[dict[str, Any]],
    listing_memory: dict[str, Any],
) -> dict[str, Any] | None:
    template_key = str(listing_memory.get("template_key", "")).strip()
    if template_key:
        for profile in profiles:
            if str(profile.get("template_key", "")).strip() == template_key:
                return profile

    template_slug = str(listing_memory.get("template_slug", "")).strip()
    if template_slug:
        for profile in profiles:
            if str(profile.get("_slug", "")).strip() == template_slug:
                return profile

    return None


def build_christmas_group_target_profiles(
    profiles: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    target_keys = {
        "tshirt": "GENERIC_SHIRTS",
        "sweatshirt": "GENERIC_SWEATSHIRTS",
        "hoodie": "GENERIC_HOODIES",
    }
    targets: dict[str, dict[str, Any]] = {}
    for member_key, template_key in target_keys.items():
        matches = [
            candidate
            for candidate in profiles
            if str(candidate.get("template_key", "") or "").strip() == template_key
        ]
        if len(matches) != 1:
            raise ValueError(
                f"Christmas {member_key} requires exactly one {template_key} profile; "
                f"found {len(matches)}."
            )
        targets[member_key] = matches[0]
    return targets


def build_variants_summary(selected_variants: dict[str, list[str]]) -> str:
    parts = [
        f"{dim_name}: {len(values)}"
        for dim_name, values in selected_variants.items()
        if values
    ]
    return ", ".join(parts) if parts else "No variants"


def build_price_summary(size_price_map: dict[str, float]) -> str:
    if not size_price_map:
        return "No pricing"

    prices: list[float] = []
    missing_count = 0
    for raw_price in dict(size_price_map or {}).values():
        try:
            price = float(raw_price)
        except (TypeError, ValueError):
            price = 0.0
        if price > 0:
            prices.append(price)
        else:
            missing_count += 1

    if not prices:
        return f"No valid pricing ({missing_count} missing)"

    if len(set(prices)) == 1:
        summary = f"{len(prices)} price key(s) at {prices[0]:.2f}"
    else:
        summary = f"{len(prices)} price key(s) from {min(prices):.2f} to {max(prices):.2f}"

    if missing_count:
        summary += f" ({missing_count} missing)"
    return summary


def build_review_price_rows(
    profile: dict[str, Any] | None,
    listing_memory: dict[str, Any],
) -> list[dict[str, Any]]:
    profile = profile or {}
    selected_variants = dict(listing_memory.get("selected_variants", {}) or {})
    size_price_map = dict(listing_memory.get("size_price_map", {}) or {})
    rows: list[dict[str, Any]] = []

    if has_design_size_pricing(profile, selected_variants):
        variant_combos = build_variant_combinations(
            profile,
            {
                "design": list(selected_variants.get("design", []) or []),
                "size": list(selected_variants.get("size", []) or []),
            },
        )
        for combo in variant_combos:
            price_key = build_variant_price_key(combo)
            try:
                price = float(get_variant_price_from_map(profile, size_price_map, combo, fallback=0) or 0)
            except (TypeError, ValueError):
                price = 0.0
            rows.append({
                "Garment": combo.get("design", ""),
                "Size": combo.get("size", ""),
                "Price key": price_key,
                "Price": price,
                "Status": "OK" if price > 0 else "Missing",
            })
        return rows

    for key, raw_price in size_price_map.items():
        try:
            price = float(raw_price)
        except (TypeError, ValueError):
            price = 0.0
        rows.append({
            "Size": str(key),
            "Price key": str(key),
            "Price": price,
            "Status": "OK" if price > 0 else "Missing",
        })
    return rows


def safe_widget_key_part(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_]+", "_", str(value or "")).strip("_") or "value"


def build_review_memory_fingerprint(
    listing_memory: dict[str, Any],
    profile: dict[str, Any] | None = None,
) -> str:
    payload = {
        "template": profile.get("template_key", profile.get("_slug", "")) if profile else "",
        "sku_decoration_code": listing_memory.get("sku_decoration_code", ""),
        "manual_sku_listing_code": listing_memory.get("manual_sku_listing_code", ""),
        "generated_sku_listing_code": listing_memory.get("generated_sku_listing_code", ""),
        "sku_listing_code": listing_memory.get("sku_listing_code", ""),
        "parent_sku": listing_memory.get("parent_sku", ""),
        "merchant_shipping_group_name": normalize_merchant_shipping_group(
            listing_memory.get("merchant_shipping_group_name", "")
        ),
        "size_price_map": listing_memory.get("size_price_map", {}),
        "price_input_mode": listing_memory.get("price_input_mode", ""),
        "title": listing_memory.get("title", ""),
        "product_description": listing_memory.get("product_description", ""),
        "generic_keywords": listing_memory.get("generic_keywords", ""),
        "bullet_points": listing_memory.get("bullet_points", []),
    }
    raw = json.dumps(payload, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:10]


def normalize_design_size_pricing_mode(value: Any, default: str = "Manual price by garment/size") -> str:
    pricing_modes = {
        "Use one price for all",
        "Use one price per cluster",
        "Manual price by garment/size",
    }
    legacy_pricing_modes = {
        "One price for all": "Use one price for all",
        "One price per cluster": "Use one price per cluster",
        "Manual by garment/size": "Manual price by garment/size",
    }
    value = legacy_pricing_modes.get(str(value or ""), str(value or ""))
    return value if value in pricing_modes else default


def get_review_edit_state_keys(review_key_prefix: str) -> dict[str, str]:
    return {
        "context": f"{review_key_prefix}_edit_context",
        "sku_decoration_choice": f"{review_key_prefix}_sku_decoration_choice",
        "custom_sku_decoration_code": f"{review_key_prefix}_custom_sku_decoration_code",
        "manual_sku_listing_code": f"{review_key_prefix}_manual_sku_listing_code",
        "generated_sku_listing_code": f"{review_key_prefix}_generated_sku_listing_code",
        "merchant_shipping_group_name": f"{review_key_prefix}_merchant_shipping_group_name",
        "price_input_mode": f"{review_key_prefix}_price_input_mode",
    }


def has_complete_review_sku_price_state(
    review_key_prefix: str,
    listing_memory: dict[str, Any],
    profile: dict[str, Any],
) -> bool:
    keys = get_review_edit_state_keys(review_key_prefix)
    selected_variants = dict(listing_memory.get("selected_variants", {}))
    saved_prices = normalize_saved_price_map_for_profile(
        profile,
        listing_memory.get("size_price_map", {}),
    )

    required_keys = [
        keys["context"],
        keys["sku_decoration_choice"],
        keys["custom_sku_decoration_code"],
        keys["manual_sku_listing_code"],
        keys["generated_sku_listing_code"],
        keys["merchant_shipping_group_name"],
        keys["price_input_mode"],
    ]
    if has_design_size_pricing(profile, selected_variants):
        variant_combos = build_variant_combinations(
            profile,
            {
                "design": list(selected_variants.get("design", []) or []),
                "size": list(selected_variants.get("size", []) or []),
            },
        )
        for combo in variant_combos:
            variant_price_key = build_variant_price_key(combo)
            required_keys.append(f"{review_key_prefix}_price_{safe_widget_key_part(variant_price_key)}")
    else:
        for size in saved_prices:
            required_keys.append(f"{review_key_prefix}_price_{safe_widget_key_part(size)}")

    return all(key in st.session_state for key in required_keys)


def clear_review_editor_state(review_key_prefix: str) -> None:
    prefixes = [
        f"{review_key_prefix}_edit_",
        f"{review_key_prefix}_content_",
        f"{review_key_prefix}_price_",
        f"{review_key_prefix}_sku_",
        f"{review_key_prefix}_custom_sku_",
        f"{review_key_prefix}_manual_sku_",
        f"{review_key_prefix}_generated_sku_",
        f"{review_key_prefix}_active_editor_",
    ]
    review_state_tokens = [
        "_edit_",
        "_content_",
        "_price_",
        "_sku_",
        "_custom_sku_",
        "_manual_sku_",
        "_generated_sku_",
    ]
    exact_keys = set(get_review_edit_state_keys(review_key_prefix).values())
    exact_keys.update(get_review_content_state_keys(review_key_prefix).values())
    exact_keys.add(f"{review_key_prefix}_active_editor_prefix")
    for key in list(st.session_state.keys()):
        key_text = str(key)
        fingerprinted_review_key = (
            key_text.startswith(f"{review_key_prefix}_")
            and any(token in key_text for token in review_state_tokens)
        )
        if key in exact_keys or fingerprinted_review_key or any(key_text.startswith(prefix) for prefix in prefixes):
            st.session_state.pop(key, None)


def get_review_content_state_keys(review_key_prefix: str) -> dict[str, str]:
    return {
        "context": f"{review_key_prefix}_content_context",
        "title": f"{review_key_prefix}_content_title",
        "description": f"{review_key_prefix}_content_description",
        "keywords": f"{review_key_prefix}_content_keywords",
        "bullet_1": f"{review_key_prefix}_content_bullet_1",
        "bullet_2": f"{review_key_prefix}_content_bullet_2",
        "bullet_3": f"{review_key_prefix}_content_bullet_3",
        "bullet_4": f"{review_key_prefix}_content_bullet_4",
        "bullet_5": f"{review_key_prefix}_content_bullet_5",
    }


def has_complete_review_content_state(review_key_prefix: str) -> bool:
    keys = get_review_content_state_keys(review_key_prefix)
    required_keys = [
        keys["context"],
        keys["title"],
        keys["description"],
        keys["keywords"],
        *[keys[f"bullet_{idx}"] for idx in range(1, 6)],
    ]
    return all(key in st.session_state for key in required_keys)


def initialize_review_content_state(
    review_key_prefix: str,
    listing_memory: dict[str, Any],
) -> None:
    keys = get_review_content_state_keys(review_key_prefix)
    bullet_points = (list(listing_memory.get("bullet_points", [])) + ["", "", "", "", ""])[:5]
    context = json.dumps(
        {
            "title": listing_memory.get("title", ""),
            "product_description": listing_memory.get("product_description", ""),
            "generic_keywords": listing_memory.get("generic_keywords", ""),
            "bullet_points": bullet_points,
        },
        sort_keys=True,
    )
    required_content_keys = [
        keys["title"],
        keys["description"],
        keys["keywords"],
        *[keys[f"bullet_{idx}"] for idx in range(1, 6)],
    ]
    if (
        st.session_state.get(keys["context"]) == context
        and all(key in st.session_state for key in required_content_keys)
    ):
        return

    st.session_state[keys["title"]] = str(listing_memory.get("title", "") or "")
    st.session_state[keys["description"]] = str(listing_memory.get("product_description", "") or "")
    st.session_state[keys["keywords"]] = str(listing_memory.get("generic_keywords", "") or "")
    for idx, bullet in enumerate(bullet_points, start=1):
        st.session_state[keys[f"bullet_{idx}"]] = str(bullet or "")
    st.session_state[keys["context"]] = context


def get_review_content_edits(review_key_prefix: str) -> dict[str, Any]:
    keys = get_review_content_state_keys(review_key_prefix)
    return {
        "title": str(st.session_state.get(keys["title"], "") or "").strip(),
        "product_description": str(st.session_state.get(keys["description"], "") or "").strip(),
        "generic_keywords": str(st.session_state.get(keys["keywords"], "") or "").strip(),
        "bullet_points": [
            str(st.session_state.get(keys[f"bullet_{idx}"], "") or "").strip()
            for idx in range(1, 6)
        ],
    }


def apply_review_content_edits(
    payload: dict[str, Any],
    edits: dict[str, Any],
) -> dict[str, Any]:
    payload = dict(payload)
    for field_name in ["title", "product_description", "generic_keywords", "bullet_points"]:
        payload[field_name] = edits.get(field_name, payload.get(field_name, "" if field_name != "bullet_points" else []))
    return payload


def render_review_content_editor(
    review_key_prefix: str,
    listing_memory: dict[str, Any],
) -> dict[str, Any]:
    initialize_review_content_state(review_key_prefix, listing_memory)
    keys = get_review_content_state_keys(review_key_prefix)

    st.text_input("Product title", key=keys["title"])
    title_chars = len(str(st.session_state.get(keys["title"], "") or "").strip())
    st.caption(f"Title: {title_chars}/200 chars")

    st.markdown("**Bullets**")
    for idx in range(1, 6):
        st.text_input(f"Bullet {idx}", key=keys[f"bullet_{idx}"])

    st.text_area("Product description", height=180, key=keys["description"])
    description_chars = len(str(st.session_state.get(keys["description"], "") or "").strip())
    st.caption(f"Description: {description_chars}/2000 chars")

    st.text_area("Search terms", height=100, key=keys["keywords"])
    keywords_bytes = len(str(st.session_state.get(keys["keywords"], "") or "").encode("utf-8"))
    st.caption(f"Search terms: {keywords_bytes}/249 bytes")

    return get_review_content_edits(review_key_prefix)


CHRISTMAS_GROUP_FOLDER_SUFFIX_BY_MEMBER = {
    "tshirt": "TSHIRT",
    "sweatshirt": "SWEATSHIRT",
    "hoodie": "HOODIE",
}

CHRISTMAS_GROUP_PARENT_SUFFIX_BY_MEMBER = {
    "tshirt": "T",
    "sweatshirt": "S",
    "hoodie": "H",
}


def resolve_grouped_child_listing_code_for_review(
    listing_memory: dict[str, Any],
    value: Any,
    *,
    prefer_source_listing_code: bool = False,
    profile: dict[str, Any] | None = None,
) -> str:
    normalized = sanitize_sku(str(value or "")).upper()
    source_group = listing_memory.get("source_group")
    source_group = source_group if isinstance(source_group, dict) else {}
    member_key = resolve_christmas_grouped_child_member_key(
        listing_memory,
        profile or {},
    )
    if not member_key:
        return normalized

    if prefer_source_listing_code:
        explicit_source_code = sanitize_sku(
            str(source_group.get("source_listing_code", "") or "")
        ).upper()
        if explicit_source_code:
            return explicit_source_code

    folder_suffix = CHRISTMAS_GROUP_FOLDER_SUFFIX_BY_MEMBER.get(member_key, "")
    has_saved_group_identity = (
        str(source_group.get("group_type", "") or "").strip() == "christmas_project"
    )
    if profile and not has_saved_group_identity:
        try:
            folder_suffix = str(
                derive_christmas_group_members(profile)[member_key]["folder_suffix"]
            ).strip().upper()
        except (KeyError, TypeError, ValueError):
            return normalized
    if not normalized or not folder_suffix:
        return normalized

    suffix_marker = f"-{folder_suffix}"
    if normalized.endswith(suffix_marker):
        return normalized[:-len(suffix_marker)].rstrip("-")

    return normalized


def normalize_saved_variant_values_for_profile(
    profile: dict[str, Any],
    selected_variants: dict[str, list[str]],
) -> dict[str, list[str]]:
    aliases_by_dimension = profile.get("saved_variant_value_aliases", {})
    normalized = {
        str(dimension): list(values or [])
        for dimension, values in dict(selected_variants or {}).items()
    }
    if not isinstance(aliases_by_dimension, dict):
        return normalized

    for dimension, values in list(normalized.items()):
        aliases = aliases_by_dimension.get(str(dimension).strip().lower(), {})
        if not isinstance(aliases, dict):
            continue
        lookup = {
            str(source).strip().casefold(): str(target).strip()
            for source, target in aliases.items()
            if str(source).strip() and str(target).strip()
        }
        normalized[dimension] = [
            lookup.get(str(value).strip().casefold(), value)
            for value in values
        ]
    return normalized


def normalize_saved_price_map_for_profile(
    profile: dict[str, Any],
    size_price_map: dict[str, Any],
) -> dict[str, Any]:
    aliases_by_dimension = profile.get("saved_variant_value_aliases", {})
    size_aliases = (
        aliases_by_dimension.get("size", {})
        if isinstance(aliases_by_dimension, dict)
        else {}
    )
    if not isinstance(size_aliases, dict):
        return dict(size_price_map or {})
    lookup = {
        str(source).strip().casefold(): str(target).strip()
        for source, target in size_aliases.items()
        if str(source).strip() and str(target).strip()
    }
    normalized: dict[str, Any] = {}
    for raw_key, value in dict(size_price_map or {}).items():
        key = str(raw_key)
        design, separator, size = key.partition("||")
        source_size = size if separator else design
        translated_size = lookup.get(source_size.strip().casefold(), source_size)
        translated_key = f"{design}{separator}{translated_size}" if separator else translated_size
        normalized[translated_key] = value
    return normalized


def resolve_christmas_grouped_child_member_key(
    listing_memory: dict[str, Any],
    profile: dict[str, Any],
) -> str:
    source_group = listing_memory.get("source_group")
    if isinstance(source_group, dict) and (
        str(source_group.get("group_type", "") or "").strip() == "christmas_project"
    ):
        member_key = str(source_group.get("member_key", "") or "").strip().casefold()
        if member_key in CHRISTMAS_GROUP_PARENT_SUFFIX_BY_MEMBER:
            return member_key

    if str(profile.get("template_key", "") or "").strip().upper() != "CP":
        return ""
    if str(listing_memory.get("template_key", "") or "").strip().upper() != "CP":
        return ""

    selected_designs = {
        str(design).strip().casefold()
        for design in listing_memory.get("selected_variants", {}).get("design", [])
        if str(design).strip()
    }
    if not selected_designs:
        return ""

    try:
        members = derive_christmas_group_members(profile)
    except (TypeError, ValueError):
        return ""
    matches = [
        member_key
        for member_key, member in members.items()
        if selected_designs
        == {
            str(design).strip().casefold()
            for design in member.get("designs", [])
            if str(design).strip()
        }
    ]
    return matches[0] if len(matches) == 1 else ""


def build_review_parent_sku(
    listing_memory: dict[str, Any],
    profile: dict[str, Any],
    sku_decoration_code: str,
    sku_listing_code: str,
) -> str:
    explicit_parent_sku = sanitize_sku(
        str(listing_memory.get("parent_sku_override", "") or "")
    ).upper()
    if explicit_parent_sku:
        return explicit_parent_sku

    parent_sku = build_parent_sku_from_context(
        profile,
        sku_decoration_code,
        sku_listing_code,
    )

    member_key = resolve_christmas_grouped_child_member_key(listing_memory, profile)
    parent_suffix = CHRISTMAS_GROUP_PARENT_SUFFIX_BY_MEMBER.get(member_key, "")
    if not parent_suffix:
        return parent_sku

    suffix_marker = f"-{parent_suffix}"
    if parent_sku.upper().endswith(suffix_marker):
        return parent_sku

    return f"{parent_sku}-{parent_suffix}"



def get_grouped_child_generation_parent_sku_override(
    listing_memory: dict[str, Any],
    profile: dict[str, Any],
) -> str:
    member_key = resolve_christmas_grouped_child_member_key(listing_memory, profile)
    if member_key not in {"tshirt", "sweatshirt", "hoodie"}:
        return ""

    # Exact compatibility repair for the original released CHRTST canary:
    # CHRTST-TSHIRT / CHRTST-SWEATSHIRT / CHRTST-HOODIE -> CHRTST.
    # Normal reviewed codes remain unchanged.
    listing_code = resolve_grouped_child_listing_code_for_review(
        listing_memory,
        listing_memory.get("sku_listing_code", ""),
        profile=profile,
    )
    if not listing_code:
        return ""

    decoration_code = get_default_sku_decoration_code(profile, listing_memory)

    return build_review_parent_sku(
        listing_memory,
        profile,
        decoration_code,
        listing_code,
    )


def get_listing_generation_parent_sku_override(
    listing_memory: dict[str, Any],
    profile: dict[str, Any],
) -> str:
    explicit_parent_sku = sanitize_sku(
        str(listing_memory.get("parent_sku_override", "") or "")
    ).upper()
    if explicit_parent_sku:
        return explicit_parent_sku
    return get_grouped_child_generation_parent_sku_override(listing_memory, profile)


def preserve_grouped_child_generation_context(
    listing_memory: dict[str, Any],
    generation_payload: dict[str, Any],
) -> dict[str, Any]:
    if isinstance(listing_memory.get("source_group"), dict):
        generation_payload["source_group"] = deepcopy(listing_memory["source_group"])
    return generation_payload

def initialize_review_edit_state(
    review_key_prefix: str,
    listing_memory: dict[str, Any],
    profile: dict[str, Any],
) -> None:
    keys = get_review_edit_state_keys(review_key_prefix)

    review_manual_sku_listing_code = resolve_grouped_child_listing_code_for_review(
        listing_memory,
        listing_memory.get("manual_sku_listing_code", ""),
        profile=profile,
    )
    review_generated_sku_listing_code = resolve_grouped_child_listing_code_for_review(
        listing_memory,
        listing_memory.get("generated_sku_listing_code", ""),
        profile=profile,
    )
    review_sku_listing_code = resolve_grouped_child_listing_code_for_review(
        listing_memory,
        listing_memory.get("sku_listing_code", ""),
        prefer_source_listing_code=True,
        profile=profile,
    )

    context = json.dumps(
        {
            "template": profile.get("template_key", profile.get("_slug", "")),
            "sku_decoration_code": listing_memory.get("sku_decoration_code", ""),
            "manual_sku_listing_code": review_manual_sku_listing_code,
            "generated_sku_listing_code": review_generated_sku_listing_code,
            "sku_listing_code": review_sku_listing_code,
            "merchant_shipping_group_name": normalize_merchant_shipping_group(
                listing_memory.get("merchant_shipping_group_name", "")
            ),
            "size_price_map": listing_memory.get("size_price_map", {}),
            "price_input_mode": listing_memory.get("price_input_mode", ""),
        },
        sort_keys=True,
    )
    sku_decoration_code = get_default_sku_decoration_code(profile, listing_memory)
    selected_variants = dict(listing_memory.get("selected_variants", {}))
    saved_prices = normalize_saved_price_map_for_profile(
        profile,
        listing_memory.get("size_price_map", {}),
    )

    if (
        st.session_state.get(keys["context"]) == context
        and has_complete_review_sku_price_state(review_key_prefix, listing_memory, profile)
    ):
        return

    st.session_state[keys["sku_decoration_choice"]] = (
        sku_decoration_code if sku_decoration_code in SKU_DECORATION_OPTIONS else "Custom"
    )
    st.session_state[keys["custom_sku_decoration_code"]] = (
        "" if sku_decoration_code in SKU_DECORATION_OPTIONS else sku_decoration_code
    )
    st.session_state[keys["manual_sku_listing_code"]] = review_manual_sku_listing_code
    st.session_state[keys["generated_sku_listing_code"]] = (
        review_generated_sku_listing_code
        or f"D{generate_unique_sku(5)}"
    )
    st.session_state[keys["merchant_shipping_group_name"]] = normalize_merchant_shipping_group(
        listing_memory.get("merchant_shipping_group_name", "")
    )
    st.session_state[keys["price_input_mode"]] = normalize_design_size_pricing_mode(
        listing_memory.get("price_input_mode", ""),
        default="Manual price by garment/size",
    )

    if has_design_size_pricing(profile, selected_variants):
        variant_combos = build_variant_combinations(
            profile,
            {
                "design": list(selected_variants.get("design", []) or []),
                "size": list(selected_variants.get("size", []) or []),
            },
        )
        for combo in variant_combos:
            variant_price_key = build_variant_price_key(combo)
            widget_key = f"{review_key_prefix}_price_{safe_widget_key_part(variant_price_key)}"
            st.session_state[widget_key] = float(
                get_variant_price_from_map(profile, saved_prices, combo, fallback=0) or 0
            )
    else:
        for size, price in saved_prices.items():
            price_key = f"{review_key_prefix}_price_{safe_widget_key_part(size)}"
            st.session_state[price_key] = float(price)

    st.session_state[keys["context"]] = context


def get_review_sku_and_price_edits(
    review_key_prefix: str,
    listing_memory: dict[str, Any],
    profile: dict[str, Any],
) -> dict[str, Any]:
    keys = get_review_edit_state_keys(review_key_prefix)
    sku_decoration_choice = st.session_state.get(keys["sku_decoration_choice"], "")
    if sku_decoration_choice == "Custom":
        sku_decoration_code = sanitize_sku(str(st.session_state.get(keys["custom_sku_decoration_code"], ""))).upper()
    else:
        sku_decoration_code = sanitize_sku(str(sku_decoration_choice)).upper()

    generated_sku_listing_code = sanitize_sku(str(st.session_state.get(keys["generated_sku_listing_code"], ""))).upper()
    if not generated_sku_listing_code:
        generated_sku_listing_code = f"D{generate_unique_sku(5)}"
        st.session_state[keys["generated_sku_listing_code"]] = generated_sku_listing_code

    manual_sku_listing_code = sanitize_sku(str(st.session_state.get(keys["manual_sku_listing_code"], ""))).upper()
    sku_listing_code = manual_sku_listing_code or generated_sku_listing_code
    parent_sku = build_review_parent_sku(
        listing_memory,
        profile,
        sku_decoration_code,
        sku_listing_code,
    )

    selected_variants = dict(listing_memory.get("selected_variants", {}))
    existing_prices = dict(listing_memory.get("size_price_map", {}))
    size_price_map: dict[str, float] = {}
    price_input_mode = normalize_design_size_pricing_mode(st.session_state.get(keys["price_input_mode"], ""))
    if has_design_size_pricing(profile, selected_variants):
        variant_combos = build_variant_combinations(
            profile,
            {
                "design": list(selected_variants.get("design", []) or []),
                "size": list(selected_variants.get("size", []) or []),
            },
        )
        if price_input_mode == "Use one price for all":
            first_combo = variant_combos[0] if variant_combos else {}
            fallback_price = float(get_variant_price_from_map(profile, existing_prices, first_combo, fallback=0) or 0)
            shared_key = f"{review_key_prefix}_price_all"
            shared_price = float(st.session_state.get(shared_key, fallback_price))
            for combo in variant_combos:
                size_price_map[build_variant_price_key(combo)] = shared_price
        elif price_input_mode == "Use one price per cluster":
            selected_designs = list(selected_variants.get("design", []) or [])
            selected_sizes = list(selected_variants.get("size", []) or [])
            design_size_map = profile.get("design_size_map", {})
            for design in selected_designs:
                valid_sizes = [
                    size
                    for size in selected_sizes
                    if not design_size_map.get(design) or size in design_size_map.get(design, [])
                ]
                clusters: dict[str, list[str]] = {}
                for size in valid_sizes:
                    clusters.setdefault(get_design_size_price_cluster_label(design, size), []).append(size)
                for cluster_label, cluster_sizes in clusters.items():
                    first_combo = {"design": design, "size": cluster_sizes[0]}
                    fallback_price = float(get_variant_price_from_map(profile, existing_prices, first_combo, fallback=0) or 0)
                    cluster_key = (
                        f"{review_key_prefix}_cluster_price_"
                        f"{safe_widget_key_part(design)}_{safe_widget_key_part(cluster_label)}"
                    )
                    cluster_price = float(st.session_state.get(cluster_key, fallback_price))
                    for size in cluster_sizes:
                        size_price_map[build_variant_price_key({"design": design, "size": size})] = cluster_price
        else:
            for combo in variant_combos:
                variant_price_key = build_variant_price_key(combo)
                widget_key = f"{review_key_prefix}_price_{safe_widget_key_part(variant_price_key)}"
                fallback_price = float(get_variant_price_from_map(profile, existing_prices, combo, fallback=0) or 0)
                size_price_map[variant_price_key] = float(st.session_state.get(widget_key, fallback_price))
    else:
        price_sizes = list(selected_variants.get("size", [])) or list(existing_prices.keys()) or ["default"]
        for size in price_sizes:
            widget_key = f"{review_key_prefix}_price_{safe_widget_key_part(size)}"
            fallback_price = float(existing_prices.get(size, 0) or 0)
            size_price_map[str(size)] = float(st.session_state.get(widget_key, fallback_price))

    return {
        "sku_decoration_code": sku_decoration_code,
        "manual_sku_listing_code": manual_sku_listing_code,
        "generated_sku_listing_code": generated_sku_listing_code,
        "sku_listing_code": sku_listing_code,
        "parent_sku": parent_sku,
        "merchant_shipping_group_name": normalize_merchant_shipping_group(
            st.session_state.get(keys["merchant_shipping_group_name"], "")
        ),
        "size_price_map": size_price_map,
        "price_input_mode": price_input_mode if has_design_size_pricing(profile, selected_variants) else "",
    }


def apply_review_sku_and_price_edits(
    payload: dict[str, Any],
    edits: dict[str, Any],
) -> dict[str, Any]:
    payload = dict(payload)
    for field_name in [
        "sku_decoration_code",
        "manual_sku_listing_code",
        "generated_sku_listing_code",
        "sku_listing_code",
        "parent_sku",
        "merchant_shipping_group_name",
        "size_price_map",
        "price_input_mode",
    ]:
        payload[field_name] = edits.get(field_name, payload.get(field_name, ""))
    return payload


def render_review_sku_price_editor(
    review_key_prefix: str,
    profile: dict[str, Any],
    listing_memory: dict[str, Any],
) -> dict[str, Any]:
    initialize_review_edit_state(review_key_prefix, listing_memory, profile)
    keys = get_review_edit_state_keys(review_key_prefix)

    sku_col1, sku_col2 = st.columns(2)
    with sku_col1:
        sku_decoration_choice = st.selectbox(
            "Decoration code",
            SKU_DECORATION_OPTIONS,
            key=keys["sku_decoration_choice"],
        )
        if sku_decoration_choice == "Custom":
            st.text_input(
                "Custom decoration code",
                key=keys["custom_sku_decoration_code"],
            )
    with sku_col2:
        generated_code = str(st.session_state.get(keys["generated_sku_listing_code"], ""))
        st.text_input(
            "Listing/design code (optional)",
            key=keys["manual_sku_listing_code"],
            placeholder=generated_code,
            help="Leave blank to use the generated unique design identifier.",
        )
        st.caption(f"Generated code: `{generated_code or '-'}`")

    st.markdown("**Fulfillment**")
    st.selectbox(
        "Merchant Shipping Group",
        MERCHANT_SHIPPING_GROUP_OPTIONS,
        key=keys["merchant_shipping_group_name"],
        help="Leave empty to keep Amazon/default fulfilment behavior.",
    )

    edits = get_review_sku_and_price_edits(review_key_prefix, listing_memory, profile)
    st.caption(f"Parent SKU: `{edits['parent_sku']}`")

    selected_variants = dict(listing_memory.get("selected_variants", {}))
    sku_length_report = build_child_sku_length_report(
        profile,
        edits["parent_sku"],
        selected_variants,
        sku_decoration_code=edits["sku_decoration_code"],
        sku_listing_code=edits["sku_listing_code"],
    )
    st.caption(
        f"Child SKU length: max `{sku_length_report['max_length']}/{MAX_AMAZON_SKU_LENGTH}` "
        f"across `{sku_length_report['count']}` variant(s)."
    )
    if sku_length_report["oversized"]:
        sample = ", ".join(
            f"{row['sku']} ({row['length']})"
            for row in sku_length_report["oversized"][:5]
        )
        st.error(
            f"SKU is too long for review/export. Keep every child SKU <= {MAX_AMAZON_SKU_LENGTH} chars. "
            f"Shorten the MPN/listing code. Too long: {sample}"
        )
    existing_prices = dict(listing_memory.get("size_price_map", {}))
    if has_design_size_pricing(profile, selected_variants):
        st.markdown("**Price by garment and size**")
        pricing_modes = [
            "Use one price for all",
            "Use one price per cluster",
            "Manual price by garment/size",
        ]
        if keys["price_input_mode"] not in st.session_state:
            st.session_state[keys["price_input_mode"]] = normalize_design_size_pricing_mode(
                listing_memory.get("price_input_mode", ""),
            )
        pricing_mode = st.radio(
            "Pricing mode",
            pricing_modes,
            horizontal=True,
            key=keys["price_input_mode"],
        )

        selected_designs = list(selected_variants.get("design", []) or [])
        selected_sizes = list(selected_variants.get("size", []) or [])
        design_size_map = profile.get("design_size_map", {})
        variant_combos = build_variant_combinations(
            profile,
            {
                "design": selected_designs,
                "size": selected_sizes,
            },
        )

        if pricing_mode == "Use one price for all":
            first_combo = variant_combos[0] if variant_combos else {}
            shared_key = f"{review_key_prefix}_price_all"
            if shared_key not in st.session_state:
                st.session_state[shared_key] = float(
                    get_variant_price_from_map(profile, existing_prices, first_combo, fallback=0) or 0
                )
            st.number_input(
                "Price for all garment/size combinations",
                min_value=0.0,
                step=0.5,
                format="%.2f",
                key=shared_key,
            )
        elif pricing_mode == "Use one price per cluster":
            for design in selected_designs:
                valid_sizes = [
                    size
                    for size in selected_sizes
                    if not design_size_map.get(design) or size in design_size_map.get(design, [])
                ]
                if not valid_sizes:
                    continue

                clusters: dict[str, list[str]] = {}
                for size in valid_sizes:
                    clusters.setdefault(get_design_size_price_cluster_label(design, size), []).append(size)

                with st.expander(f"{design} cluster prices", expanded=True):
                    price_cols = st.columns(min(3, max(1, len(clusters))))
                    for idx, (cluster_label, cluster_sizes) in enumerate(clusters.items()):
                        first_combo = {"design": design, "size": cluster_sizes[0]}
                        cluster_key = (
                            f"{review_key_prefix}_cluster_price_"
                            f"{safe_widget_key_part(design)}_{safe_widget_key_part(cluster_label)}"
                        )
                        if cluster_key not in st.session_state:
                            st.session_state[cluster_key] = float(
                                get_variant_price_from_map(profile, existing_prices, first_combo, fallback=0) or 0
                            )
                        with price_cols[idx % len(price_cols)]:
                            st.number_input(
                                f"{cluster_label} price",
                                min_value=0.0,
                                step=0.5,
                                format="%.2f",
                                key=cluster_key,
                            )
                            st.caption(", ".join(cluster_sizes))
        else:
            for design in selected_designs:
                valid_sizes = [
                    size
                    for size in selected_sizes
                    if not design_size_map.get(design) or size in design_size_map.get(design, [])
                ]
                if not valid_sizes:
                    continue
                with st.expander(f"{design} prices", expanded=True):
                    price_cols = st.columns(min(4, max(1, len(valid_sizes))))
                    for idx, size in enumerate(valid_sizes):
                        combo = {"design": design, "size": size}
                        variant_price_key = build_variant_price_key(combo)
                        widget_key = f"{review_key_prefix}_price_{safe_widget_key_part(variant_price_key)}"
                        if widget_key not in st.session_state:
                            st.session_state[widget_key] = float(
                                get_variant_price_from_map(profile, existing_prices, combo, fallback=0) or 0
                            )
                        with price_cols[idx % len(price_cols)]:
                            st.number_input(
                                str(size),
                                min_value=0.0,
                                step=0.5,
                                format="%.2f",
                                key=widget_key,
                            )
    else:
        price_sizes = list(selected_variants.get("size", [])) or list(edits["size_price_map"].keys()) or ["default"]
        st.markdown("**Price by size**")
        price_cols = st.columns(min(4, max(1, len(price_sizes))))
        for idx, size in enumerate(price_sizes):
            price_key = f"{review_key_prefix}_price_{safe_widget_key_part(size)}"
            if price_key not in st.session_state:
                st.session_state[price_key] = float(edits["size_price_map"].get(str(size), 0) or 0)
            with price_cols[idx % len(price_cols)]:
                st.number_input(
                    str(size),
                    min_value=0.0,
                    step=0.5,
                    format="%.2f",
                    key=price_key,
                )

    edits = get_review_sku_and_price_edits(review_key_prefix, listing_memory, profile)
    render_variant_combinations_preview(
        profile=profile,
        parent_sku=edits["parent_sku"],
        selected_variants=selected_variants,
        base_title=str(listing_memory.get("title", "")),
        sku_decoration_code=edits["sku_decoration_code"],
        sku_listing_code=edits["sku_listing_code"],
    )
    return edits


def append_workflow_event(
    payload: dict[str, Any],
    action: str,
    actor: str = "",
    from_state: str = "",
    to_state: str = "",
    folder_path: str = "",
    details: dict[str, Any] | None = None,
) -> None:
    events = list(payload.get("workflow_events", []))

    event = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "action": action,
        "actor": str(actor or ""),
        "from_state": str(from_state or ""),
        "to_state": str(to_state or ""),
        "folder_path": str(folder_path or ""),
    }

    if details:
        event["details"] = dict(details)

    events.append(event)
    payload["workflow_events"] = events[-100:]


def build_review_snapshot(
    profile: dict[str, Any],
    payload: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    folder_path: str,
    quality_report: dict[str, Any] | None = None,
    preview_errors: list[str] | None = None,
) -> dict[str, Any]:
    selected_variants = dict(payload.get("selected_variants", {}))
    size_price_map = dict(payload.get("size_price_map", {}))
    quality_report = dict(quality_report or {})
    preview_errors = list(preview_errors or [])

    try:
        dropbox_overview = get_cached_dropbox_overview(profile, dropbox_cfg)
    except Exception:
        dropbox_overview = {}

    try:
        variant_combos = build_variant_combinations(profile, selected_variants)
    except Exception:
        variant_combos = []

    blockers = list(quality_report.get("blockers", []))
    warnings = list(quality_report.get("warnings", []))

    return {
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "listing_folder_path": str(folder_path or ""),
        "template": profile.get("label", profile.get("_slug", "")) if profile else "",
        "template_key": profile.get("template_key", "") if profile else "",
        "resource_root": str(dropbox_overview.get("resource_root", "") or ""),
        "garment_resource_root": str(dropbox_overview.get("garment_resource_root", "") or ""),
        "variants_summary": build_variants_summary(selected_variants),
        "price_summary": build_price_summary(size_price_map),
        "quantity": normalize_variant_quantity(payload.get("quantity", DEFAULT_VARIANT_QUANTITY)),
        "fulfillment": {
            "handling_time_days": normalize_handling_time_days(
                payload.get("handling_time_days", DEFAULT_HANDLING_TIME_DAYS)
            ),
            "merchant_shipping_group_name": normalize_merchant_shipping_group(
                payload.get("merchant_shipping_group_name", "")
            ),
        },
        "image_summary": {
            "selected_color_count": len(get_selected_colors_for_image_resolution(profile, selected_variants)) if profile else 0,
            "selected_size_count": len(selected_variants.get("size", [])),
            "expected_child_variants": len(variant_combos),
            "support_images_configured": len(dropbox_overview.get("garment_resource_images", [])) + len(dropbox_overview.get("shared_resource_images", [])),
            "preview_error_count": len(preview_errors),
        },
        "quality_summary": {
            "score": quality_report.get("score", 0),
            "blocker_count": len(blockers),
            "warning_count": len(warnings),
        },
        "prepared_by": {
            "assets_prepared_by": str(payload.get("assets_prepared_by", "") or ""),
            "content_prepared_by": str(payload.get("content_prepared_by", "") or ""),
        },
    }


def build_ready_review_data(
    profile: dict[str, Any] | None,
    listing_memory: dict[str, Any],
    ready_folder_name: str,
    dropbox_cfg: dict[str, Any],
    source_folder_path: str | None = None,
    include_images: bool = False,
    include_quality: bool = False,
) -> dict[str, Any]:
    review_data = {
        "folder_name": ready_folder_name,
        "template": listing_memory.get("template_label", "") or listing_memory.get("template_slug", "") or "Unknown",
        "assets_prepared_by": listing_memory.get("assets_prepared_by", ""),
        "content_prepared_by": listing_memory.get("content_prepared_by", ""),
        "reviewed_by": listing_memory.get("reviewed_by", ""),
        "prepared_at": listing_memory.get("prepared_at", ""),
        "reviewed_at": listing_memory.get("reviewed_at", ""),
        "title": listing_memory.get("title", ""),
        "bullet_points": (list(listing_memory.get("bullet_points", [])) + ["", "", "", "", ""])[:5],
        "product_description": listing_memory.get("product_description", ""),
        "generic_keywords": listing_memory.get("generic_keywords", ""),
        "variants_summary": build_variants_summary(dict(listing_memory.get("selected_variants", {}))),
        "quantity": normalize_variant_quantity(listing_memory.get("quantity", DEFAULT_VARIANT_QUANTITY)),
        "price_summary": build_price_summary(dict(listing_memory.get("size_price_map", {}))),
        "parent_main_image_url": "",
        "support_images": [],
        "child_image_rows": [],
        "quality_report": {"blockers": [], "warnings": [], "score": 0, "breakdown": {}},
        "errors": [],
        "image_review_loaded": include_images or include_quality,
        "quality_check_loaded": include_quality,
        "review_snapshot": dict(listing_memory.get("review_snapshot", {})) if isinstance(listing_memory.get("review_snapshot"), dict) else {},
        "workflow_events": list(listing_memory.get("workflow_events", [])) if isinstance(listing_memory.get("workflow_events"), list) else [],
    }

    if not profile:
        review_data["errors"].append("Template profile could not be resolved for this ready listing.")
        return review_data

    review_data["template"] = profile.get("label", profile.get("_slug", review_data["template"]))

    bullets = review_data["bullet_points"]
    selected_variants = dict(listing_memory.get("selected_variants", {}))
    size_price_map = {
        str(size): float(price)
        for size, price in dict(listing_memory.get("size_price_map", {})).items()
    }

    generation_prep = prepare_generation_payload(
        profile=profile,
        title=str(listing_memory.get("title", "")),
        bullets=bullets,
        product_description=str(listing_memory.get("product_description", "")),
        generic_keywords=str(listing_memory.get("generic_keywords", "")),
        selected_variants=selected_variants,
        size_price_map=size_price_map,
        sku_decoration_code=get_default_sku_decoration_code(profile, listing_memory),
        sku_listing_code=str(listing_memory.get("sku_listing_code", "") or get_saved_generated_sku_listing_code(listing_memory)),
        manual_sku_listing_code=str(listing_memory.get("manual_sku_listing_code", "") or ""),
        generated_sku_listing_code=get_saved_generated_sku_listing_code(listing_memory),
        quantity=review_data["quantity"],
        staged_folder_name=ready_folder_name,
        handling_time_days=normalize_handling_time_days(
            listing_memory.get("handling_time_days", DEFAULT_HANDLING_TIME_DAYS)
        ),
        merchant_shipping_group_name=normalize_merchant_shipping_group(
            listing_memory.get("merchant_shipping_group_name", "")
        ),
        parent_main_image_choice=str(listing_memory.get("parent_main_image_choice", "") or ""),
        parent_main_image_url=str(listing_memory.get("parent_main_image_url", "") or ""),
        parent_sku_override=get_listing_generation_parent_sku_override(
            listing_memory,
            profile,
        ),
    )

    review_data["errors"].extend(generation_prep["errors"])
    payload = dict(generation_prep["payload"])
    selected_variants = dict(payload.get("selected_variants", {}))

    if not (include_images or include_quality):
        return review_data

    dropbox_overview = get_cached_dropbox_overview(profile, dropbox_cfg)
    ready_folder_path = source_folder_path or build_ready_folder_path(dropbox_cfg, ready_folder_name)
    selected_colors = payload.get("colors", [])

    try:
        (
            payload["parent_main_image_url"],
            payload["other_images"],
            payload["color_image_map"],
            payload["design_color_image_url_map"],
        ) = resolve_folder_image_urls(
            profile,
            selected_variants,
            selected_colors,
            dropbox_overview,
            ready_folder_path,
            selected_parent_main_image_label=str(payload.get("parent_main_image_choice", "") or ""),
            selected_parent_main_image_url=str(payload.get("selected_parent_main_image_url", "") or payload.get("parent_main_image_url", "") or ""),
        )
    except Exception as exc:
        review_data["errors"].append(str(exc))

    review_data["parent_main_image_url"] = payload.get("parent_main_image_url", "")
    review_data["support_images"] = [
        {
            "label": f"{idx}. {Path(image_url).name}",
            "filename": Path(image_url).name,
            "url": image_url,
        }
        for idx, image_url in enumerate(payload.get("other_images", []), start=1)
        if image_url
    ]

    child_image_rows: list[dict[str, str]] = []
    color_image_map = payload.get("color_image_map", {}) or {}
    design_color_image_url_map = payload.get("design_color_image_url_map", {}) or {}

    for color, image_url in color_image_map.items():
        child_image_rows.append({
            "variant": color,
            "filename": Path(image_url).name if image_url else "",
            "url": image_url,
        })

    for color, design_map in design_color_image_url_map.items():
        for design, image_url in design_map.items():
            child_image_rows.append({
                "variant": f"{color} / {design}",
                "filename": Path(image_url).name if image_url else "",
                "url": image_url,
            })

    review_data["child_image_rows"] = child_image_rows

    if include_quality:
        review_data["quality_report"] = validate_listing_quality(profile, payload)

    return review_data


def render_ready_review_panel(
    item: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    key_prefix: str = "ready_review",
    source_folder_path: str | None = None,
) -> None:
    folder_key = str(item.get("folder_name", "listing")).replace("/", "_").replace("\\", "_").replace(" ", "_")
    review_key_prefix = f"{key_prefix}_{folder_key}"
    listing_memory = dict(item.get("listing_memory", {}) or {})
    profile = item.get("profile")
    memory_fingerprint = build_review_memory_fingerprint(listing_memory, profile) if listing_memory else "empty"
    editor_key_prefix = f"{review_key_prefix}_{memory_fingerprint}"
    st.session_state[f"{review_key_prefix}_active_editor_prefix"] = editor_key_prefix
    image_review_loaded = bool(st.session_state.get(f"{review_key_prefix}_load_image_review", False))
    quality_check_loaded = bool(st.session_state.get(f"{review_key_prefix}_run_full_quality", False))

    review_data = build_ready_review_data(
        profile=item.get("profile"),
        listing_memory=item.get("listing_memory", {}),
        ready_folder_name=item.get("folder_name", ""),
        dropbox_cfg=dropbox_cfg,
        source_folder_path=source_folder_path,
        include_images=image_review_loaded or quality_check_loaded,
        include_quality=quality_check_loaded,
    )

    reload_col, _ = st.columns([1, 4])
    with reload_col:
        if st.button(
            "Reload from listing_inputs.json",
            key=f"{review_key_prefix}_reload_listing_inputs",
            width="stretch",
        ):
            clear_review_editor_state(review_key_prefix)
            clear_cached_listing_memory(source_folder_path or "")
            st.session_state.pop("ready_queue_items_cache", None)
            st.session_state.pop("approved_queue_items_cache", None)
            st.rerun()

    review_sections = ["Overview", "Content", "SKU & Price", "Images", "Quality"]
    active_review_section = st.segmented_control(
        "Review section",
        review_sections,
        key=f"{review_key_prefix}_active_section",
        selection_mode="single",
        width="stretch",
        label_visibility="collapsed",
    )

    if active_review_section == "Overview":
        col1, col2 = st.columns(2)
        with col1:
            st.write(f"Folder: `{review_data['folder_name']}`")
            st.write(f"Template: `{review_data['template']}`")
            st.write(f"Assets prepared by: `{review_data['assets_prepared_by'] or '-'}`")
            st.write(f"Content prepared by: `{review_data['content_prepared_by'] or '-'}`")
            st.write(f"Reviewed by: `{review_data['reviewed_by'] or '-'}`")
        with col2:
            st.write(f"Prepared at: `{review_data['prepared_at'] or '-'}`")
            st.write(f"Reviewed at: `{review_data['reviewed_at'] or '-'}`")
            st.write(f"Variants: {review_data['variants_summary']}")
            st.write(f"Quantity: {review_data['quantity']}")
            st.write(
                "Shipping group: "
                f"`{normalize_merchant_shipping_group(listing_memory.get('merchant_shipping_group_name', '')) or '-'}`"
            )
            st.write(f"Pricing: {review_data['price_summary']}")

        price_rows = build_review_price_rows(item.get("profile"), item.get("listing_memory", {}))
        if price_rows:
            missing_price_count = sum(1 for row in price_rows if row.get("Status") == "Missing")
            if missing_price_count:
                st.warning(f"{missing_price_count} price key(s) are missing or zero.")
            with st.expander("Price overview", expanded=False):
                st.dataframe(price_rows, width="stretch", hide_index=True)

        snapshot = review_data.get("review_snapshot", {}) or {}
        if snapshot:
            st.markdown("**Review snapshot**")
            image_summary = snapshot.get("image_summary", {}) or {}
            quality_summary = snapshot.get("quality_summary", {}) or {}
            st.caption(
                f"Snapshot: {snapshot.get('created_at', '-')} | "
                f"Expected child variants: {image_summary.get('expected_child_variants', '-')} | "
                f"Support images configured: {image_summary.get('support_images_configured', '-')} | "
                f"Quality blockers: {quality_summary.get('blocker_count', 0)} | "
                f"Warnings: {quality_summary.get('warning_count', 0)}"
            )

        workflow_events = review_data.get("workflow_events", []) or []
        if workflow_events:
            with st.expander("Workflow history", expanded=False):
                st.dataframe(workflow_events[-20:], hide_index=True, width="stretch")

    if active_review_section == "Content":
        if not profile or not listing_memory:
            st.warning("This listing could not be loaded for content review.")
        else:
            st.caption("Edit content here using the current listing values as placeholders.")
            content_edits = render_review_content_editor(
                review_key_prefix=editor_key_prefix,
                listing_memory=listing_memory,
            )
            if source_folder_path and st.button(
                "Save content edits",
                key=f"{review_key_prefix}_save_content_edits",
                width="content",
            ):
                payload = apply_review_content_edits(listing_memory, content_edits)
                try:
                    save_listing_inputs_json_to_dropbox(
                        profile=profile,
                        payload=payload,
                        folder_path=source_folder_path,
                    )
                    st.session_state.pop("ready_queue_items_cache", None)
                    st.session_state.pop("approved_queue_items_cache", None)
                    clear_review_editor_state(review_key_prefix)
                    item["listing_memory"] = payload
                    set_workflow_flash("success", "Saved content edits.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"Could not save content edits: {exc}")

    if active_review_section == "SKU & Price":
        if not profile or not listing_memory:
            st.warning("This listing could not be loaded for SKU and price review.")
        else:
            loaded_prices = dict(listing_memory.get("size_price_map", {}) or {})
            positive_price_count = 0
            for raw_price in loaded_prices.values():
                try:
                    if float(raw_price) > 0:
                        positive_price_count += 1
                except (TypeError, ValueError):
                    pass
            saved_listing_code = (
                listing_memory.get("manual_sku_listing_code")
                or listing_memory.get("sku_listing_code")
                or listing_memory.get("generated_sku_listing_code")
                or "-"
            )
            st.caption(
                f"Loaded from listing_inputs.json: `{positive_price_count}/{len(loaded_prices)}` positive price key(s); "
                f"saved listing code: `{saved_listing_code}`"
            )
            st.caption("Reviewer edits here are saved when the listing is approved or denied.")
            review_edits = render_review_sku_price_editor(
                review_key_prefix=editor_key_prefix,
                profile=profile,
                listing_memory=listing_memory,
            )
            if source_folder_path and st.button(
                "Save SKU, shipping, and price edits",
                key=f"{review_key_prefix}_save_sku_price_edits",
                width="content",
            ):
                payload = apply_review_sku_and_price_edits(listing_memory, review_edits)
                try:
                    save_listing_inputs_json_to_dropbox(
                        profile=profile,
                        payload=payload,
                        folder_path=source_folder_path,
                    )
                    st.session_state.pop("ready_queue_items_cache", None)
                    st.session_state.pop("approved_queue_items_cache", None)
                    clear_review_editor_state(review_key_prefix)
                    item["listing_memory"] = payload
                    set_workflow_flash("success", "Saved SKU, shipping, and price edits.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"Could not save SKU and price edits: {exc}")

    if active_review_section == "Images":
        dropbox_overview = get_cached_dropbox_overview(item.get("profile", {}), dropbox_cfg)
        render_dropbox_folder_links(source_folder_path, dropbox_overview)

        st.markdown("**Add resource image**")
        uploaded_resource_image = st.file_uploader(
            "Resource image",
            type=["png", "jpg", "jpeg"],
            accept_multiple_files=False,
            key=f"{review_key_prefix}_resource_image_upload",
            help="Uploads this image to the selected listing's resources folder.",
        )
        if st.button(
            "Upload resource image",
            key=f"{review_key_prefix}_resource_image_upload_btn",
            width="content",
            disabled=not bool(source_folder_path and uploaded_resource_image),
        ):
            try:
                uploaded_path = upload_ready_review_resource_image(
                    source_folder_path or "",
                    uploaded_resource_image.name,
                    uploaded_resource_image.getvalue(),
                )
                clear_resource_image_caches()
                st.session_state[f"{review_key_prefix}_load_image_review"] = True
                review_data = build_ready_review_data(
                    profile=item.get("profile"),
                    listing_memory=item.get("listing_memory", {}),
                    ready_folder_name=item.get("folder_name", ""),
                    dropbox_cfg=dropbox_cfg,
                    source_folder_path=source_folder_path,
                    include_images=True,
                    include_quality=quality_check_loaded,
                )
                st.success(f"Uploaded resource image: {Path(uploaded_path).name}")
            except Exception as exc:
                st.error(f"Could not upload resource image: {exc}")

        if not review_data["image_review_loaded"]:
            st.info("Image mappings are not loaded for this review yet. Use the Dropbox folder links above for normal review.")
            if st.button("Load image review", key=f"{review_key_prefix}_load_image_review_btn", width="content"):
                st.session_state["active_perf_action_label"] = "load image review"
                st.session_state[f"{review_key_prefix}_load_image_review"] = True
                review_data = build_ready_review_data(
                    profile=item.get("profile"),
                    listing_memory=item.get("listing_memory", {}),
                    ready_folder_name=item.get("folder_name", ""),
                    dropbox_cfg=dropbox_cfg,
                    source_folder_path=source_folder_path,
                    include_images=True,
                    include_quality=quality_check_loaded,
                )

        if review_data["image_review_loaded"]:
            support_images = review_data.get("support_images", [])
            child_image_rows = review_data.get("child_image_rows", [])

            st.success("Image review data loaded.")

            st.markdown("**Parent main image**")
            if review_data["parent_main_image_url"]:
                st.image(review_data["parent_main_image_url"], width=240)
                st.caption(Path(review_data["parent_main_image_url"]).name)
            else:
                st.caption("No resolved parent main image.")

            st.markdown("**Support image order**")
            if support_images:
                cols = st.columns(min(4, len(support_images)))
                for idx, image_entry in enumerate(support_images):
                    with cols[idx % len(cols)]:
                        st.image(image_entry["url"], width=170)
                        st.caption(image_entry["label"])
            else:
                st.caption("No support images found.")

            st.markdown("**Child variant image mapping**")
            if child_image_rows:
                cols_per_row = 3
                cols = st.columns(cols_per_row)
                for idx, image_entry in enumerate(child_image_rows):
                    with cols[idx % cols_per_row]:
                        st.markdown(f"**{image_entry['variant']}**")
                        if image_entry.get("url"):
                            st.image(image_entry["url"], width=180)
                            st.caption(image_entry.get("filename", ""))
                        else:
                            st.caption("No resolved image URL.")
            else:
                st.caption("No child image mappings found.")

            with st.expander("Technical image URLs and filenames", expanded=False):
                st.markdown("**Parent main image URL**")
                if review_data["parent_main_image_url"]:
                    st.code(review_data["parent_main_image_url"], language=None)
                else:
                    st.caption("No resolved parent main image.")

                st.markdown("**Support image order**")
                if support_images:
                    for image_entry in support_images:
                        st.write(image_entry["label"])
                        st.code(image_entry["url"], language=None)
                else:
                    st.caption("No support images found.")

                st.markdown("**Child variant image mapping**")
                if child_image_rows:
                    st.dataframe(child_image_rows, width="stretch", hide_index=True)
                else:
                    st.caption("No child image mappings found.")

    if active_review_section == "Quality":
        if not review_data["quality_check_loaded"]:
            snapshot = review_data.get("review_snapshot", {}) or {}
            quality_summary = snapshot.get("quality_summary", {}) or {}
            image_summary = snapshot.get("image_summary", {}) or {}

            if snapshot:
                st.markdown("**Saved review snapshot**")
                q_col1, q_col2, q_col3, q_col4 = st.columns(4)
                q_col1.metric("Score", quality_summary.get("score", 0))
                q_col2.metric("Blockers", quality_summary.get("blocker_count", 0))
                q_col3.metric("Warnings", quality_summary.get("warning_count", 0))
                q_col4.metric("Expected variants", image_summary.get("expected_child_variants", 0))

                st.caption(
                    f"Snapshot created: {snapshot.get('created_at', '-')} | "
                    f"Support images configured: {image_summary.get('support_images_configured', 0)} | "
                    f"Preview errors at submit: {image_summary.get('preview_error_count', 0)}"
                )

                if int(quality_summary.get("blocker_count", 0) or 0) == 0:
                    st.success("Saved submit-time quality snapshot has no blockers.")
                else:
                    st.error("Saved submit-time quality snapshot had blockers.")
            else:
                st.info("No saved review snapshot found for this older listing.")

            st.caption(
                "Full image quality check is optional. It loads current Dropbox image mappings again "
                "only when you need a deeper live verification."
            )
            if st.button("Run full live image quality check", key=f"{review_key_prefix}_run_full_quality_btn", width="content"):
                st.session_state["active_perf_action_label"] = "run full live image quality check"
                st.session_state[f"{review_key_prefix}_run_full_quality"] = True
                review_data = build_ready_review_data(
                    profile=item.get("profile"),
                    listing_memory=item.get("listing_memory", {}),
                    ready_folder_name=item.get("folder_name", ""),
                    dropbox_cfg=dropbox_cfg,
                    source_folder_path=source_folder_path,
                    include_images=True,
                    include_quality=True,
                )

        if review_data["quality_check_loaded"]:
            if review_data["errors"]:
                st.error("Preflight issues found")
                for error in review_data["errors"]:
                    st.write(f"- {error}")
            else:
                st.success("No preflight issues found.")

            blockers = review_data["quality_report"].get("blockers", [])
            warnings = review_data["quality_report"].get("warnings", [])

            st.markdown("**Quality blockers**")
            if blockers:
                for blocker in blockers:
                    st.write(f"- {blocker}")
            else:
                st.write("None")

            st.markdown("**Quality warnings**")
            if warnings:
                for warning in warnings:
                    st.write(f"- {warning}")
            else:
                st.write("None")


def build_queue_items(
    folder_names: list[str],
    profiles: list[dict[str, Any]],
    dropbox_cfg: dict[str, Any],
    folder_path_builder: Callable[[dict[str, Any], str], str],
    ready_label: str,
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []

    for folder_name in folder_names:
        folder_path = folder_path_builder(dropbox_cfg, folder_name)
        load_error = ""
        listing_memory: dict[str, Any] = {}

        try:
            listing_memory = load_listing_memory_from_dropbox(folder_path)
        except Exception as exc:
            load_error = str(exc)

        profile = find_profile_for_listing_memory(profiles, listing_memory) if listing_memory else None
        template_label = (
            profile.get("label", profile.get("_slug", ""))
            if profile else
            listing_memory.get("template_label", "") or listing_memory.get("template_slug", "") or "Unknown"
        )
        selected_variants = listing_memory.get("selected_variants", {}) if listing_memory else {}

        items.append({
            "folder_name": folder_name,
            "template": template_label,
            "title": listing_memory.get("title", "") if listing_memory else "",
            "variants_summary": build_variants_summary(selected_variants),
            "load_status": ready_label if listing_memory and not load_error else "Missing or invalid inputs",
            "profile": profile,
            "listing_memory": listing_memory,
            "load_error": load_error,
        })

    return items


def build_ready_queue_items(
    ready_folder_names: list[str],
    profiles: list[dict[str, Any]],
    dropbox_cfg: dict[str, Any],
) -> list[dict[str, Any]]:
    return [
        item
        for item in build_queue_items(
            ready_folder_names,
            profiles,
            dropbox_cfg,
            build_ready_folder_path,
            "Ready for approval",
        )
        if is_ready_listing_visible(item.get("listing_memory", {}))
    ]


def build_approved_queue_items(
    approved_folder_names: list[str],
    profiles: list[dict[str, Any]],
    dropbox_cfg: dict[str, Any],
) -> list[dict[str, Any]]:
    return build_queue_items(
        approved_folder_names,
        profiles,
        dropbox_cfg,
        build_approved_folder_path,
        "Approved",
    )


CHRISTMAS_GROUP_MEMBER_LABELS = {
    "tshirt": "T-Shirt",
    "sweatshirt": "Sweatshirt",
    "hoodie": "Hoodie",
}


def classify_finished_listing_origin(listing_memory: dict[str, Any]) -> tuple[str, str, str]:
    source_group = listing_memory.get("source_group")
    source_group = source_group if isinstance(source_group, dict) else {}
    if str(source_group.get("group_type", "") or "").strip().casefold() == "christmas_project":
        member_key = str(source_group.get("member_key", "") or "").strip().casefold()
        member_label = CHRISTMAS_GROUP_MEMBER_LABELS.get(member_key, member_key or "Unknown")
        task_id = str(source_group.get("task_id", "") or "").strip()
        return "Grouped Christmas", member_label, task_id

    template_key = str(listing_memory.get("template_key", "") or "").strip().upper()
    if template_key in {"GENERIC_SHIRTS", "GENERIC_SWEATSHIRTS", "GENERIC_HOODIES"}:
        return "Generic listing", "", ""
    if template_key == "CP":
        return "Christmas Project (single)", "", ""
    return "Standard listing", "", ""


def build_finished_generation_history_rows(
    finished_folder_names: list[str],
    profiles: list[dict[str, Any]],
    dropbox_cfg: dict[str, Any],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for folder_name in finished_folder_names:
        folder_path = build_finished_folder_path(dropbox_cfg, folder_name)
        try:
            listing_memory = load_listing_memory_from_dropbox(folder_path)
            if not listing_memory:
                raise FileNotFoundError("listing_inputs.json is missing or empty.")

            profile = find_profile_for_listing_memory(profiles, listing_memory)
            template_label = str(
                (profile or {}).get("label")
                or listing_memory.get("template_label")
                or listing_memory.get("template_key")
                or "Unknown"
            )
            generated_outputs = listing_memory.get("generated_outputs")
            generated_outputs = generated_outputs if isinstance(generated_outputs, list) else []
            latest_output = next(
                (output for output in reversed(generated_outputs) if isinstance(output, dict)),
                {},
            )
            sku_manifest = listing_memory.get("sku_manifest")
            sku_manifest = sku_manifest if isinstance(sku_manifest, dict) else {}
            generated_at = str(
                latest_output.get("created_at")
                or sku_manifest.get("created_at")
                or ""
            ).strip()
            origin, christmas_member, task_id = classify_finished_listing_origin(listing_memory)
            history_group = task_id or (generated_at[:16] if generated_at else "Unknown time")
            rows.append({
                "folder_name": folder_name,
                "generated_at": generated_at,
                "generated_date": generated_at[:10] if len(generated_at) >= 10 else "Unknown",
                "history_group": history_group,
                "origin": origin,
                "christmas_member": christmas_member,
                "christmas_task_id": task_id,
                "template": template_label,
                "parent_sku": str(listing_memory.get("parent_sku", "") or ""),
                "title": str(listing_memory.get("title", "") or ""),
                "workbook": str(
                    latest_output.get("workbook_name")
                    or sku_manifest.get("output_workbook_name")
                    or ""
                ),
                "generation_status": str(listing_memory.get("generation_status", "") or "Finished"),
                "load_status": "Loaded",
            })
        except Exception as exc:
            rows.append({
                "folder_name": folder_name,
                "generated_at": "",
                "generated_date": "Unknown",
                "history_group": "Unknown time",
                "origin": "Unknown",
                "christmas_member": "",
                "christmas_task_id": "",
                "template": "Unknown",
                "parent_sku": "",
                "title": "",
                "workbook": "",
                "generation_status": "Unknown",
                "load_status": f"Could not load: {exc}",
            })

    return sorted(
        rows,
        key=lambda row: (str(row.get("generated_at", "")), str(row.get("folder_name", ""))),
        reverse=True,
    )



def build_sku_manifest(
    profile: dict[str, Any],
    payload: dict[str, Any],
    finished_folder_path: str,
    output_path: Path | None = None,
) -> dict[str, Any]:
    selected_variants = dict(payload.get("selected_variants", {}))
    size_price_map = dict(payload.get("size_price_map", {}))
    parent_sku = str(payload.get("parent_sku", "") or "").strip()
    variant_combos = build_variant_combinations(profile, selected_variants)
    variant_combos = sort_variant_combinations_by_price(profile, variant_combos, size_price_map)
    effective_profile = apply_sku_context_to_profile(
        profile,
        payload.get("sku_decoration_code", ""),
        payload.get("sku_listing_code", ""),
    )

    children: list[dict[str, Any]] = []

    for variant_values in variant_combos:
        sku_details = build_child_sku_details(effective_profile, parent_sku, variant_values)
        seller_sku = sku_details["amazon_seller_sku"]
        size_value = str(variant_values.get("size", "") or "")
        price_value = get_variant_price_from_map(profile, size_price_map, variant_values, fallback="")

        child_row: dict[str, Any] = {
            "amazon_seller_sku": seller_sku,
            "canonical_sku": sku_details.get("canonical_sku", seller_sku),
            "parent_sku": parent_sku,
            "supplier": sku_details.get("supplier", ""),
            "supplier_stock_key": sku_details.get("supplier_stock_key", ""),
            "supplier_stock_key_status": sku_details.get("supplier_stock_key_status", ""),
            "supplier_stock_key_reason": sku_details.get("supplier_stock_key_reason", ""),
            "design_or_listing_code": sku_details.get("design_or_listing_code", ""),
            "variant_values": dict(variant_values),
            "quantity": normalize_variant_quantity(payload.get("quantity", DEFAULT_VARIANT_QUANTITY)),
            "price": price_value,
        }

        for variant_key, variant_value in variant_values.items():
            child_row[str(variant_key)] = variant_value

        children.append(child_row)

    suppliers = sorted({
        str(child.get("supplier", "") or "")
        for child in children
        if str(child.get("supplier", "") or "")
    })
    missing_supplier_stock_key_count = sum(
        1
        for child in children
        if child.get("supplier_stock_key_status") == "missing"
    )

    return {
        "schema_version": 2,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "marketplace": "amazon",
        "template_label": profile.get("label", profile.get("_slug", "")),
        "template_slug": profile.get("_slug", ""),
        "template_key": profile.get("template_key", ""),
        "stock_reference_key": profile.get("stock_reference_key", ""),
        "sku_decoration_code": payload.get("sku_decoration_code", ""),
        "strict_stock_ready": is_strict_stock_ready(profile),
        "suppliers": suppliers,
        "parent_sku": parent_sku,
        "finished_folder_path": str(finished_folder_path or ""),
        "output_workbook_name": output_path.name if output_path else "",
        "child_sku_count": len(children),
        "missing_supplier_stock_key_count": missing_supplier_stock_key_count,
        "children": children,
    }


def save_generated_artifacts_to_dropbox(
    profile: dict[str, Any],
    payload: dict[str, Any],
    finished_folder_path: str,
    output_path: Path,
) -> dict[str, Any]:
    finished_folder_path = str(finished_folder_path or "").rstrip("/")
    if not finished_folder_path:
        raise ValueError("Finished folder path is required before saving generated artifacts.")

    workbook_dropbox_path = f"{finished_folder_path}/{output_path.name}"
    with output_path.open("rb") as workbook_file:
        upload_binary_file(workbook_dropbox_path, workbook_file.read())

    sku_manifest = build_sku_manifest(
        profile=profile,
        payload=payload,
        finished_folder_path=finished_folder_path,
        output_path=output_path,
    )

    sku_manifest_path = f"{finished_folder_path}/sku_manifest.json"
    upload_text_file(
        sku_manifest_path,
        json.dumps(sku_manifest, indent=2, ensure_ascii=False),
    )

    artifact = {
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "workbook_name": output_path.name,
        "workbook_dropbox_path": workbook_dropbox_path,
        "sku_manifest_dropbox_path": sku_manifest_path,
        "child_sku_count": sku_manifest.get("child_sku_count", 0),
        "missing_supplier_stock_key_count": sku_manifest.get("missing_supplier_stock_key_count", 0),
    }

    payload["sku_manifest"] = sku_manifest
    generated_outputs = list(payload.get("generated_outputs", []))
    generated_outputs.append(artifact)
    payload["generated_outputs"] = generated_outputs[-50:]

    return artifact

def generate_approved_listing(
    profile: dict[str, Any],
    listing_memory: dict[str, Any],
    approved_folder_name: str,
    dropbox_cfg: dict[str, Any],
) -> dict[str, Any]:
    if not profile:
        raise ValueError("Could not find a template profile for this approved listing.")

    title = str(listing_memory.get("title", ""))
    bullets = list(listing_memory.get("bullet_points", []))
    bullets = (bullets + ["", "", "", "", ""])[:5]
    product_description = str(listing_memory.get("product_description", ""))
    generic_keywords = str(listing_memory.get("generic_keywords", ""))
    selected_variants = dict(listing_memory.get("selected_variants", {}))
    size_price_map = {
        str(size): float(price)
        for size, price in dict(listing_memory.get("size_price_map", {})).items()
    }
    quantity = normalize_variant_quantity(listing_memory.get("quantity", DEFAULT_VARIANT_QUANTITY))

    generation_prep = prepare_generation_payload(
        profile=profile,
        title=title,
        bullets=bullets,
        product_description=product_description,
        generic_keywords=generic_keywords,
        selected_variants=selected_variants,
        size_price_map=size_price_map,
        sku_decoration_code=get_default_sku_decoration_code(profile, listing_memory),
        sku_listing_code=str(listing_memory.get("sku_listing_code", "") or get_saved_generated_sku_listing_code(listing_memory)),
        manual_sku_listing_code=str(listing_memory.get("manual_sku_listing_code", "") or ""),
        generated_sku_listing_code=get_saved_generated_sku_listing_code(listing_memory),
        quantity=quantity,
        staged_folder_name=approved_folder_name,
        handling_time_days=normalize_handling_time_days(
            listing_memory.get("handling_time_days", DEFAULT_HANDLING_TIME_DAYS)
        ),
        merchant_shipping_group_name=normalize_merchant_shipping_group(
            listing_memory.get("merchant_shipping_group_name", "")
        ),
        parent_main_image_choice=str(listing_memory.get("parent_main_image_choice", "") or ""),
        parent_main_image_url=str(listing_memory.get("parent_main_image_url", "") or ""),
        parent_sku_override=get_listing_generation_parent_sku_override(
            listing_memory,
            profile,
        ),
    )
    generation_payload = generation_prep["payload"]
    selected_variants = dict(generation_payload.get("selected_variants", {}))
    if "mpn" in listing_memory:
        generation_payload["mpn"] = listing_memory.get("mpn")
    original_finished_folder_name = str(listing_memory.get("original_finished_folder_name", "")).strip()
    if original_finished_folder_name:
        generation_payload["original_finished_folder_name"] = original_finished_folder_name
    generation_payload["assets_prepared_by"] = listing_memory.get("assets_prepared_by", "")
    generation_payload["content_prepared_by"] = listing_memory.get("content_prepared_by", "")
    generation_payload["reviewed_by"] = listing_memory.get("reviewed_by", "")
    generation_payload["prepared_at"] = listing_memory.get("prepared_at", "")
    generation_payload["reviewed_at"] = listing_memory.get("reviewed_at", "") or format_workflow_timestamp()
    if isinstance(listing_memory.get("review_snapshot"), dict):
        generation_payload["review_snapshot"] = dict(listing_memory.get("review_snapshot", {}))
    if isinstance(listing_memory.get("workflow_events"), list):
        generation_payload["workflow_events"] = list(listing_memory.get("workflow_events", []))
    preserve_grouped_child_generation_context(listing_memory, generation_payload)
    generation_errors = generation_prep["errors"]
    if generation_errors:
        raise ValueError("; ".join(generation_errors))

    dropbox_overview = get_cached_dropbox_overview(profile, dropbox_cfg)
    approved_folder_path = build_approved_folder_path(dropbox_cfg, approved_folder_name)
    selected_colors = generation_payload["colors"]
    selected_parent_main_image_label = str(generation_payload.get("parent_main_image_choice", "") or "")
    selected_parent_main_image_url = str(
        generation_payload.get("selected_parent_main_image_url", "")
        or generation_payload.get("parent_main_image_url", "")
        or ""
    )

    generation_timings: dict[str, float] = {}

    step_started_at = time.perf_counter()
    template_path = resolve_template_path(profile)
    wb = load_workbook(template_path, keep_vba=True, read_only=True)
    wb.close()
    generation_timings["template_check"] = round(time.perf_counter() - step_started_at, 4)

    step_started_at = time.perf_counter()
    resolve_folder_image_urls(
        profile,
        selected_variants,
        selected_colors,
        dropbox_overview,
        approved_folder_path,
        selected_parent_main_image_label=selected_parent_main_image_label,
        selected_parent_main_image_url=selected_parent_main_image_url,
    )
    generation_timings["pre_move_image_check"] = round(time.perf_counter() - step_started_at, 4)

    final_sku = ""
    finished_folder_path = ""

    try:
        step_started_at = time.perf_counter()
        final_sku, finished_folder_path = choose_finished_folder_target(
            dropbox_cfg=dropbox_cfg,
            parent_sku=generation_payload["parent_sku"],
            reuse_finished_folder_name=original_finished_folder_name,
        )
        generation_timings["choose_finished_target"] = round(time.perf_counter() - step_started_at, 4)

        step_started_at = time.perf_counter()
        parent_main_image_url, other_images, color_image_map, design_color_image_url_map = resolve_folder_image_urls(
            profile,
            selected_variants,
            selected_colors,
            dropbox_overview,
            approved_folder_path,
            selected_parent_main_image_label=selected_parent_main_image_label,
            selected_parent_main_image_url=selected_parent_main_image_url,
        )
        generation_timings["image_resolve"] = round(time.perf_counter() - step_started_at, 4)

        payload = dict(generation_payload)
        payload["finished_folder_sku"] = final_sku
        payload["pending_finished_folder_path"] = finished_folder_path
        payload["parent_main_image_url"] = parent_main_image_url
        payload["other_images"] = other_images
        payload["color_image_map"] = color_image_map
        payload["design_color_image_url_map"] = design_color_image_url_map

        step_started_at = time.perf_counter()
        output_path, workbook_timings = build_workbook(profile, payload)
        generation_timings["build_workbook_total"] = round(time.perf_counter() - step_started_at, 4)
        for workbook_step, workbook_seconds in workbook_timings.items():
            generation_timings[f"workbook_{workbook_step}"] = round(float(workbook_seconds), 4)

        if not output_path.exists() or output_path.stat().st_size <= 0:
            raise ValueError(f"Workbook was not created correctly: {output_path.name}")

        step_started_at = time.perf_counter()
        generated_artifact = save_generated_artifacts_to_dropbox(
            profile=profile,
            payload=payload,
            finished_folder_path=approved_folder_path,
            output_path=output_path,
        )
        generation_timings["save_generated_artifacts"] = round(time.perf_counter() - step_started_at, 4)

        append_workflow_event(
            payload,
            action="generate_approved_listing",
            actor=str(payload.get("reviewed_by", "") or ""),
            from_state="approved",
            to_state="approved",
            folder_path=approved_folder_path,
            details={
                "output_name": output_path.name,
                "workbook_dropbox_path": generated_artifact.get("workbook_dropbox_path", ""),
                "sku_manifest_dropbox_path": generated_artifact.get("sku_manifest_dropbox_path", ""),
                "child_sku_count": generated_artifact.get("child_sku_count", 0),
                "missing_supplier_stock_key_count": generated_artifact.get("missing_supplier_stock_key_count", 0),
                "pending_finished_folder_path": finished_folder_path,
            },
        )

        step_started_at = time.perf_counter()
        save_listing_inputs_json_to_dropbox(profile=profile, payload=payload, folder_path=approved_folder_path)
        generation_timings["save_approved_listing_inputs"] = round(time.perf_counter() - step_started_at, 4)

        return {
            "folder_name": approved_folder_name,
            "status": "Success",
            "message": f"Generated {output_path.name}. Moving to Finished before download.",
            "output_path": str(output_path),
            "output_name": output_path.name,
            "approved_folder_name": approved_folder_name,
            "approved_folder_path": approved_folder_path,
            "pending_finished_folder_path": finished_folder_path,
            "pending_finished_folder_sku": final_sku,
            "parent_sku": payload.get("parent_sku", ""),
            "generated_artifact": generated_artifact,
            "timings": generation_timings,
        }
    except Exception as exc:
        raise


def move_generated_approved_listing_to_finished(
    result: dict[str, Any],
    profiles: list[dict[str, Any]],
    dropbox_cfg: dict[str, Any],
) -> dict[str, Any]:
    approved_folder_name = str(result.get("approved_folder_name") or result.get("folder_name") or "").strip()
    if not approved_folder_name or approved_folder_name == "Combined workbook":
        raise ValueError("Only generated single approved listings can be moved to Finished.")

    approved_folder_path = build_approved_folder_path(dropbox_cfg, approved_folder_name)
    if not path_exists(approved_folder_path):
        raise FileNotFoundError(f"Approved folder not found: {approved_folder_path}")

    listing_memory = load_listing_memory_from_dropbox(approved_folder_path)
    if not listing_memory:
        raise ValueError(f"listing_inputs.json could not be loaded from {approved_folder_name}.")

    profile = find_profile_for_listing_memory(profiles, listing_memory)
    if not profile:
        raise ValueError(f"Template profile could not be resolved for {approved_folder_name}.")

    parent_sku = str(listing_memory.get("parent_sku") or result.get("parent_sku") or "").strip()
    if not parent_sku:
        raise ValueError(f"Parent SKU is missing for {approved_folder_name}.")

    reuse_finished_folder_name = str(
        listing_memory.get("finished_folder_sku")
        or result.get("pending_finished_folder_sku")
        or listing_memory.get("original_finished_folder_name")
        or ""
    ).strip()

    final_sku, finished_folder_path = finalize_approved_dropbox_folder(
        dropbox_cfg=dropbox_cfg,
        approved_folder_name=approved_folder_name,
        parent_sku=parent_sku,
        reuse_finished_folder_name=reuse_finished_folder_name,
    )

    payload = dict(listing_memory)
    payload["finished_folder_sku"] = final_sku
    payload["pending_finished_folder_path"] = ""
    generated_outputs = list(payload.get("generated_outputs", []))
    if generated_outputs:
        latest_output = dict(generated_outputs[-1])
        old_workbook_name = Path(str(latest_output.get("workbook_dropbox_path", "") or result.get("output_name", ""))).name
        old_manifest_name = Path(str(latest_output.get("sku_manifest_dropbox_path", "") or "sku_manifest.json")).name
        latest_output["status"] = "finished"
        latest_output["finished_at"] = format_workflow_timestamp()
        latest_output["finished_folder_path"] = finished_folder_path
        if old_workbook_name:
            latest_output["workbook_dropbox_path"] = f"{finished_folder_path}/{old_workbook_name}"
        if old_manifest_name:
            latest_output["sku_manifest_dropbox_path"] = f"{finished_folder_path}/{old_manifest_name}"
        generated_outputs[-1] = latest_output
        payload["generated_outputs"] = generated_outputs

    if isinstance(payload.get("sku_manifest"), dict):
        sku_manifest = dict(payload.get("sku_manifest", {}))
        sku_manifest["finished_folder_path"] = finished_folder_path
        payload["sku_manifest"] = sku_manifest
        upload_text_file(
            f"{finished_folder_path}/sku_manifest.json",
            json.dumps(sku_manifest, indent=2, ensure_ascii=False),
        )

    append_workflow_event(
        payload,
        action="move_generated_approved_to_finished",
        actor=str(payload.get("reviewed_by", "") or ""),
        from_state="approved",
        to_state="finished",
        folder_path=finished_folder_path,
        details={
            "old_approved_folder_name": approved_folder_name,
            "output_name": result.get("output_name", ""),
        },
    )
    save_listing_inputs_json_to_dropbox(profile=profile, payload=payload, folder_path=finished_folder_path)

    return {
        "folder_name": approved_folder_name,
        "status": "Success",
        "message": f"Moved to Finished: {Path(finished_folder_path).name}",
        "finished_folder_path": finished_folder_path,
    }


def move_successful_generation_results_to_finished(
    results: list[dict[str, Any]],
    profiles: list[dict[str, Any]],
    dropbox_cfg: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    updated_results: list[dict[str, Any]] = []
    move_results: list[dict[str, Any]] = []

    for result in results:
        should_move = (
            result.get("status") == "Success"
            and result.get("approved_folder_name")
            and result.get("approved_folder_name") != "Combined workbook"
            and not result.get("finished_folder_path")
        )
        if not should_move:
            updated_results.append(result)
            continue

        try:
            move_result = move_generated_approved_listing_to_finished(
                result=result,
                profiles=profiles,
                dropbox_cfg=dropbox_cfg,
            )
            move_results.append(move_result)
            updated_result = dict(result)
            updated_result["finished_folder_path"] = move_result.get("finished_folder_path", "")
            updated_result["message"] = (
                f"Generated {result.get('output_name', 'workbook')} and moved to Finished. "
                "Download is ready."
            )
            updated_results.append(updated_result)
        except Exception as exc:
            move_result = {
                "folder_name": result.get("approved_folder_name") or result.get("folder_name", ""),
                "status": "Failed",
                "message": str(exc),
            }
            move_results.append(move_result)
            failed_result = dict(result)
            failed_result["status"] = "Failed"
            failed_result["message"] = (
                f"Generated {result.get('output_name', 'workbook')}, but could not move to Finished: {exc}"
            )
            failed_result.pop("output_path", None)
            updated_results.append(failed_result)

    if any(row.get("status") == "Success" for row in move_results):
        refresh_cached_folder_names("approved", "finished")
        st.session_state.pop("approved_queue_items_cache", None)

    return updated_results, move_results


def render_generation_results(
    results: list[dict[str, Any]],
    download_key_prefix: str,
) -> None:
    if not results:
        return

    summary_rows = [
        {
            "folder_name": result.get("folder_name", ""),
            "status": result.get("status", ""),
            "message": result.get("message", ""),
        }
        for result in results
    ]
    st.dataframe(summary_rows, width="stretch", hide_index=True)

    success_results = [
        result for result in results
        if result.get("status") == "Success" and result.get("output_path")
    ]
    if not success_results:
        return

    st.markdown("**Downloads**")
    for result in success_results:
        output_path = Path(result["output_path"])
        if not output_path.exists():
            st.warning(f"Workbook not found for {result.get('folder_name', '')}: {output_path.name}")
            continue

        output_name = result.get("output_name", output_path.name)
        download_key = f"{download_key_prefix}_{result.get('folder_name', '')}_{output_name}"
        with output_path.open("rb") as f:
            st.download_button(
                label=f"Download {output_name}",
                data=f.read(),
                file_name=output_name,
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                key=download_key,
            )


def process_ready_review_decision(
    review_item: dict[str, Any],
    dropbox_cfg: dict[str, Any],
    reviewed_by: str,
    *,
    approve: bool,
    review_edit_prefix: str = "",
) -> dict[str, Any]:
    folder_name = str(review_item.get("folder_name", "") or "")
    profile = review_item.get("profile")
    listing_memory = dict(review_item.get("listing_memory", {}) or {})
    if not folder_name or not profile or not listing_memory:
        raise ValueError("This ready listing could not be loaded for review.")

    ready_folder_path = build_ready_folder_path(dropbox_cfg, folder_name)
    payload = dict(listing_memory)

    if review_edit_prefix:
        if has_complete_review_content_state(review_edit_prefix):
            content_edits = get_review_content_edits(review_edit_prefix)
            payload = apply_review_content_edits(payload, content_edits)

        if has_complete_review_sku_price_state(review_edit_prefix, payload, profile):
            review_edits = get_review_sku_and_price_edits(
                review_edit_prefix,
                payload,
                profile,
            )
            payload = apply_review_sku_and_price_edits(payload, review_edits)

    payload["reviewed_by"] = reviewed_by
    payload["reviewed_at"] = format_workflow_timestamp()
    append_workflow_event(
        payload,
        action="approve_ready_listing" if approve else "deny_ready_listing",
        actor=reviewed_by,
        from_state="ready",
        to_state="approved" if approve else "_stage",
        folder_path=ready_folder_path,
        details={"review_folder": folder_name},
    )

    save_listing_inputs_json_to_dropbox(
        profile=profile,
        payload=payload,
        folder_path=ready_folder_path,
    )

    if approve:
        target_path = move_ready_dropbox_folder_to_approved(
            dropbox_cfg=dropbox_cfg,
            ready_folder_name=folder_name,
            approved_folder_name=folder_name,
        )
        refresh_cached_folder_names("ready", "approved")
        clear_cached_listing_memory(ready_folder_path, target_path)
        return {
            "folder_name": folder_name,
            "status": "Success",
            "action": "approved",
            "message": f"Approved successfully: {Path(target_path).name}",
            "target_path": target_path,
        }

    target_path = move_ready_dropbox_folder_to_denied_stage(
        dropbox_cfg=dropbox_cfg,
        ready_folder_name=folder_name,
    )
    refresh_cached_folder_names("ready", "stage")
    clear_cached_listing_memory(ready_folder_path, target_path)
    return {
        "folder_name": folder_name,
        "status": "Success",
        "action": "denied",
        "message": f"Denied and returned to staging: {Path(target_path).name}",
        "target_path": target_path,
    }


def render_review_queue_view(
    ready_folder_names: list[str],
    profiles: list[dict[str, Any]],
    dropbox_cfg: dict[str, Any],
) -> None:
    st.subheader("Review queue")

    queue_items = build_ready_queue_items(ready_folder_names, profiles, dropbox_cfg)
    summary_rows = [
        {
            "folder_name": item["folder_name"],
            "template": item["template"],
            "title": item["title"],
            "variants_summary": item["variants_summary"],
            "load_status": item["load_status"],
        }
        for item in queue_items
    ]

    if not summary_rows:
        st.info("No listings are currently waiting for review.")
        return

    ready_lookup = {item["folder_name"]: item for item in queue_items}
    review_folder_options = [item["folder_name"] for item in queue_items if item["listing_memory"]]
    actionable_folder_options = [
        item["folder_name"]
        for item in queue_items
        if item["profile"] and item["listing_memory"] and not item["load_error"]
    ]
    load_issue_count = sum(1 for item in queue_items if item["load_error"])
    missing_memory_count = sum(1 for item in queue_items if not item["listing_memory"])
    bulk_results = list(st.session_state.get("review_queue_bulk_approve_results", []))

    metric_col1, metric_col2, metric_col3 = st.columns(3)
    metric_col1.metric("Waiting review", len(queue_items))
    metric_col2.metric("Reviewable", len(actionable_folder_options))
    metric_col3.metric("Load issues", load_issue_count + missing_memory_count)

    with st.expander("Review queue summary", expanded=False):
        st.dataframe(summary_rows, width="stretch", hide_index=True)

    review_queue_sections = ["Review / decide", "Bulk approve", "Results"]
    active_review_queue_section = st.segmented_control(
        "Review queue section",
        review_queue_sections,
        key="review_queue_section",
        selection_mode="single",
        width="stretch",
        label_visibility="collapsed",
    )
    if active_review_queue_section not in review_queue_sections:
        active_review_queue_section = review_queue_sections[0]

    if active_review_queue_section == "Bulk approve":
        st.caption("Use this only for listings that do not need detailed review edits.")
        existing_bulk_selection = list(st.session_state.get("review_queue_bulk_approve_folders", []))
        valid_bulk_selection = [
            folder_name for folder_name in existing_bulk_selection
            if folder_name in actionable_folder_options
        ]
        if valid_bulk_selection != existing_bulk_selection:
            st.session_state["review_queue_bulk_approve_folders"] = valid_bulk_selection

        with st.container(border=True):
            with st.form("review_queue_bulk_approve_form"):
                selected_bulk_folders = st.multiselect(
                    "Ready listings to approve",
                    actionable_folder_options,
                    key="review_queue_bulk_approve_folders",
                    placeholder="Choose listings that are ready without further edits",
                )
                bulk_reviewed_by = st.selectbox(
                    "Reviewed by",
                    WORKFLOW_ASSIGNEES,
                    key="review_queue_bulk_reviewed_by",
                )
                bulk_approve_clicked = st.form_submit_button(
                    "Approve selected for generation",
                    width="stretch",
                    disabled=not bool(actionable_folder_options),
                )

            if bulk_approve_clicked:
                if not selected_bulk_folders:
                    st.warning("Select at least one ready listing to approve.")
                elif not bulk_reviewed_by:
                    st.warning("Select who reviewed these listings before approving them.")
                else:
                    st.session_state["pending_perf_action_label"] = "bulk approve ready listings"
                    results: list[dict[str, Any]] = []
                    for folder_name in selected_bulk_folders:
                        review_item = ready_lookup.get(folder_name)
                        try:
                            result = process_ready_review_decision(
                                review_item or {},
                                dropbox_cfg,
                                bulk_reviewed_by,
                                approve=True,
                            )
                        except Exception as exc:
                            result = {
                                "folder_name": folder_name,
                                "status": "Failed",
                                "action": "approved",
                                "message": str(exc),
                            }
                        results.append(result)

                    st.session_state["review_queue_bulk_approve_results"] = results
                    success_results = [row for row in results if row.get("status") == "Success"]
                    failed_results = [row for row in results if row.get("status") == "Failed"]
                    if len(success_results) == 1:
                        st.session_state["last_approved_folder_path"] = success_results[0].get("target_path", "")
                    st.session_state.pop("review_queue_bulk_approve_folders", None)
                    clear_runtime_caches()
                    set_workflow_flash(
                        "success" if not failed_results else "warning",
                        f"Approved {len(success_results)} of {len(selected_bulk_folders)} selected listing(s).",
                        f"{len(failed_results)} failed." if failed_results else "",
                    )
                    st.rerun()

        return

    if active_review_queue_section == "Results":
        with st.container(border=True):
            if bulk_results:
                st.dataframe(bulk_results, width="stretch", hide_index=True)
            else:
                st.caption("No recent bulk approval results yet.")
        return

    if active_review_queue_section != "Review / decide":
        return

    with st.container(border=True):
        st.caption("Open a listing only when you need detailed content, price, image, or quality review.")
        if not review_folder_options:
            st.caption("No reviewable ready listings found.")
            return

        current_review_folder = st.session_state.get("ready_queue_review_folder", review_folder_options[0])
        if current_review_folder not in review_folder_options:
            current_review_folder = review_folder_options[0]
            st.session_state["ready_queue_review_folder"] = current_review_folder

        selected_review_folder = st.selectbox(
            "Listing folder",
            review_folder_options,
            key="ready_queue_review_folder",
        )
        review_item = ready_lookup.get(selected_review_folder)
        if review_item:
            review_panel_key_suffix = selected_review_folder.replace("/", "_").replace("\\", "_").replace(" ", "_")
            review_panel_open_key = f"review_queue_panel_open_{review_panel_key_suffix}"

            panel_col1, panel_col2 = st.columns([1, 3])
            with panel_col1:
                if st.button("Open review panel", key=f"{review_panel_open_key}_open_btn", width="stretch"):
                    st.session_state["active_perf_action_label"] = "open ready review panel"
                    clear_review_editor_state(f"review_queue_{review_panel_key_suffix}")
                    st.session_state[review_panel_open_key] = True
            with panel_col2:
                if st.session_state.get(review_panel_open_key, False):
                    if st.button("Hide review panel", key=f"{review_panel_open_key}_hide_btn"):
                        st.session_state["active_perf_action_label"] = "hide ready review panel"
                        st.session_state[review_panel_open_key] = False
                else:
                    st.caption("Review panel is closed.")

            if st.session_state.get(review_panel_open_key, False):
                with st.expander("Review panel", expanded=True):
                    render_ready_review_panel(
                        review_item,
                        dropbox_cfg,
                        key_prefix="review_queue",
                        source_folder_path=build_ready_folder_path(dropbox_cfg, review_item["folder_name"]),
                    )

            review_edit_prefix = f"review_queue_{review_panel_key_suffix}"
            active_review_edit_prefix = st.session_state.get(
                f"{review_edit_prefix}_active_editor_prefix",
                review_edit_prefix,
            )
            if review_item.get("profile") and review_item.get("listing_memory"):
                initialize_review_edit_state(
                    review_edit_prefix,
                    review_item["listing_memory"],
                    review_item["profile"],
                )

            default_reviewer = review_item.get("listing_memory", {}).get("reviewed_by", "")
            review_reviewer_key = st.session_state.get("review_queue_review_folder_reviewer_key", "")
            reviewer_context_key = f"{selected_review_folder}|{default_reviewer}"
            if review_reviewer_key != reviewer_context_key:
                st.session_state["review_queue_reviewed_by"] = default_reviewer if default_reviewer in WORKFLOW_ASSIGNEES else ""
                st.session_state["review_queue_review_folder_reviewer_key"] = reviewer_context_key

            with st.form("review_queue_decision_form"):
                st.selectbox(
                    "Reviewed by",
                    WORKFLOW_ASSIGNEES,
                    key="review_queue_reviewed_by",
                )
                action_col1, action_col2 = st.columns(2)
                with action_col1:
                    approve_clicked = st.form_submit_button(
                        "Approve for generation",
                        width="stretch",
                    )
                with action_col2:
                    deny_clicked = st.form_submit_button(
                        "Deny and return to staging",
                        width="stretch",
                    )

            if approve_clicked or deny_clicked:
                st.session_state["pending_perf_action_label"] = (
                    "approve ready listing" if approve_clicked else "deny ready listing"
                )
                reviewed_by = st.session_state.get("review_queue_reviewed_by", "")
                if not reviewed_by:
                    st.warning("Select who reviewed this listing before approving or denying it.")
                    return
                if not review_item.get("profile") or not review_item.get("listing_memory"):
                    st.error("This ready listing could not be loaded for review.")
                    return

                try:
                    result = process_ready_review_decision(
                        review_item,
                        dropbox_cfg,
                        reviewed_by,
                        approve=approve_clicked,
                        review_edit_prefix=active_review_edit_prefix,
                    )
                    if approve_clicked:
                        st.session_state["last_approved_folder_path"] = result.get("target_path", "")
                        clear_runtime_caches()
                        set_workflow_flash(
                            "success",
                            result.get("message", "Approved successfully."),
                        )
                    else:
                        target_path = result.get("target_path", "")
                        if target_path:
                            st.session_state["pending_staged_folder_selection_on_rerun"] = Path(target_path).name
                        st.session_state["auto_switch_to_staged"] = True
                        clear_runtime_caches()
                        set_workflow_flash(
                            "warning",
                            result.get("message", "Denied and returned to staging."),
                        )
                    st.rerun()
                except Exception as exc:
                    if approve_clicked:
                        st.error(f"Could not approve the listing: {exc}")
                    else:
                        st.error(f"Could not deny the listing: {exc}")

    return


def generate_approved_output_batch(
    *,
    target_folders: list[str],
    approved_lookup: dict[str, dict[str, Any]],
    profiles: list[dict[str, Any]],
    dropbox_cfg: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    results: list[dict[str, Any]] = []
    for folder_name in target_folders:
        item = approved_lookup.get(folder_name)
        if not item:
            results.append({
                "folder_name": folder_name,
                "status": "Failed",
                "message": "Approved folder could not be loaded.",
            })
            continue

        try:
            results.append(generate_approved_listing(
                profile=item["profile"],
                listing_memory=item["listing_memory"],
                approved_folder_name=folder_name,
                dropbox_cfg=dropbox_cfg,
            ))
        except Exception as exc:
            results.append({
                "folder_name": folder_name,
                "status": "Failed",
                "message": str(exc),
            })

    if len(target_folders) >= 2:
        selected_items = [
            approved_lookup.get(folder_name)
            for folder_name in target_folders
            if approved_lookup.get(folder_name)
        ]
        combined_groups: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
        skipped_combined_items: list[str] = []
        for item in selected_items:
            item_profile = item.get("profile") if item else None
            folder_name = str(item.get("folder_name", "") if item else "")
            if not item_profile or not item.get("listing_memory") or item.get("load_error"):
                skipped_combined_items.append(
                    f"{folder_name or 'unknown'} (profile or saved listing memory unavailable)"
                )
                continue
            try:
                group_key = get_combined_workbook_group_identity(item_profile)
            except Exception as exc:
                skipped_combined_items.append(f"{folder_name or 'unknown'} ({exc})")
                continue
            combined_groups.setdefault(group_key, []).append(item)

        for group_items in combined_groups.values():
            if len(group_items) < 2:
                continue
            group_label = build_combined_workbook_group_label(group_items)
            try:
                results.append(generate_approved_listings_combined(group_items, dropbox_cfg))
            except Exception as exc:
                results.append({
                    "folder_name": f"Combined workbook - {group_label}",
                    "status": "Failed",
                    "message": f"Separate workbooks were generated. Grouped workbook was skipped: {exc}",
                })

        if skipped_combined_items:
            results.append({
                "folder_name": "Combined workbook",
                "status": "Skipped",
                "message": (
                    "Grouped workbook skipped for incompatible listing(s): "
                    + ", ".join(skipped_combined_items[:10])
                ),
            })

    return move_successful_generation_results_to_finished(
        results=results,
        profiles=profiles,
        dropbox_cfg=dropbox_cfg,
    )


def render_approved_queue_view(
    approved_folder_names: list[str],
    profiles: list[dict[str, Any]],
    dropbox_cfg: dict[str, Any],
) -> None:
    st.subheader("Approved output")

    queue_items = build_approved_queue_items(approved_folder_names, profiles, dropbox_cfg)
    stored_results = list(st.session_state.get("approved_queue_generation_results", []))
    approved_lookup = {item["folder_name"]: item for item in queue_items}
    review_folder_options = [item["folder_name"] for item in queue_items if item["listing_memory"]]
    actionable_folder_options = [
        item["folder_name"]
        for item in queue_items
        if item["profile"] and item["listing_memory"] and not item["load_error"]
    ]
    summary_rows = [
        {
            "folder_name": item["folder_name"],
            "template": item["template"],
            "title": item["title"],
            "variants_summary": item["variants_summary"],
            "load_status": item["load_status"],
        }
        for item in queue_items
    ]

    metric_col1, metric_col2, metric_col3 = st.columns(3)
    metric_col1.metric("Approved", len(queue_items))
    metric_col2.metric("Ready", len(actionable_folder_options))
    metric_col3.metric(
        "Downloads",
        sum(1 for result in stored_results if result.get("status") == "Success" and result.get("output_path")),
    )

    if summary_rows:
        with st.expander("Approved folder summary", expanded=False):
            st.dataframe(summary_rows, width="stretch", hide_index=True)
    else:
        st.info("No approved folders found.")
        render_generation_results(stored_results, "approved_download")
        return

    approved_output_sections = ["Generate", "Review / edit", "Return"]
    active_approved_output_section = st.segmented_control(
        "Approved output section",
        approved_output_sections,
        key="approved_output_section",
        selection_mode="single",
        width="stretch",
        label_visibility="collapsed",
    )
    if active_approved_output_section not in approved_output_sections:
        active_approved_output_section = approved_output_sections[0]

    generate_selected = False
    generate_all = False
    selected_approved_folders: list[str] = []

    if active_approved_output_section == "Generate":
        with st.container(border=True):
            st.caption("Generate Excel files. Successful single listings are moved to Finished before downloads appear.")
            with st.form("approved_output_generation_form"):
                selected_approved_folders = st.multiselect(
                    "Approved folders",
                    actionable_folder_options,
                    key="approved_queue_selected_folders",
                    placeholder="Choose approved folders to generate",
                )

                col1, col2 = st.columns(2)
                with col1:
                    generate_selected = st.form_submit_button("Generate selected", width="stretch")
                with col2:
                    generate_all = st.form_submit_button("Generate all approved", width="stretch")

        if stored_results:
            stored_results, move_results = move_successful_generation_results_to_finished(
                results=stored_results,
                profiles=profiles,
                dropbox_cfg=dropbox_cfg,
            )
            if move_results:
                st.session_state["approved_queue_generation_results"] = stored_results
                st.session_state["approved_queue_move_generated_results"] = move_results
                clear_runtime_caches()
            render_generation_results(stored_results, "approved_download")

    if active_approved_output_section == "Review / edit":
        with st.container(border=True):
            st.caption("Open detailed approved listing review only when you need to inspect or edit it.")
            if review_folder_options:
                current_review_folder = st.session_state.get("approved_queue_review_folder", review_folder_options[0])
                if current_review_folder not in review_folder_options:
                    current_review_folder = review_folder_options[0]
                    st.session_state["approved_queue_review_folder"] = current_review_folder

                selected_review_folder = st.selectbox(
                    "Approved listing",
                    review_folder_options,
                    key="approved_queue_review_folder",
                )
                review_item = approved_lookup.get(selected_review_folder)
                if review_item:
                    approved_panel_key_suffix = selected_review_folder.replace("/", "_").replace("\\", "_").replace(" ", "_")
                    approved_panel_open_key = f"approved_output_panel_open_{approved_panel_key_suffix}"

                    panel_col1, panel_col2 = st.columns([1, 3])
                    with panel_col1:
                        if st.button("Open review panel", key=f"{approved_panel_open_key}_open_btn", width="stretch"):
                            st.session_state["active_perf_action_label"] = "open approved review panel"
                            clear_review_editor_state(f"approved_output_{approved_panel_key_suffix}")
                            st.session_state[approved_panel_open_key] = True
                    with panel_col2:
                        if st.session_state.get(approved_panel_open_key, False):
                            if st.button("Hide review panel", key=f"{approved_panel_open_key}_hide_btn"):
                                st.session_state["active_perf_action_label"] = "hide approved review panel"
                                st.session_state[approved_panel_open_key] = False
                        else:
                            st.caption("Review panel is closed.")

                    if st.session_state.get(approved_panel_open_key, False):
                        with st.expander("Review panel", expanded=True):
                            render_ready_review_panel(
                                review_item,
                                dropbox_cfg,
                                key_prefix="approved_output",
                                source_folder_path=build_approved_folder_path(dropbox_cfg, review_item["folder_name"]),
                            )
            else:
                st.caption("No approved listings are available for review.")
        return

    if active_approved_output_section == "Return":
        with st.container(border=True):
            st.caption("Move approved listings back when they need content, price, image, or setup changes.")
            existing_return_selection = list(st.session_state.get("approved_queue_return_folders", []))
            valid_return_selection = [
                folder_name for folder_name in existing_return_selection
                if folder_name in actionable_folder_options
            ]
            if valid_return_selection != existing_return_selection:
                st.session_state["approved_queue_return_folders"] = valid_return_selection

            with st.form("approved_queue_return_form"):
                selected_return_folders = st.multiselect(
                    "Approved listings to move",
                    actionable_folder_options,
                    key="approved_queue_return_folders",
                )
                return_target_label = st.radio(
                    "Move selected listings to",
                    ["Review queue", "Stage folder"],
                    horizontal=True,
                    key="approved_queue_return_target",
                )
                return_actor = st.selectbox(
                    "Returned by",
                    WORKFLOW_ASSIGNEES,
                    key="approved_queue_return_actor",
                )
                return_reason = st.text_input(
                    "Reason",
                    key="approved_queue_return_reason",
                    placeholder="Needs price/content/image changes",
                )
                return_clicked = st.form_submit_button(
                    "Move selected approved listing(s)",
                    width="stretch",
                    disabled=not bool(actionable_folder_options),
                )

            if return_clicked:
                if not selected_return_folders:
                    st.warning("Select at least one approved listing to move.")
                elif not return_actor:
                    st.warning("Select who is returning these listings.")
                else:
                    target_state = "ready" if return_target_label == "Review queue" else "stage"
                    action_label = (
                        "return approved to review"
                        if target_state == "ready"
                        else "return approved to stage"
                    )
                    st.session_state["active_perf_action_label"] = action_label
                    st.session_state["pending_perf_action_label"] = action_label

                    return_results = [
                        return_approved_listing(
                            dropbox_cfg=dropbox_cfg,
                            profiles=profiles,
                            fallback_profile=profiles[0] if profiles else {},
                            approved_folder_name=folder_name,
                            target_state=target_state,
                            actor=return_actor,
                            reason=return_reason.strip(),
                        )
                        for folder_name in selected_return_folders
                    ]
                    success_results = [
                        row for row in return_results
                        if row.get("status") == "Success"
                    ]
                    failed_results = [
                        row for row in return_results
                        if row.get("status") == "Failed"
                    ]

                    st.session_state["approved_queue_return_results"] = return_results
                    st.session_state.pop("approved_queue_return_folders", None)
                    clear_runtime_caches()
                    set_workflow_flash(
                        "success" if not failed_results else "warning",
                        f"Moved {len(success_results)} of {len(selected_return_folders)} approved listing(s).",
                        (
                            "They are back in the review queue."
                            if target_state == "ready" and not failed_results
                            else "They are back in staging."
                            if target_state == "stage" and not failed_results
                            else f"{len(failed_results)} folder(s) failed."
                        ),
                    )
                    st.rerun()

            return_results = list(st.session_state.get("approved_queue_return_results", []))
            if return_results:
                st.dataframe(return_results, width="stretch", hide_index=True)
        return

    if active_approved_output_section != "Generate":
        return

    if generate_selected:
        st.session_state["pending_perf_action_label"] = "generate selected approved"
    elif generate_all:
        st.session_state["pending_perf_action_label"] = "generate all approved"

    target_folders = selected_approved_folders if generate_selected else [
        item["folder_name"] for item in queue_items if item["profile"] and item["listing_memory"] and not item["load_error"]
    ] if generate_all else []
    if generate_selected and not target_folders:
        st.warning("Select at least one approved folder to generate.")
        return
    if not target_folders:
        return

    approved_generation_started_at = time.perf_counter()
    approved_generation_target_count = len(target_folders)

    results, move_results = generate_approved_output_batch(
        target_folders=target_folders,
        approved_lookup=approved_lookup,
        profiles=profiles,
        dropbox_cfg=dropbox_cfg,
    )
    if move_results:
        st.session_state["approved_queue_move_generated_results"] = move_results

    approved_generation_elapsed_ms = round(
        (time.perf_counter() - approved_generation_started_at) * 1000,
        1,
    )
    approved_generation_failures = sum(
        1 for result in results if result.get("status") == "Failed"
    )

    generation_step_rows: list[dict[str, Any]] = []
    for result in results:
        timings = result.get("timings", {}) if isinstance(result, dict) else {}
        for step_name, seconds in dict(timings).items():
            try:
                generation_step_rows.append({
                    "folder_name": result.get("folder_name", ""),
                    "step": step_name,
                    "ms": round(float(seconds) * 1000, 1),
                })
            except Exception:
                pass

    if generation_step_rows:
        slowest_generation_step = max(generation_step_rows, key=lambda row: row["ms"])
        slowest_generation_event = (
            f"Approved generation: {slowest_generation_step['step']} "
            f"({slowest_generation_step['folder_name']})"
        )
        slowest_generation_ms = slowest_generation_step["ms"]
    else:
        slowest_generation_event = f"Approved generation: {approved_generation_target_count} folder(s)"
        slowest_generation_ms = approved_generation_elapsed_ms

    st.session_state["approved_generation_step_rows"] = generation_step_rows

    perf_history = st.session_state.setdefault("perf_history", [])
    perf_history.append({
        "run": len(perf_history) + 1,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "action": (
            "generate selected approved actual"
            if generate_selected
            else "generate all approved actual"
        ),
        "full_rerun_ms": approved_generation_elapsed_ms,
        "recorded_load_ms": approved_generation_elapsed_ms,
        "estimated_ui_build_ms": 0,
        "slowest_event": slowest_generation_event,
        "slowest_ms": slowest_generation_ms,
        "event_count": approved_generation_target_count,
    })

    if len(perf_history) > 300:
        st.session_state["perf_history"] = perf_history[-300:]

    # Prevent the next display rerun from inheriting the generation label.
    st.session_state.pop("pending_perf_action_label", None)
    st.session_state.pop("active_perf_action_label", None)

    st.session_state["approved_queue_generation_results"] = results
    st.session_state.pop("approved_queue_selected_folders", None)
    clear_runtime_caches()
    st.rerun()

def debug_size_headers(header_map: dict[str, int]) -> None:
    if not st.session_state.get("show_header_debug", False):
        return

    patterns = [
        "size_system",
        "size_class",
        "size_value",
        "apparel_size",
        "body_type",
        "height_type",
    ]

    st.write("Detailed size/header matches")
    for pattern in patterns:
        matches = [key for key in header_map.keys() if pattern.lower() in key.lower()]
        st.write({pattern: matches})


def render_widget_visibility_patch() -> None:
    st.markdown(
        """
        <style>
        input,
        textarea,
        [contenteditable="true"],
        [role="textbox"] {
            color: #f8fafc !important;
            -webkit-text-fill-color: #f8fafc !important;
            caret-color: #f8fafc !important;
            opacity: 1 !important;
            visibility: visible !important;
            text-indent: 0 !important;
            letter-spacing: 0 !important;
            text-shadow: none !important;
            filter: none !important;
        }

        input *,
        textarea *,
        [role="textbox"] * {
            color: inherit !important;
            -webkit-text-fill-color: inherit !important;
            opacity: 1 !important;
            visibility: visible !important;
        }

        [data-testid="stTextInput"] input,
        [data-testid="stTextArea"] textarea,
        [data-testid="stNumberInput"] input,
        [data-testid="stSelectbox"] input,
        [data-testid="stMultiSelect"] input,
        [data-testid="stTextInput"] div[data-baseweb="input"] input,
        [data-testid="stTextArea"] div[data-baseweb="textarea"] textarea,
        [data-testid="stNumberInput"] div[data-baseweb="input"] input,
        [data-testid="stSelectbox"] div[data-baseweb="select"] input,
        [data-testid="stMultiSelect"] div[data-baseweb="select"] input,
        div[data-baseweb="base-input"] input,
        div[data-baseweb="input"] input,
        div[data-baseweb="textarea"] textarea,
        div[data-baseweb="select"] input,
        .stTextInput input,
        .stTextArea textarea,
        .stNumberInput input {
            color: #f8fafc !important;
            -webkit-text-fill-color: #f8fafc !important;
            caret-color: #f8fafc !important;
            opacity: 1 !important;
            text-shadow: none !important;
        }

        [data-testid="stTextInput"] input::placeholder,
        [data-testid="stTextArea"] textarea::placeholder,
        [data-testid="stNumberInput"] input::placeholder,
        [data-testid="stSelectbox"] input::placeholder,
        [data-testid="stMultiSelect"] input::placeholder,
        div[data-baseweb="input"] input::placeholder,
        div[data-baseweb="textarea"] textarea::placeholder,
        .stTextInput input::placeholder,
        .stTextArea textarea::placeholder {
            color: #94a3b8 !important;
            -webkit-text-fill-color: #94a3b8 !important;
            opacity: 1 !important;
        }

        [data-testid="stTextInput"] input:disabled,
        [data-testid="stTextArea"] textarea:disabled,
        [data-testid="stNumberInput"] input:disabled,
        div[data-baseweb="input"] input:disabled,
        div[data-baseweb="textarea"] textarea:disabled,
        .stTextInput input:disabled,
        .stTextArea textarea:disabled {
            color: #cbd5e1 !important;
            -webkit-text-fill-color: #cbd5e1 !important;
        }

        [data-testid="stTextInput"] input[value]:not([value=""]),
        [data-testid="stTextArea"] textarea:not(:placeholder-shown),
        [data-testid="stNumberInput"] input[value]:not([value=""]) {
            color: #f8fafc !important;
            -webkit-text-fill-color: #f8fafc !important;
        }

        input:-webkit-autofill,
        input:-webkit-autofill:hover,
        input:-webkit-autofill:focus,
        textarea:-webkit-autofill,
        textarea:-webkit-autofill:hover,
        textarea:-webkit-autofill:focus {
            -webkit-text-fill-color: #f8fafc !important;
            transition: background-color 9999s ease-out 0s !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def main() -> None:
    st.set_page_config(page_title="Amazon Listing Generator", layout="wide")
    render_widget_visibility_patch()
    st.title("Amazon Listing Generator")
    st.caption("Template-based Amazon flat file generator.")
    render_workflow_flash()
    reset_load_events()
    consume_pending_perf_action_label()
    capture_rerun_cause()

    started_at = time.perf_counter()
    profiles = list_template_profiles()
    record_load_event("Template profiles", started_at)

    started_at = time.perf_counter()
    dropbox_cfg = load_dropbox_templates_config()
    record_load_event("Dropbox template config", started_at)

    stage_root = dropbox_cfg.get("stage_root", "")
    ready_root = dropbox_cfg.get("ready_root", "")
    approved_root = dropbox_cfg.get("approved_root", "")
    finished_root = dropbox_cfg.get("finished_root", "")

    if not stage_root or not ready_root or not approved_root or not finished_root:
        st.error("stage_root, ready_root, approved_root, and finished_root must be set in config/dropbox_templates.json")
        st.stop()

    workflow_tab_labels = [
        "Product setup",
        "Listing content",
        "Review queue",
        "Approved output",
    ]
    pending_workflow_tab = st.session_state.pop("pending_workflow_active_tab", "")
    if pending_workflow_tab in workflow_tab_labels:
        st.session_state["workflow_active_tab"] = pending_workflow_tab
    if st.session_state.get("workflow_active_tab") not in workflow_tab_labels:
        st.session_state["workflow_active_tab"] = workflow_tab_labels[0]
    active_workflow_tab = st.segmented_control(
        "Workflow section",
        workflow_tab_labels,
        key="workflow_active_tab",
        selection_mode="single",
        width="stretch",
        label_visibility="collapsed",
    )

    staged_folder_names: list[str] = []
    ready_folder_names: list[str] = []
    approved_folder_names: list[str] = []
    finished_folder_names: list[str] = []

    try:
        if (
            active_workflow_tab == "Product setup"
            and st.session_state.get("stage_folder_list_loaded", False)
        ):
            staged_folder_names = get_cached_folder_names("stage", stage_root, "_stage folders")
        if (
            active_workflow_tab == "Product setup"
            and st.session_state.get("active_folder_source_mode") == "Restage finished folder"
        ):
            finished_folder_names = get_cached_folder_names("finished", finished_root, "finished folders")
        elif active_workflow_tab == "Review queue" and st.session_state.get("review_queue_tab_loaded", False):
            ready_folder_names = get_cached_folder_names("ready", ready_root, "ready folders")
        elif active_workflow_tab == "Approved output":
            finished_folder_names = get_cached_folder_names("finished", finished_root, "finished folders")
            if st.session_state.get("approved_output_tab_loaded", False):
                approved_folder_names = get_cached_folder_names("approved", approved_root, "approved folders")
    except Exception as exc:
        st.error(f"Could not read Dropbox folders: {exc}")
        st.stop()

    dropbox_folder_load_errors = dict(st.session_state.get("dropbox_folder_load_errors", {}))
    if dropbox_folder_load_errors:
        st.warning(
            "Some Dropbox folder lists could not be refreshed. "
            "The app is using cached lists where possible; empty sections can be retried."
        )
        with st.expander("Dropbox folder load errors", expanded=False):
            for message in dropbox_folder_load_errors.values():
                st.write(message)
            if st.button("Retry Dropbox folder lists", key="retry_dropbox_folder_lists_btn"):
                st.session_state["pending_perf_action_label"] = "retry Dropbox folder lists"
                refresh_cached_folder_names("stage", "ready", "approved", "finished")
                st.rerun()

    if not profiles:
        st.error("No template profiles found. Create family folders under templates/ with schema.json, a shared workbook, and garment subfolders containing config.json.")
        st.stop()

    families = sorted({profile.get("_family_slug", "") for profile in profiles if profile.get("_family_slug")})
    detection_message = ""
    detection_level = ""

    if st.session_state.pop("auto_switch_to_staged", False):
        st.session_state["folder_source_mode"] = "Use staged folder"
        st.session_state["active_folder_source_mode"] = "Use staged folder"

    pending_staged_folder_selection = st.session_state.pop("pending_staged_folder_selection_on_rerun", None)
    if pending_staged_folder_selection:
        st.session_state["staged_folder_select"] = pending_staged_folder_selection
        st.session_state["active_staged_folder_select"] = pending_staged_folder_selection
        st.session_state.pop("last_detected_template_folder", None)
        st.session_state.pop("applied_listing_memory_key_v2", None)
        st.session_state.pop("applied_listing_memory_widget_key_v2", None)
        st.session_state.pop("initialized_listing_context_key", None)
        st.session_state.pop("last_loaded_listing_memory_signature", None)

    if st.session_state.pop("clear_staged_folder_selection_on_rerun", False):
        st.session_state["staged_folder_select"] = None
        st.session_state["active_staged_folder_select"] = ""
        st.session_state.pop("last_detected_template_folder", None)
        st.session_state.pop("applied_listing_memory_key_v2", None)
        st.session_state.pop("applied_listing_memory_widget_key_v2", None)
        st.session_state.pop("initialized_listing_context_key", None)
        st.session_state.pop("last_loaded_listing_memory_signature", None)

    # Product setup widgets are not rendered on the other workflow sections.
    # Keep durable active selections separate so hidden/cleaned widget state
    # cannot reset a loaded staged folder or template while editing content.
    if active_workflow_tab == "Product setup":
        if "folder_source_mode" in st.session_state:
            st.session_state["active_folder_source_mode"] = st.session_state.get("folder_source_mode") or "Use staged folder"
        if "staged_folder_select" in st.session_state:
            st.session_state["active_staged_folder_select"] = st.session_state.get("staged_folder_select") or ""
        if "finished_folder_select" in st.session_state:
            st.session_state["active_finished_folder_select"] = st.session_state.get("finished_folder_select") or ""
        if "template_family_select" in st.session_state:
            st.session_state["active_template_family_select"] = st.session_state.get("template_family_select") or ""
        if "listing_template_select" in st.session_state:
            st.session_state["active_listing_template_select"] = st.session_state.get("listing_template_select") or ""
    else:
        if st.session_state.get("active_folder_source_mode"):
            st.session_state["folder_source_mode"] = st.session_state.get("active_folder_source_mode")
        if st.session_state.get("active_staged_folder_select"):
            st.session_state["staged_folder_select"] = st.session_state.get("active_staged_folder_select")
        if st.session_state.get("active_finished_folder_select"):
            st.session_state["finished_folder_select"] = st.session_state.get("active_finished_folder_select")
        if st.session_state.get("active_template_family_select"):
            st.session_state["template_family_select"] = st.session_state.get("active_template_family_select")
        if st.session_state.get("active_listing_template_select"):
            st.session_state["listing_template_select"] = st.session_state.get("active_listing_template_select")

    folder_source = st.session_state.get("active_folder_source_mode", "Use staged folder")
    initial_staged_folder_name = st.session_state.get("active_staged_folder_select", "") if folder_source == "Use staged folder" else ""
    listing_memory: dict[str, Any] = {}
    authoritative_profile: dict[str, Any] | None = None
    grouped_state_load_error = ""
    listing_memory_location = ""

    if initial_staged_folder_name:
        memory_context = resolve_active_staged_listing_memory(
            dropbox_cfg=dropbox_cfg,
            staged_folder_name=initial_staged_folder_name,
            load_stage_memory=load_listing_memory_from_dropbox,
            destination_exists=path_exists_strict,
            load_fresh_memory=load_listing_memory_from_dropbox_fresh,
        )
        listing_memory = dict(memory_context.get("memory", {}) or {})
        listing_memory_location = str(memory_context.get("location", "") or "")
        grouped_state_load_error = str(memory_context.get("error", "") or "")
        known_grouped_sources = set(st.session_state.get("known_grouped_source_folders", []))
        if is_grouped_christmas_memory(listing_memory):
            known_grouped_sources.add(initial_staged_folder_name)
            st.session_state["known_grouped_source_folders"] = sorted(known_grouped_sources)
        elif (
            initial_staged_folder_name in known_grouped_sources
            or "listing_group" in listing_memory
            or isinstance(listing_memory.get("group_submission"), dict)
        ):
            grouped_state_load_error = (
                "Grouped Christmas state could not be loaded. This staged task cannot "
                "be edited or submitted until its saved grouped state is recovered."
            )

        authoritative_profile = find_profile_for_listing_memory(profiles, listing_memory) if listing_memory else None
        if authoritative_profile:
            st.session_state["template_family_select"] = authoritative_profile.get("_family_slug", "")
            st.session_state["listing_template_select"] = authoritative_profile.get("label", authoritative_profile.get("_slug", ""))
            st.session_state["active_template_family_select"] = authoritative_profile.get("_family_slug", "")
            st.session_state["active_listing_template_select"] = authoritative_profile.get("label", authoritative_profile.get("_slug", ""))

    current_folder_source_mode = st.session_state.get("active_folder_source_mode", "Use staged folder")
    current_detect_folder = st.session_state.get("active_staged_folder_select", "") if current_folder_source_mode == "Use staged folder" else ""

    if current_detect_folder and not authoritative_profile:
        last_detect_folder = st.session_state.get("last_detected_template_folder", "")
        matches: list[dict[str, Any]] = []
        should_detect_folder_template = last_detect_folder != current_detect_folder
        if not should_detect_folder_template:
            matches = find_template_matches_for_staged_folder(current_detect_folder, profiles)
            if len(matches) == 1:
                matched = matches[0]
                current_template_label = st.session_state.get("active_listing_template_select") or st.session_state.get("listing_template_select", "")
                current_template_family = st.session_state.get("active_template_family_select") or st.session_state.get("template_family_select", "")
                should_detect_folder_template = (
                    current_template_family != matched.get("_family_slug", "")
                    or current_template_label != matched.get("label", matched.get("_slug", ""))
                )

        if should_detect_folder_template:
            if not matches:
                matches = find_template_matches_for_staged_folder(current_detect_folder, profiles)
            st.session_state["last_detected_template_folder"] = current_detect_folder

            if len(matches) == 1:
                matched = matches[0]
                st.session_state["template_family_select"] = matched.get("_family_slug", "")
                st.session_state["listing_template_select"] = matched.get("label", matched.get("_slug", ""))
                st.session_state["active_template_family_select"] = matched.get("_family_slug", "")
                st.session_state["active_listing_template_select"] = matched.get("label", matched.get("_slug", ""))
                st.session_state["template_detection_message"] = (
                    f"Auto-detected template `{matched.get('label', matched.get('_slug', ''))}` from staged folder `{current_detect_folder}`."
                )
                st.session_state["template_detection_level"] = "info"
            elif len(matches) > 1:
                matched_families = {match.get("_family_slug", "") for match in matches}
                if len(matched_families) == 1:
                    matched_family = next(iter(matched_families))
                    st.session_state["template_family_select"] = matched_family
                    st.session_state["active_template_family_select"] = matched_family
                    match_labels = ", ".join(match.get("label", match.get("_slug", "")) for match in matches)
                    st.session_state["template_detection_message"] = (
                        f"Detected family `{matched_family}` from staged folder `{current_detect_folder}`. "
                        f"Please confirm which template to use: {match_labels}."
                    )
                    st.session_state["template_detection_level"] = "warning"
                else:
                    st.session_state["template_detection_message"] = (
                        f"Found multiple possible template matches for `{current_detect_folder}`. Please choose manually."
                    )
                    st.session_state["template_detection_level"] = "warning"
            else:
                st.session_state.pop("template_detection_message", None)
                st.session_state.pop("template_detection_level", None)

    detection_message = st.session_state.get("template_detection_message", "")
    detection_level = st.session_state.get("template_detection_level", "")

    if authoritative_profile:
        selected_family = authoritative_profile.get("_family_slug", "")
        selected_label = authoritative_profile.get("label", authoritative_profile.get("_slug", ""))
    else:
        selected_family = st.session_state.get("active_template_family_select") or st.session_state.get("template_family_select", families[0] if families else "")
    if families and selected_family not in families:
        selected_family = families[0]
        st.session_state["template_family_select"] = selected_family
        st.session_state["active_template_family_select"] = selected_family

    family_profiles = [
        profile for profile in profiles
        if profile.get("_family_slug") == selected_family
    ]

    family_labels = [profile.get("label", profile["_slug"]) for profile in family_profiles]

    if not authoritative_profile:
        selected_label = st.session_state.get("active_listing_template_select") or st.session_state.get("listing_template_select", family_labels[0] if family_labels else "")
    if family_labels and selected_label not in family_labels:
        selected_label = family_labels[0]
        st.session_state["listing_template_select"] = selected_label
        st.session_state["active_listing_template_select"] = selected_label

    profile = family_profiles[family_labels.index(selected_label)]
    active_staged_folder_name = initial_staged_folder_name
    active_listing_memory = listing_memory
    active_profile = profile
    active_profile_slug = active_profile.get("_slug", "")
    active_family_slug = active_profile.get("_family_slug", "")
    active_template_label = active_profile.get("label", active_profile_slug)

    st.sidebar.markdown("### Active template")
    st.sidebar.write(f"Family: `{active_family_slug}`")
    st.sidebar.write(f"Template: `{active_profile_slug}`")
    st.sidebar.write(f"Workbook: `{active_profile.get('template_file', '')}`")
    st.sidebar.write(f"Variation theme: `{active_profile.get('variation_theme', '')}`")
    st.sidebar.checkbox("Show troubleshooting debug", key="show_header_debug", value=False)
    st.sidebar.checkbox("Copy row styles", key="copy_row_styles", value=True)
    st.sidebar.checkbox("Auto-load image mappings", key="auto_load_image_mappings", value=False)
    if st.sidebar.button("Refresh Dropbox queues", key="refresh_dropbox_queues_btn", width="stretch"):
        st.session_state["pending_perf_action_label"] = "refresh Dropbox queues"
        refresh_cached_folder_names("stage", "ready", "approved", "finished")
        clear_cached_listing_memory()
        st.rerun()

    colors_available = get_profile_color_options(active_profile)
    sizes_available = active_profile.get("sizes", [])
    include_garment_resource_overview = False
    dropbox_overview_cache_hit = (
        st.session_state.get("dropbox_overview_cache", {}).get("key")
        == build_dropbox_overview_cache_key(
            active_profile,
            dropbox_cfg,
            include_garment_resource_images=include_garment_resource_overview,
        )
    )
    t_dropbox_overview_start = time.perf_counter()
    dropbox_overview = get_cached_dropbox_overview(
        active_profile,
        dropbox_cfg,
        include_garment_resource_images=include_garment_resource_overview,
    )
    t_dropbox_overview_end = time.perf_counter()

    with st.sidebar.expander("Dropbox debug"):
        st.caption("Runs only when clicked so it cannot slow normal app loading.")
        if st.button("Test shared preview link", key="test_dropbox_preview_link_btn"):
            try:
                test_path = ""
                if dropbox_overview.get("shared_resource_images"):
                    test_path = dropbox_overview["shared_resource_images"][0]

                st.write("Test path:", test_path)
                if test_path:
                    st.write("Preview URL:", dropbox_preview_url(test_path))
            except Exception as exc:
                st.error(f"Dropbox debug failed: {exc}")


    parent_sku_from_config = str(get_default(active_profile, "parent_sku", "")).strip()

    staged_folder_name = None
    selected_finished_folder = None
    content_debug_container = None
    content_preflight_container = None

    if active_workflow_tab == "Product setup":
        folder_source, staged_folder_name, selected_finished_folder = render_product_setup_controls(
            staged_folder_names=staged_folder_names,
            finished_folder_names=finished_folder_names,
            finished_root=finished_root,
            dropbox_cfg=dropbox_cfg,
            profiles=profiles,
            profile=profile,
            families=families,
            family_labels=family_labels,
            detection_message=detection_message,
            detection_level=detection_level,
            selected_family=selected_family,
            selected_label=selected_label,
            workflow_assignees=WORKFLOW_ASSIGNEES,
            selectbox_index_without_state_conflict=selectbox_index_without_state_conflict,
            get_cached_folder_names=get_cached_folder_names,
            refresh_cached_folder_names=refresh_cached_folder_names,
            clear_cached_listing_memory=clear_cached_listing_memory,
            clear_runtime_caches=clear_runtime_caches,
            restage_finished_listing_for_review=restage_finished_listing_for_review,
            set_workflow_flash=set_workflow_flash,
            reset_restaged_selection_state=reset_restaged_selection_state,
            list_folder_names=list_folder_names,
            scan_staged_folder_readiness=scan_staged_folder_readiness,
            merchant_shipping_group_options=MERCHANT_SHIPPING_GROUP_OPTIONS,
            sku_decoration_options=SKU_DECORATION_OPTIONS,
            default_variant_quantity=DEFAULT_VARIANT_QUANTITY,
            get_default_sku_decoration_code=get_default_sku_decoration_code,
            sanitize_sku=sanitize_sku,
            generate_unique_sku=generate_unique_sku,
            get_default=get_default,
            build_parent_sku_from_context=build_parent_sku_from_context,
            create_listing_task=create_staged_listing_task_in_dropbox,
        )

    folder_source = st.session_state.get("active_folder_source_mode", folder_source)
    if folder_source == "Use staged folder":
        staged_folder_name = st.session_state.get("active_staged_folder_select") or active_staged_folder_name
    elif folder_source == "Restage finished folder":
        selected_finished_folder = st.session_state.get("active_finished_folder_select")

    listing_memory = dict(active_listing_memory)
    listing_context_key = ""
    if staged_folder_name:
        listing_context_key = f"{staged_folder_name}|{active_profile.get('template_key', active_profile_slug)}"

    if staged_folder_name and listing_memory:
        memory_fingerprint = json.dumps(
            {
                "folder": staged_folder_name,
                "profile": active_profile.get("template_key", active_profile_slug),
                "template_key": listing_memory.get("template_key", ""),
                "template_slug": listing_memory.get("template_slug", ""),
                "title": listing_memory.get("title", ""),
                "bullet_points": listing_memory.get("bullet_points", []),
                "product_description": listing_memory.get("product_description", ""),
                "generic_keywords": listing_memory.get("generic_keywords", ""),
                "selected_variants": listing_memory.get("selected_variants", {}),
                "size_price_map": listing_memory.get("size_price_map", {}),
                "listing_group": listing_memory.get("listing_group", {}),
                "sku_decoration_code": listing_memory.get("sku_decoration_code", ""),
                "manual_sku_listing_code": listing_memory.get("manual_sku_listing_code", ""),
                "generated_sku_listing_code": listing_memory.get("generated_sku_listing_code", ""),
                "sku_listing_code": listing_memory.get("sku_listing_code", ""),
                "quantity": normalize_variant_quantity(
                    listing_memory.get("quantity", DEFAULT_VARIANT_QUANTITY)
                ),
                "handling_time_days": normalize_handling_time_days(
                    listing_memory.get("handling_time_days", DEFAULT_HANDLING_TIME_DAYS)
                ),
                "merchant_shipping_group_name": normalize_merchant_shipping_group(
                    listing_memory.get("merchant_shipping_group_name", "")
                ),
            },
            sort_keys=True,
        )

        applied_memory_key = st.session_state.get("applied_listing_memory_key_v2", "")
        applied_widget_memory_key = st.session_state.get("applied_listing_memory_widget_key_v2", "")
        content_widget_keys = [
            "title_input",
            "bullet_1",
            "bullet_2",
            "bullet_3",
            "bullet_4",
            "bullet_5",
            "product_description",
            "generic_keywords",
        ]
        content_widgets_missing = any(key not in st.session_state for key in content_widget_keys)
        memory_has_content = any(
            str(value or "").strip()
            for value in [
                listing_memory.get("title", ""),
                listing_memory.get("product_description", ""),
                listing_memory.get("generic_keywords", ""),
                *listing_memory.get("bullet_points", []),
            ]
        )
        content_widgets_empty = all(
            not str(st.session_state.get(key, "") or "").strip()
            for key in content_widget_keys
        )
        listing_content_widgets_missing = (
            active_workflow_tab == "Listing content"
            and listing_content_widget_state_is_missing(
                st.session_state,
                listing_memory,
            )
        )
        should_apply_memory = (
            applied_memory_key != memory_fingerprint
            or content_widgets_missing
            or (memory_has_content and content_widgets_empty)
            or (
                active_workflow_tab == "Listing content"
                and (
                    applied_widget_memory_key != memory_fingerprint
                    or listing_content_widgets_missing
                )
            )
        )

        if should_apply_memory:
            apply_listing_memory_to_session(listing_memory, active_profile)
            st.session_state["applied_listing_memory_key_v2"] = memory_fingerprint
            if active_workflow_tab == "Listing content":
                st.session_state["applied_listing_memory_widget_key_v2"] = memory_fingerprint
            st.session_state["initialized_listing_context_key"] = listing_context_key
            st.session_state["last_loaded_listing_memory_signature"] = f"{staged_folder_name}|{active_profile_slug}"
    elif listing_context_key:
        initialized_context_key = st.session_state.get("initialized_listing_context_key", "")
        if initialized_context_key != listing_context_key:
            initialize_listing_context_defaults(active_profile)
            st.session_state["initialized_listing_context_key"] = listing_context_key

    if listing_memory:
        st.sidebar.info("Loaded saved listing inputs from staged folder.")

    for field_name in ["assets_prepared_by", "content_prepared_by", "reviewed_by", "prepared_at", "reviewed_at"]:
        if field_name not in st.session_state:
            current_value = str(listing_memory.get(field_name, ""))
            if field_name in {"assets_prepared_by", "content_prepared_by", "reviewed_by"} and current_value not in WORKFLOW_ASSIGNEES:
                current_value = ""
            st.session_state[field_name] = current_value

    title = st.session_state.get("title_input", listing_memory.get("title", ""))

    saved_bullets = listing_memory.get("bullet_points", [])
    saved_bullets = (saved_bullets + ["", "", "", "", ""])[:5]
    bullets = [
        st.session_state.get("bullet_1", saved_bullets[0]),
        st.session_state.get("bullet_2", saved_bullets[1]),
        st.session_state.get("bullet_3", saved_bullets[2]),
        st.session_state.get("bullet_4", saved_bullets[3]),
        st.session_state.get("bullet_5", saved_bullets[4]),
    ]

    product_description = st.session_state.get("product_description", listing_memory.get("product_description", ""))
    generic_keywords = st.session_state.get("generic_keywords", listing_memory.get("generic_keywords", ""))

    st.session_state.setdefault("title_input", title)
    for idx, bullet_value in enumerate(bullets, start=1):
        st.session_state.setdefault(f"bullet_{idx}", bullet_value)
    st.session_state.setdefault("product_description", product_description)
    st.session_state.setdefault("generic_keywords", generic_keywords)
    st.session_state.setdefault(CONTENT_EDITOR_KEYS["title"], title)
    for idx, bullet_value in enumerate(bullets):
        st.session_state.setdefault(CONTENT_EDITOR_KEYS["bullets"][idx], bullet_value)
    st.session_state.setdefault(CONTENT_EDITOR_KEYS["description"], product_description)
    st.session_state.setdefault(CONTENT_EDITOR_KEYS["keywords"], generic_keywords)
    st.session_state.setdefault(
        "variant_quantity",
        normalize_variant_quantity(listing_memory.get("quantity", DEFAULT_VARIANT_QUANTITY)),
    )
    st.session_state.setdefault(
        "handling_time_days",
        normalize_handling_time_days(listing_memory.get("handling_time_days", DEFAULT_HANDLING_TIME_DAYS)),
    )
    st.session_state.setdefault(
        "merchant_shipping_group_name",
        normalize_merchant_shipping_group(listing_memory.get("merchant_shipping_group_name", "")),
    )

    profile = active_profile
    variant_dimensions = active_profile.get("variant_dimensions", [])
    saved_selected_variants = listing_memory.get("selected_variants", {})
    selected_variants = normalize_selected_variants_session_state(active_profile, listing_memory)

    auto_load_image_mappings = bool(st.session_state.get("auto_load_image_mappings", False))
    load_image_mappings_now = bool(st.session_state.pop("load_image_mappings_now", False))
    scan_mapped_colours_now = bool(st.session_state.pop("scan_mapped_colours_now", False))
    manual_image_load_requested = bool(load_image_mappings_now and staged_folder_name)
    auto_apply_mapped_colors = False

    # Image mappings should persist while editing listing content.
    # Treat mappings as loaded for the staged folder + template, not for every selected colour/size change.
    image_mapping_context_key = json.dumps(
        {
            "folder": staged_folder_name or "",
            "template_slug": active_profile.get("_slug", ""),
            "template_key": active_profile.get("template_key", ""),
        },
        sort_keys=True,
    )

    persisted_image_mappings_loaded = bool(
        staged_folder_name
        and st.session_state.get("image_mappings_loaded_folder") == staged_folder_name
        and st.session_state.get("image_mappings_loaded_context") == image_mapping_context_key
    )

    image_mappings_stale = bool(
        staged_folder_name
        and st.session_state.get("image_mappings_loaded_folder") == staged_folder_name
        and st.session_state.get("image_mappings_loaded_context") != image_mapping_context_key
    )

    if manual_image_load_requested:
        st.session_state["image_mappings_loaded_folder"] = staged_folder_name
        st.session_state["image_mappings_loaded_context"] = image_mapping_context_key
        st.session_state.pop("preview_image_cache", None)
        st.session_state.pop("preview_image_mapping_cache", None)
        st.session_state.pop("resolved_image_bundle_cache", None)
        persisted_image_mappings_loaded = True
        image_mappings_stale = False

    # Only explicit load/auto-load may build image mappings.
    # Previously-loaded mappings are included only when the current folder/template/variant context matches.
    should_load_image_mappings = bool(staged_folder_name) and (
        auto_load_image_mappings
        or manual_image_load_requested
        or persisted_image_mappings_loaded
        or auto_apply_mapped_colors
        or scan_mapped_colours_now
    )

    if scan_mapped_colours_now and staged_folder_name:
        image_resolution_reason = "mapped_colour_scan"
    elif auto_apply_mapped_colors and staged_folder_name:
        image_resolution_reason = "auto_mapped_colours"
    elif auto_load_image_mappings and staged_folder_name:
        image_resolution_reason = "auto_load"
    elif manual_image_load_requested:
        image_resolution_reason = "manual_load"
    elif persisted_image_mappings_loaded and staged_folder_name:
        image_resolution_reason = "cache_reuse"
    elif image_mappings_stale:
        image_resolution_reason = "stale_context"
    else:
        image_resolution_reason = ""

    image_preview_variants = selected_variants
    resolve_preview_image_urls = bool(
        should_load_image_mappings
        and image_resolution_reason not in {"auto_mapped_colours", "mapped_colour_scan"}
    )

    if should_load_image_mappings and image_resolution_reason not in {"auto_mapped_colours", "mapped_colour_scan"}:
        dropbox_overview = get_cached_dropbox_overview(
            active_profile,
            dropbox_cfg,
            include_garment_resource_images=True,
        )

    preview_image_cache_hit = False
    t_preview_image_start = time.perf_counter()
    preview_image_data: dict[str, Any] = {}
    if should_load_image_mappings:
        preview_image_cache_hit = (
            st.session_state.get("preview_image_cache", {}).get("key")
            == build_preview_image_cache_key(
                profile,
                dropbox_cfg,
                staged_folder_name or "",
                image_preview_variants,
                should_load_image_mappings,
                resolve_preview_image_urls,
            )
        )
        preview_image_data = get_cached_preview_image_data(
            profile=profile,
            dropbox_cfg=dropbox_cfg,
            staged_folder_name=staged_folder_name or "",
            selected_variants=image_preview_variants,
            dropbox_overview=dropbox_overview,
            include_mappings=should_load_image_mappings,
            resolve_preview_urls=resolve_preview_image_urls,
        )
    t_preview_image_end = time.perf_counter()
    record_load_event(
        "Images: preview/mapping data",
        t_preview_image_start,
        "with mappings" if should_load_image_mappings else "paths only",
    )
    staged_preview_paths = preview_image_data.get("staged_preview_paths", [])
    staged_preview_entries = preview_image_data.get("staged_preview_entries", [])
    staged_resource_paths = preview_image_data.get("staged_resource_paths", [])
    staged_resource_entries = preview_image_data.get("staged_resource_entries", [])
    design_color_preview_entries = preview_image_data.get("design_color_preview_entries", [])
    parent_main_image_options = preview_image_data.get("parent_main_image_options", [])
    garment_resource_entries = preview_image_data.get("garment_resource_entries", [])
    garment_resource_group_root_entries = preview_image_data.get("garment_resource_group_root_entries", [])
    garment_resource_group_entries = preview_image_data.get("garment_resource_group_entries", [])
    global_resource_entries = preview_image_data.get("global_resource_entries", [])
    staged_variant_entries = preview_image_data.get("staged_variant_entries", [])
    preview_color_image_map = preview_image_data.get("color_image_map", {})
    preview_design_color_image_url_map = preview_image_data.get("design_color_image_url_map", {})
    full_preview_color_image_map = preview_image_data.get("full_color_image_map", preview_color_image_map)
    full_preview_design_color_image_url_map = preview_image_data.get(
        "full_design_color_image_url_map",
        preview_design_color_image_url_map,
    )

    pending_mapped_colour_key = st.session_state.get("apply_mapped_colours_widget_key", "")
    if pending_mapped_colour_key and should_load_image_mappings:
        pending_valid_colours = (
            get_profile_color_options(profile)
            or list(selected_variants.get("color", []))
            or list(selected_variants.get("colour", []))
        )
        mapped_colours_to_apply = get_mapped_color_options(
            pending_valid_colours,
            full_preview_color_image_map,
            full_preview_design_color_image_url_map,
        )
        if mapped_colours_to_apply:
            st.session_state[pending_mapped_colour_key] = list(mapped_colours_to_apply)
            st.session_state.pop("apply_mapped_colours_widget_key", None)
            st.rerun()

    price_dimension_values = selected_variants.get("size", ["default"])
    saved_prices = normalize_saved_price_map_for_profile(
        profile,
        listing_memory.get("size_price_map", {}),
    )
    existing_values = [saved_prices.get(size) for size in price_dimension_values if size in saved_prices]
    unique_existing_values = {v for v in existing_values if v is not None}
    default_same_price = bool(price_dimension_values) and len(unique_existing_values) == 1 and len(existing_values) == len(price_dimension_values)
    use_same_price = st.session_state.get("use_same_price_for_all_sizes", default_same_price)

    if use_same_price:
        fallback_price = (
            float(saved_prices.get(price_dimension_values[0], 29.99))
            if default_same_price and price_dimension_values
            else get_default_price_for_size(profile, price_dimension_values[0], saved_prices)
            if price_dimension_values
            else 29.99
        )
        shared_price = float(st.session_state.get("shared_price_all_sizes", fallback_price))
        size_price_map = {size: shared_price for size in price_dimension_values}
    else:
        size_price_map = {
            size: float(
                st.session_state.get(
                    f"price_{size}",
                    get_default_price_for_size(profile, size, saved_prices),
                )
            )
            for size in price_dimension_values
        }

    quantity = normalize_variant_quantity(
        st.session_state.get(
            "variant_quantity",
            listing_memory.get("quantity", DEFAULT_VARIANT_QUANTITY),
        )
    )
    st.session_state["variant_quantity"] = quantity
    selected_parent_main_label = st.session_state.get("parent_main_image_choice", "Automatic (recommended)")
    selected_parent_main_image_url = resolve_selected_parent_main_image_url(
        parent_main_image_options,
        selected_parent_main_image_label=selected_parent_main_label,
    )
    st.session_state.setdefault("use_resource_fallback_images", False)
    use_resource_fallback_images = bool(st.session_state.get("use_resource_fallback_images", False))
    preview_parent_main_image_url = (
        selected_parent_main_image_url
        or (parent_main_image_options[0][1] if parent_main_image_options else "")
    )
    resolved_image_bundle = {
        "parent_main_image_url": preview_parent_main_image_url if preview_parent_main_image_url else "",
        "other_images": [],
        "color_image_map": preview_color_image_map,
        "design_color_image_url_map": preview_design_color_image_url_map,
    }
    resolved_image_error = ""
    partial_image_mappings_loaded = False
    image_mappings_loaded_this_run = False
    current_resolved_image_cache_key = ""
    if staged_folder_name:
        current_resolved_image_cache_key = build_resolved_image_bundle_cache_key(
            profile=profile,
            dropbox_cfg=dropbox_cfg,
            staged_folder_name=staged_folder_name,
            selected_variants=image_preview_variants,
            selected_parent_main_image_label=selected_parent_main_label,
            selected_parent_main_image_url=selected_parent_main_image_url,
            use_resource_fallback_images=use_resource_fallback_images,
        )
    resolved_image_bundle_cache_hit = bool(
        current_resolved_image_cache_key
        and st.session_state.get("resolved_image_bundle_cache", {}).get("key") == current_resolved_image_cache_key
    )
    should_resolve_image_bundle = bool(
        staged_folder_name
        and should_load_image_mappings
        and image_resolution_reason not in {"auto_mapped_colours", "mapped_colour_scan"}
    )
    t_resolved_image_start = time.perf_counter()
    if staged_folder_name and (should_resolve_image_bundle or resolved_image_bundle_cache_hit):
        try:
            resolved_image_bundle = get_cached_resolved_image_bundle(
                profile=profile,
                dropbox_cfg=dropbox_cfg,
                staged_folder_name=staged_folder_name,
                selected_variants=image_preview_variants,
                dropbox_overview=dropbox_overview,
                selected_parent_main_image_label=selected_parent_main_label,
                selected_parent_main_image_url=selected_parent_main_image_url,
                use_resource_fallback_images=use_resource_fallback_images,
            )
            image_mappings_loaded_this_run = should_resolve_image_bundle and not resolved_image_bundle_cache_hit
        except Exception as exc:
            resolved_image_bundle = {
                "parent_main_image_url": preview_parent_main_image_url if preview_parent_main_image_url else "",
                "other_images": [],
                "color_image_map": preview_color_image_map,
                "design_color_image_url_map": preview_design_color_image_url_map,
            }
            partial_image_mappings_loaded = bool(
                preview_color_image_map or preview_design_color_image_url_map
            )
            if not partial_image_mappings_loaded:
                resolved_image_error = str(exc)
    t_resolved_image_end = time.perf_counter()
    record_load_event(
        "Images: resolved image bundle",
        t_resolved_image_start,
        image_resolution_reason or "cache/no-load",
    )
    preview_parent_main_image_url = resolved_image_bundle.get("parent_main_image_url", preview_parent_main_image_url)
    preview_other_images = list(resolved_image_bundle.get("other_images", []))
    preview_color_image_map = dict(resolved_image_bundle.get("color_image_map", preview_color_image_map))
    preview_design_color_image_url_map = dict(
        resolved_image_bundle.get("design_color_image_url_map", preview_design_color_image_url_map)
    )
    image_mappings_loaded = bool(
        (
            current_resolved_image_cache_key
            and st.session_state.get("resolved_image_bundle_cache", {}).get("key") == current_resolved_image_cache_key
        )
        or partial_image_mappings_loaded
    )
    if not staged_folder_name:
        image_mapping_status = "not_loaded"
        image_mapping_detail = "Select a staged folder to load image mappings."
    elif resolved_image_error:
        image_mapping_status = "error"
        image_mapping_detail = resolved_image_error
    elif image_mappings_loaded:
        image_mapping_status = "loaded"
        if partial_image_mappings_loaded:
            image_mapping_detail = "Partial image mappings loaded from staged filenames. Use Mapped colours to select only those colours."
        else:
            image_mapping_detail = "Image mappings loaded."
    elif image_mappings_stale:
        image_mapping_status = "not_loaded"
        image_mapping_detail = "Image mappings need refresh because the folder or template changed. Click Load / refresh image mappings to update them."
    else:
        image_mapping_status = "not_loaded"
        image_mapping_detail = "Image mappings not loaded yet. Use Load / refresh image mappings when you need image review or full checks."

    if active_workflow_tab == "Product setup":
        render_product_setup(
            active_staged_folder_name=active_staged_folder_name,
            active_template_label=active_template_label,
            selected_parent_main_label=selected_parent_main_label,
            preview_parent_main_image_url=preview_parent_main_image_url,
            preview_color_image_map=preview_color_image_map,
            preview_design_color_image_url_map=preview_design_color_image_url_map,
            preview_other_images=preview_other_images,
            image_mapping_status=image_mapping_status,
            image_mapping_detail=image_mapping_detail,
            dropbox_cfg=dropbox_cfg,
            staged_folder_name=staged_folder_name,
            image_mapping_context_key=image_mapping_context_key,
            dropbox_overview=dropbox_overview,
            staged_resource_paths=staged_resource_paths,
            staged_preview_paths=staged_preview_paths,
            selected_variants=selected_variants,
            staged_resource_entries=staged_resource_entries,
            garment_resource_entries=garment_resource_entries,
            garment_resource_group_root_entries=garment_resource_group_root_entries,
            garment_resource_group_entries=garment_resource_group_entries,
            global_resource_entries=global_resource_entries,
            parent_main_image_options=parent_main_image_options,
            variant_dimensions=variant_dimensions,
            design_color_preview_entries=design_color_preview_entries,
            staged_variant_entries=staged_variant_entries,
            profile=profile,
            parent_sku_from_config=parent_sku_from_config,
            title=title,
            active_profile=active_profile,
            global_brand_name=GLOBAL_BRAND_NAME,
            render_active_product_context=render_active_product_context,
            render_stage_images_zip_upload=render_stage_images_zip_upload,
            upload_resource_images_to_folder=upload_resource_images_to_folder,
            clear_resource_image_caches=clear_resource_image_caches,
            build_variants_summary=build_variants_summary,
            render_path_grid=render_path_grid,
            selectbox_index_without_state_conflict=selectbox_index_without_state_conflict,
            render_design_color_grid=render_design_color_grid,
            render_color_grid=render_color_grid,
            render_variant_combinations_preview=render_variant_combinations_preview,
            get_default=get_default,
            get_stock_reference=get_stock_reference,
            is_strict_stock_ready=is_strict_stock_ready,
        )

    score_clicked = False
    ready_clicked = False

    if active_workflow_tab == "Listing content":
        listing_content_result = render_listing_content(
            active_staged_folder_name=active_staged_folder_name,
            active_template_label=active_template_label,
            selected_parent_main_label=selected_parent_main_label,
            preview_parent_main_image_url=preview_parent_main_image_url,
            preview_color_image_map=preview_color_image_map,
            preview_design_color_image_url_map=preview_design_color_image_url_map,
            preview_other_images=preview_other_images,
            image_mapping_status=image_mapping_status,
            image_mapping_detail=image_mapping_detail,
            staged_folder_name=staged_folder_name,
            listing_memory=listing_memory,
            active_profile=active_profile,
            profile=profile,
            memory_fingerprint=locals().get("memory_fingerprint", ""),
            grouped_state_load_error=grouped_state_load_error,
            listing_memory_location=listing_memory_location,
            CONTENT_EDITOR_KEYS=CONTENT_EDITOR_KEYS,
            MERCHANT_SHIPPING_GROUP_OPTIONS=MERCHANT_SHIPPING_GROUP_OPTIONS,
            SKU_DECORATION_OPTIONS=SKU_DECORATION_OPTIONS,
            WORKFLOW_ASSIGNEES=WORKFLOW_ASSIGNEES,
            variant_dimensions=variant_dimensions,
            selected_variants=selected_variants,
            colors_available=colors_available,
            full_preview_color_image_map=full_preview_color_image_map,
            full_preview_design_color_image_url_map=full_preview_design_color_image_url_map,
            auto_apply_mapped_colors=auto_apply_mapped_colors,
            image_mapping_context_key=image_mapping_context_key,
            render_active_product_context=render_active_product_context,
            listing_memory_has_content=listing_memory_has_content,
            apply_listing_memory_to_session=apply_listing_memory_to_session,
            words_repeated_at_least=words_repeated_at_least,
            find_forbidden_title_phrases_for_app=find_forbidden_title_phrases_for_app,
            sync_content_editor_to_canonical_state=sync_content_editor_to_canonical_state,
            trim_search_terms=trim_search_terms,
            normalize_merchant_shipping_group=normalize_merchant_shipping_group,
            selectbox_index_without_state_conflict=selectbox_index_without_state_conflict,
            get_mapped_color_options=get_mapped_color_options,
            apply_mapped_colors_to_widget_once=apply_mapped_colors_to_widget_once,
            get_available_sizes_for_selected_colors=get_available_sizes_for_selected_colors,
            normalize_multiselect_values=normalize_multiselect_values,
            get_default_sku_decoration_code=get_default_sku_decoration_code,
            sanitize_sku=sanitize_sku,
            get_or_create_generated_sku_listing_code=get_or_create_generated_sku_listing_code,
            build_parent_sku_from_context=build_parent_sku_from_context,
            render_variant_combinations_preview=render_variant_combinations_preview,
            build_size_price_inputs=build_size_price_inputs,
            load_grouped_image_manifest=lambda folder_name, selected_profile: (
                load_grouped_christmas_image_manifest_from_dropbox(
                    dropbox_cfg=dropbox_cfg,
                    staged_folder_name=folder_name,
                    profile=selected_profile,
                )
            ),
            save_grouped_draft=lambda selected_profile, payload, folder_name: (
                save_grouped_christmas_draft_to_dropbox(
                    dropbox_cfg=dropbox_cfg,
                    staged_folder_name=folder_name,
                    profile=selected_profile,
                    payload=payload,
                )
            ),
            submit_grouped_listing=lambda selected_profile, payload, folder_name: (
                submit_grouped_christmas_to_review(
                    dropbox_cfg=dropbox_cfg,
                    staged_folder_name=folder_name,
                    profile=selected_profile,
                    draft_payload=payload,
                    profiles=profiles,
                )
            ),
            dev_tools_enabled=dev_tools_enabled(os.environ, st.secrets),
            load_grouped_test_json=load_grouped_christmas_test_content,
        )
        title = listing_content_result["title"]
        bullets = listing_content_result["bullets"]
        product_description = listing_content_result["product_description"]
        generic_keywords = listing_content_result["generic_keywords"]
        handling_time_days = listing_content_result["handling_time_days"]
        selected_variants = listing_content_result["selected_variants"]
        sku_decoration_code = listing_content_result["sku_decoration_code"]
        manual_sku_listing_code = listing_content_result["manual_sku_listing_code"]
        generated_sku_listing_code = listing_content_result["generated_sku_listing_code"]
        sku_listing_code = listing_content_result["sku_listing_code"]
        parent_sku_for_listing = listing_content_result["parent_sku_for_listing"]
        size_price_map = listing_content_result["size_price_map"]
        quantity = listing_content_result["quantity"]
        score_clicked = listing_content_result["score_clicked"]
        ready_clicked = listing_content_result["ready_clicked"]
        content_debug_container = listing_content_result["content_debug_container"]
        content_preflight_container = listing_content_result["content_preflight_container"]
        content_action_result_container = listing_content_result["content_action_result_container"]

    if active_workflow_tab == "Review queue":
        render_review_queue(
            ready_folder_names=ready_folder_names,
            ready_root=ready_root,
            profiles=profiles,
            dropbox_cfg=dropbox_cfg,
            refresh_cached_folder_names=refresh_cached_folder_names,
            clear_cached_listing_memory=clear_cached_listing_memory,
            get_cached_folder_names=get_cached_folder_names,
            render_review_queue_view=render_review_queue_view,
        )

    if active_workflow_tab == "Approved output":
        approved_output_kwargs = {
            "finished_folder_names": finished_folder_names,
            "approved_folder_names": approved_folder_names,
            "approved_root": approved_root,
            "dropbox_cfg": dropbox_cfg,
            "profiles": profiles,
            "profile": profile,
            "WORKFLOW_ASSIGNEES": WORKFLOW_ASSIGNEES,
            "restage_finished_listing_for_review": restage_finished_listing_for_review,
            "mark_finished_generation_ignored": mark_finished_generation_ignored,
            "refresh_cached_folder_names": refresh_cached_folder_names,
            "clear_cached_listing_memory": clear_cached_listing_memory,
            "clear_runtime_caches": clear_runtime_caches,
            "set_workflow_flash": set_workflow_flash,
            "get_cached_folder_names": get_cached_folder_names,
            "render_approved_queue_view": render_approved_queue_view,
        }
        if "build_finished_generation_history_rows" in inspect.signature(render_approved_output).parameters:
            approved_output_kwargs["build_finished_generation_history_rows"] = (
                build_finished_generation_history_rows
            )
        render_approved_output(**approved_output_kwargs)

    render_inline_loading_debug()
    render_rerun_cause_debug()
    save_debug_state_snapshot()

    if not score_clicked and not ready_clicked:
        return

    size_price_map = normalize_variant_price_map_for_selected_variants(
        profile,
        selected_variants,
        st.session_state.get("current_size_price_map", size_price_map),
    )

    if staged_folder_name and not image_mappings_loaded:
        image_resolution_reason = "submit_review" if ready_clicked else "score_check"
        with st.spinner("Loading image mappings for quality checks..."):
            try:
                resolved_image_bundle = get_cached_resolved_image_bundle(
                    profile=profile,
                    dropbox_cfg=dropbox_cfg,
                    staged_folder_name=staged_folder_name,
                    selected_variants=selected_variants,
                    dropbox_overview=dropbox_overview,
                    selected_parent_main_image_label=selected_parent_main_label,
                    selected_parent_main_image_url=selected_parent_main_image_url,
                    use_resource_fallback_images=use_resource_fallback_images,
                )
                image_mappings_loaded_this_run = True
                resolved_image_bundle_cache_hit = False
                resolved_image_error = ""
                preview_parent_main_image_url = resolved_image_bundle.get("parent_main_image_url", preview_parent_main_image_url)
                preview_other_images = list(resolved_image_bundle.get("other_images", []))
                preview_color_image_map = dict(resolved_image_bundle.get("color_image_map", preview_color_image_map))
                preview_design_color_image_url_map = dict(
                    resolved_image_bundle.get("design_color_image_url_map", preview_design_color_image_url_map)
                )
                image_mappings_loaded = True
                image_mapping_status = "loaded"
                image_mapping_detail = "Image mappings loaded."
            except Exception as exc:
                resolved_image_error = str(exc)
                image_mapping_status = "error"
                image_mapping_detail = resolved_image_error

    preflight = build_preflight_report(
        profile=profile,
        dropbox_cfg=dropbox_cfg,
        dropbox_overview=dropbox_overview,
        staged_folder_name=staged_folder_name or "",
        title=title,
        bullets=bullets,
        product_description=product_description,
        generic_keywords=generic_keywords,
        selected_variants=selected_variants,
        size_price_map=size_price_map,
        quantity=quantity,
        sku_decoration_code=sku_decoration_code,
        sku_listing_code=sku_listing_code,
        resolved_parent_main_image_url=preview_parent_main_image_url,
        resolved_other_images=preview_other_images,
        resolved_color_image_map=preview_color_image_map,
        resolved_design_color_image_url_map=preview_design_color_image_url_map,
        allow_image_resolution_fallback=True,
        use_resource_fallback_images=use_resource_fallback_images,
        has_staged_resource_images=bool(staged_resource_paths),
    )

    preview_payload = preflight["preview_payload"]
    all_preview_errors = preflight["all_preview_errors"]
    quality_report = preflight["quality_report"]

    if content_debug_container is not None and st.session_state.get("show_header_debug", False):
        with content_debug_container:
            with st.expander("Listing content image debug", expanded=False):
                st.write(
                    "cache_timings",
                    {
                        "dropbox_overview": {
                            "cache_hit": dropbox_overview_cache_hit,
                            "seconds": round(t_dropbox_overview_end - t_dropbox_overview_start, 4),
                        },
                        "preview_image_data": {
                            "cache_hit": preview_image_cache_hit,
                            "seconds": round(t_preview_image_end - t_preview_image_start, 4),
                        },
                        "resolved_image_bundle": {
                            "cache_hit": resolved_image_bundle_cache_hit,
                            "seconds": round(t_resolved_image_end - t_resolved_image_start, 4),
                            "loaded_this_run": image_mappings_loaded_this_run,
                            "reason": image_resolution_reason,
                            "status": image_mapping_status,
                        },
                    },
                )
                st.write("selected_variants", selected_variants)
                st.write("parent_main_image_options", parent_main_image_options)
                st.write("selected_parent_main_image_url", selected_parent_main_image_url)
                st.write("preview_parent_main_image_url", preview_parent_main_image_url)
                st.write("preview_color_image_map", preview_color_image_map)
                st.write("preview_design_color_image_url_map", preview_design_color_image_url_map)
                st.write("preview_other_images", preview_other_images)
                st.write(
                    "preview_payload_image_fields",
                    {
                        "parent_main_image_url": preview_payload.get("parent_main_image_url", ""),
                        "other_images": preview_payload.get("other_images", []),
                        "color_image_map": preview_payload.get("color_image_map", {}),
                        "design_color_image_url_map": preview_payload.get("design_color_image_url_map", {}),
                    },
                )

    if (score_clicked or ready_clicked) and content_preflight_container is not None:
        with content_preflight_container:
            render_preflight_dashboard(
                quality_report=quality_report,
                all_preview_errors=all_preview_errors,
            )
            render_listing_score_result(
                quality_report=quality_report,
                all_preview_errors=all_preview_errors,
            )

    if score_clicked and not ready_clicked:
        st.stop()

    generation_prep = prepare_generation_payload(
        profile=profile,
        title=title,
        bullets=bullets,
        product_description=product_description,
        generic_keywords=generic_keywords,
        selected_variants=selected_variants,
        size_price_map=size_price_map,
        sku_decoration_code=sku_decoration_code,
        sku_listing_code=sku_listing_code,
        manual_sku_listing_code=manual_sku_listing_code,
        generated_sku_listing_code=generated_sku_listing_code,
        quantity=quantity,
        staged_folder_name=staged_folder_name or "",
        handling_time_days=normalize_handling_time_days(handling_time_days),
        merchant_shipping_group_name=normalize_merchant_shipping_group(
            st.session_state.get("merchant_shipping_group_name", "")
        ),
        parent_main_image_choice=selected_parent_main_label,
        parent_main_image_url=preview_parent_main_image_url or selected_parent_main_image_url,
    )

    generation_payload = generation_prep["payload"]
    if "mpn" in listing_memory:
        generation_payload["mpn"] = listing_memory.get("mpn")
    original_finished_folder_name = str(listing_memory.get("original_finished_folder_name", "")).strip()
    if original_finished_folder_name:
        generation_payload["original_finished_folder_name"] = original_finished_folder_name
    generation_payload["assets_prepared_by"] = st.session_state.get("assets_prepared_by", "")
    generation_payload["content_prepared_by"] = st.session_state.get("content_prepared_by", "")
    generation_payload["reviewed_by"] = st.session_state.get("reviewed_by", "")
    generation_payload["prepared_at"] = st.session_state.get("prepared_at", "")
    generation_payload["reviewed_at"] = format_workflow_timestamp()
    generation_errors = generation_prep["errors"]
    action_label = "submit this listing for review" if ready_clicked else "generate"

    if generation_errors:
        st.error(f"Fix the validation errors before trying to {action_label}.")
        st.stop()

    if quality_report["blockers"]:
        st.error(f"Fix the listing quality blockers before trying to {action_label}.")
        st.stop()

    if ready_clicked:
        try:
            staged_folder_name = staged_folder_name or ""
            actual_staged_folder_name = resolve_existing_stage_folder_name(
                dropbox_cfg,
                staged_folder_name,
                staged_folder_names,
            )
            if actual_staged_folder_name != staged_folder_name:
                staged_folder_name = actual_staged_folder_name
                st.session_state["active_staged_folder_select"] = actual_staged_folder_name
                st.session_state["pending_staged_folder_selection_on_rerun"] = actual_staged_folder_name
            ready_folder_name = actual_staged_folder_name
            stage_folder_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)
            generation_payload["prepared_at"] = format_workflow_timestamp()
            st.session_state["prepared_at"] = generation_payload["prepared_at"]

            generation_payload["review_snapshot"] = build_review_snapshot(
                profile=profile,
                payload=generation_payload,
                dropbox_cfg=dropbox_cfg,
                folder_path=stage_folder_path,
                quality_report=quality_report,
                preview_errors=all_preview_errors,
            )
            append_workflow_event(
                generation_payload,
                action="submit_for_review",
                actor=st.session_state.get("content_prepared_by", "") or st.session_state.get("assets_prepared_by", ""),
                from_state="_stage",
                to_state="ready",
                folder_path=stage_folder_path,
                details={
                    "assets_prepared_by": st.session_state.get("assets_prepared_by", ""),
                    "content_prepared_by": st.session_state.get("content_prepared_by", ""),
                },
            )

            listing_memory_path = save_listing_inputs_json_to_dropbox(
                profile=profile,
                payload=generation_payload,
                folder_path=stage_folder_path,
            )
            ready_folder_path = move_staged_dropbox_folder_to_ready(
                dropbox_cfg=dropbox_cfg,
                staged_folder_name=staged_folder_name,
                ready_folder_name=ready_folder_name,
            )

            st.session_state["last_ready_folder_path"] = ready_folder_path
            st.session_state["clear_staged_folder_selection_on_rerun"] = True
            clear_runtime_caches()
            set_workflow_flash(
                "success",
                f"Submitted for review: {Path(ready_folder_path).name}",
                f"Saved listing inputs to {listing_memory_path} and moved the folder to ready for admin review.",
            )
            st.rerun()
        except Exception as exc:
            target_container = content_action_result_container or st.container()
            with target_container:
                st.error(f"Could not submit the listing for review: {exc}")

            st.stop()

    try:
        progress_text = st.empty()
        progress_bar = st.progress(0)

        t0 = time.perf_counter()

        staged_folder_name = staged_folder_name or ""
        parent_sku_from_config = generation_payload["parent_sku"]
        selected_colors = generation_payload["colors"]
        selected_variants = generation_payload["selected_variants"]
        stage_folder_path = build_stage_folder_path(dropbox_cfg, staged_folder_name)

        if st.session_state.get("finalized_stage_folder") == staged_folder_name:
            progress_text.error("This staged folder was already finalized in the current session.")
            st.stop()

        progress_text.write("Checking workbook template...")
        progress_bar.progress(10)

        template_path = resolve_template_path(profile)
        wb = load_workbook(template_path, keep_vba=True, read_only=True)
        wb.close()
        t1 = time.perf_counter()

        progress_text.write("Checking staged Dropbox assets...")
        progress_bar.progress(20)

        resolve_folder_image_urls(
            profile,
            selected_variants,
            selected_colors,
            dropbox_overview,
            stage_folder_path,
            selected_parent_main_image_label=selected_parent_main_label,
            selected_parent_main_image_url=selected_parent_main_image_url,
        )
        t2 = time.perf_counter()

        progress_text.write("Moving staged folder into finished...")
        progress_bar.progress(35)

        final_sku, finished_folder_path = finalize_staged_dropbox_folder(
            dropbox_cfg=dropbox_cfg,
            staged_folder_name=staged_folder_name,
            parent_sku=parent_sku_from_config,
            reuse_finished_folder_name=original_finished_folder_name,
        )
        t3 = time.perf_counter()

        st.session_state["finalized_stage_folder"] = staged_folder_name
        st.session_state["finalized_finished_folder_path"] = finished_folder_path
        st.session_state["finalized_sku"] = final_sku

        progress_text.write("Fetching Dropbox image links...")
        progress_bar.progress(50)

        parent_main_image_url, other_images, color_image_map, design_color_image_url_map = resolve_folder_image_urls(
            profile,
            selected_variants,
            selected_colors,
            dropbox_overview,
            finished_folder_path,
            selected_parent_main_image_label=selected_parent_main_label,
            selected_parent_main_image_url=selected_parent_main_image_url,
        )
        t4 = time.perf_counter()

        payload = dict(generation_payload)
        payload["parent_sku"] = final_sku
        payload["parent_main_image_url"] = parent_main_image_url
        payload["other_images"] = other_images
        payload["color_image_map"] = color_image_map
        payload["design_color_image_url_map"] = design_color_image_url_map


        progress_text.write("Building workbook...")
        progress_bar.progress(75)

        output_path, workbook_timings = build_workbook(profile, payload)

        generated_artifact = save_generated_artifacts_to_dropbox(
            profile=profile,
            payload=payload,
            finished_folder_path=finished_folder_path,
            output_path=output_path,
        )
        append_workflow_event(
            payload,
            action="generate_staged_listing",
            actor=st.session_state.get("content_prepared_by", "") or st.session_state.get("assets_prepared_by", ""),
            from_state="_stage",
            to_state="finished",
            folder_path=finished_folder_path,
            details={
                "output_name": output_path.name,
                "workbook_dropbox_path": generated_artifact.get("workbook_dropbox_path", ""),
                "sku_manifest_dropbox_path": generated_artifact.get("sku_manifest_dropbox_path", ""),
                "child_sku_count": generated_artifact.get("child_sku_count", 0),
                "missing_supplier_stock_key_count": generated_artifact.get("missing_supplier_stock_key_count", 0),
            },
        )

        listing_memory_path = save_listing_memory_to_dropbox(
            profile=profile,
            payload=payload,
            folder_path=finished_folder_path,
        )

        t5 = time.perf_counter()

        progress_text.write("Finalizing output...")
        progress_bar.progress(95)

        variant_combos = build_variant_combinations(profile, selected_variants)
        child_count = len(variant_combos)

        progress_bar.progress(100)
        progress_text.success("Workbook generated successfully.")

        st.success(f"Workbook generated successfully: {output_path.name}")
        st.info(f"Generated 1 parent row and {child_count} child variants.")

        with st.expander("Performance breakdown", expanded=False):
            st.write(f"Check workbook template: {t1 - t0:.2f}s")
            st.write(f"Check staged Dropbox assets: {t2 - t1:.2f}s")
            st.write(f"Move staged folder: {t3 - t2:.2f}s")
            st.write(f"Resolve Dropbox image URLs: {t4 - t3:.2f}s")
            st.write(f"Build workbook: {t5 - t4:.2f}s")
            st.write(f"Total: {t5 - t0:.2f}s")
            st.write("---")
            st.write(f"Load workbook: {workbook_timings['load_workbook']:.2f}s")
            st.write(f"Write parent row: {workbook_timings['write_parent_row']:.2f}s")
            st.write(f"Write child rows: {workbook_timings['write_child_rows']:.2f}s")
            st.write(f"Save workbook: {workbook_timings['save_workbook']:.2f}s")

        with output_path.open("rb") as f:
            st.download_button(
                label="Download Amazon workbook",
                data=f.read(),
                file_name=output_path.name,
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
            )

    except Exception as exc:
        if "progress_bar" in locals():
            progress_bar.progress(100)
        if "progress_text" in locals():
            progress_text.error("Generation failed.")

        st.error("Workbook generation failed after finalizing Dropbox assets.")
        if "finished_folder_path" in locals():
            st.write(f"Finalized folder path: `{finished_folder_path}`")
        if "final_sku" in locals():
            st.write(f"Finalized SKU: `{final_sku}`")
        st.exception(exc)




    record_load_event(
        "Total: reached end of main",
        st.session_state.get("current_rerun_started_at", time.perf_counter()),
    )
    save_completed_load_events()


def run_app_safely() -> None:
    try:
        main()
    except Exception as exc:
        if type(exc).__name__ in {"RerunException", "StopException"}:
            raise
        error_row = record_app_error(exc)
        render_app_error_report(error_row)


if __name__ == "__main__":
    run_app_safely()
