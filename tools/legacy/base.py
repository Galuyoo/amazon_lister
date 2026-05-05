from pathlib import Path
from openpyxl import load_workbook

SOURCE_FILE = Path("HOODIE.xlsm")
OUTPUT_FILE = Path("amazon_hoodie_base.xlsm")

SHEET_NAME = "Template"
HEADER_ROW = 3
DATA_START_ROW = 4

# Clear these fields if they exist in the template
FIELDS_TO_CLEAR = {
    "item_sku",
    "external_product_id",
    "external_product_id_type",
    "item_name",
    "brand_name",
    "manufacturer",
    "part_number",
    "model",
    "model_name",
    "update_delete",
    "product_description",
    "generic_keywords",
    "bullet_point1",
    "bullet_point2",
    "bullet_point3",
    "bullet_point4",
    "bullet_point5",
    "recommended_browse_nodes",
    "standard_price",
    "list_price",
    "quantity",
    "main_image_url",
    "other_image_url1",
    "other_image_url2",
    "other_image_url3",
    "other_image_url4",
    "other_image_url5",
    "other_image_url6",
    "other_image_url7",
    "other_image_url8",
    "swatch_image_url",
    "parent_child",
    "parentage",
    "parent_sku",
    "relationship_type",
    "variation_theme",
    "color_name",
    "size_name",
    "apparel_size",
    "size_map",
    "color_map",
    "department_name",
    "feed_product_type",
    "style_name",
    "material_type",
    "shirt_size",
    "target_gender",
    "age_range_description",
    "is_autographed",
    "bottoms_size_system",
    "bottoms_size_class",
    "bottoms_size_value",
    "tops_size_system",
    "tops_size_class",
    "tops_size_value",
    "size_system",
    "size_class",
    "size_to",
    "size_from",
    "item_type_name",
    "special_features",
    "pattern_name",
    "care_instructions",
    "closure_type",
    "theme",
    "occasion_type",
    "embellishment_feature",
}

# Values that commonly appear in bad/default example rows
BAD_DEFAULT_VALUES = {
    "Accessory",
    "accessory",
    "One Size",
    "Yes",
    "No",
}


def build_header_map(ws, header_row: int) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is not None:
            key = str(value).strip()
            if key:
                header_map[key] = col
    return header_map


def clear_selected_fields(ws, header_map: dict[str, int], start_row: int) -> list[str]:
    cleared = []

    for field in sorted(FIELDS_TO_CLEAR):
        col_idx = header_map.get(field)
        if not col_idx:
            continue

        for row_idx in range(start_row, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).value = None

        cleared.append(field)

    return cleared


def clear_bad_default_values(ws, start_row: int) -> int:
    count = 0
    for row_idx in range(start_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value in BAD_DEFAULT_VALUES:
                cell.value = None
                count += 1
    return count


def main() -> None:
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(f"Source file not found: {SOURCE_FILE}")

    wb = load_workbook(SOURCE_FILE, keep_vba=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    header_map = build_header_map(ws, HEADER_ROW)

    cleared_fields = clear_selected_fields(ws, header_map, DATA_START_ROW)
    replaced_bad_defaults = clear_bad_default_values(ws, DATA_START_ROW)

    wb.save(OUTPUT_FILE)

    print(f"Saved cleaned template: {OUTPUT_FILE}")
    print(f"Sheet: {SHEET_NAME}")
    print(f"Header row: {HEADER_ROW}")
    print(f"Data start row: {DATA_START_ROW}")
    print(f"Fields cleared: {len(cleared_fields)}")
    print(f"Bad default values cleared: {replaced_bad_defaults}")

    print("\nCleared fields found in workbook:")
    for field in cleared_fields:
        print(f" - {field}")


if __name__ == "__main__":
    main()