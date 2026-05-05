from pathlib import Path
from openpyxl import load_workbook

FILE = Path("amazon_hoodie_base.xlsm")
SHEET_NAME = "Template"
HEADER_ROW = 3
DATA_START_ROW = 4

CHECK_FIELDS = [
    "item_sku",
    "item_name",
    "brand_name",
    "product_description",
    "generic_keywords",
    "recommended_browse_nodes",
    "main_image_url",
    "other_image_url1",
    "parent_child",
    "parent_sku",
    "relationship_type",
    "variation_theme",
    "color_name",
    "size_name",
    "apparel_size",
]


def build_header_map(ws, header_row: int) -> dict[str, int]:
    mapping = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is not None:
            mapping[str(v).strip()] = col
    return mapping


def main() -> None:
    wb = load_workbook(FILE, keep_vba=True)
    ws = wb[SHEET_NAME]
    header_map = build_header_map(ws, HEADER_ROW)

    print(f"Verifying: {FILE}")
    print(f"Sheet: {SHEET_NAME}\n")

    for field in CHECK_FIELDS:
        col = header_map.get(field)
        if not col:
            print(f"{field}: NOT FOUND")
            continue

        values = []
        for row in range(DATA_START_ROW, min(DATA_START_ROW + 5, ws.max_row + 1)):
            values.append(ws.cell(row=row, column=col).value)

        print(f"{field}: {values}")


if __name__ == "__main__":
    main()